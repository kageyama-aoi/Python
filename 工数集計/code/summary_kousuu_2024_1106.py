from math import nan
import pandas as pd
import datetime
import random
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import os
import glob


# カレントディレクトリの 'summary_' で始まるExcelファイルを検索
files_to_delete = glob.glob('output_summary_*.xlsx')

# ファイルを削除
for file in files_to_delete:
    os.remove(file)
    print(f'ファイルが削除されました: {file}')

def process_timesheet(file_path, input_tr_csv,target_project,temp_file01,sheet_name_detail,sheet_name_processed,sheet_name_monthly_summary):
    
    # データ読込
    df = pd.read_csv(file_path, encoding='cp932')

    # データ抽出（必要カラムと、対象案件のデータのみ）
    columns_to_keep_temp1 = [
    '社員コード', '社員名', 'プロジェクトコード', 'プロジェクト名',
    '開始日', '分', 'メモ', '工程１:管理コード', '工程１:名称', '開始時間','終了時間', '取引先名'
    ]
    df_filtered = df[df['取引先名'] == target_project][columns_to_keep_temp1]
    
    del df

    # データ加工
    print(f'★要素数:{len(df_filtered)}')
    df_filtered['開始日'] = pd.to_datetime(df_filtered['開始日'])
    df_filtered['作業時間'] = (df_filtered['分'] / 60).round(2)
    df_filtered['PJコード'] = df_filtered['プロジェクトコード'] + ':' + df_filtered['プロジェクト名']
    
    df_filtered['メモ_区分'] = None 
    df_filtered['メモ_詳細'] = None

    columns_to_keep_temp1.insert(columns_to_keep_temp1.index('プロジェクト名'),'PJコード')
    columns_to_keep_temp1.insert(columns_to_keep_temp1.index('分'),'作業時間')
    columns_to_keep_temp1.insert(columns_to_keep_temp1.index('工程１:管理コード'),'メモ_区分')
    columns_to_keep_temp1.insert(columns_to_keep_temp1.index('工程１:管理コード'),'メモ_詳細')
    columns_to_keep_temp1.remove('プロジェクト名')
    columns_to_keep_temp1.remove('プロジェクトコード')
    columns_to_keep_temp1.remove('分')
    columns_to_keep_temp1.remove('取引先名')
    
    # columns_to_keep_temp1 = [
    #     'PJコード' if col == 'プロジェクト名' else
    #     '作業時間' if col == '分' else
    #     'メモ_区分' if col == '工程１:管理コード' else
    #     'メモ_詳細' if col == '工程１:管理コード' else col
    #     for col in columns_to_keep_temp1
    #     if col not in ['プロジェクト名', 'プロジェクトコード', '分', '取引先名']
    # ]
    
    
    with open('output.txt', 'a', encoding='utf-8') as f:
        print('★-----★')
        print(f'value:{df_filtered["作業時間"]} type:{df_filtered["作業時間"]}', file=f)
        print(f'value:{df_filtered[columns_to_keep_temp1]} type:{df_filtered[columns_to_keep_temp1]}', file=f)        

    

    
    df_filtered_sorted = df_filtered[columns_to_keep_temp1].sort_values(
        by = ['社員名', 'PJコード', 'メモ','開始日'], 
        ascending  = [True, True, True, True]
    ).reset_index(drop=True)

    # 月別の集計
    monthly_summary_by_name = df_filtered_sorted.groupby(
    ['社員名', 'PJコード', df_filtered_sorted['開始日'].dt.to_period('M')]
    ).agg({
    '作業時間':'sum'
    }).reset_index()

    # 月別の集計（sort)
    columns_to_keep_month_summary = ['社員名', 'PJコード', '作業時間', '開始日']
    monthly_summary_sorted = monthly_summary_by_name[columns_to_keep_month_summary].sort_values(
    by=['社員名', '作業時間', 'PJコード'],
    ascending=[True, False, True]
    ).reset_index(drop=True)
    del monthly_summary_by_name

    
    tr_list = 'tr_list.xlsx'

    # Excelブックに保存
    sheet_name_tr_info = 'TR'

    with pd.ExcelWriter(temp_file01, engine='xlsxwriter') as writer:
        df_filtered.to_excel(writer, sheet_name=sheet_name_detail, index=False)
        df_filtered_sorted.to_excel(writer, sheet_name=sheet_name_processed, index=False)
        monthly_summary_sorted.to_excel(writer, sheet_name=sheet_name_monthly_summary, index=False)
        pd.read_csv(input_tr_csv).to_excel(writer, sheet_name=sheet_name_tr_info, index=False)

    with pd.ExcelWriter(tr_list, engine='xlsxwriter') as writer:
        pd.read_csv(input_tr_csv).to_excel(writer, sheet_name=sheet_name_tr_info, index=False)

    return temp_file01, tr_list,df_filtered_sorted



# datafrmae様に修正が必要(2024/1022)
def edit_and_save_df_remake(df_filtered_sorted, output_path,sheet_name_processed):
    """
    Excelファイルを読み込み、シートを編集して新しいファイルに保存します。

    :param df_filtered_sorted: 編集するデータフレーム
    :param output_path: 保存する新しいExcelファイルのパス
    """
    
    sum_name_label = '集計_社員名'
    sum_pj_label = '集計_PJコード'
    sum_time_label = '集計_工数(h)'
    sum_block1_label = '集計_ブロック'
    sum_time_block1_label = '集計_工数(h)_番外'
            
    # 新しい列を追加し、ヘッダーを設定
    # df_filtered_sorted.insert(0, sum_name_label, '')
    # df_filtered_sorted.insert(1, sum_pj_label, '')
    # df_filtered_sorted.insert(2, sum_time_label, '')
    # df_filtered_sorted.insert(3, sum_block1_label, '')
    # df_filtered_sorted.insert(4, sum_time_block1_label, '')
    
    
    labels = [sum_name_label, sum_pj_label, sum_time_label, sum_block1_label, sum_time_block1_label]

    for i, label in enumerate(labels):
        df_filtered_sorted.insert(i, label, '')

    # 「社員名」「PJコード」「時間」の列番号を特定
    col_num = {'社員名': None, 'PJコード': None, '作業時間': None, sum_name_label: None, sum_pj_label: None, \
        sum_time_label: None, 'メモ_区分':None,'メモ_詳細':None, 'メモ':None}
    for col in df_filtered_sorted.columns:
        if col in col_num:
            col_num[col] = df_filtered_sorted.columns.get_loc(col)+1

    # 必要なヘッダーが見つからない場合のエラーチェック
    if not all(col_num.values()):
        raise ValueError("必要なヘッダーが見つかりませんでした。")

    
    one_line_ago_syainmmei_value = ''
    one_line_ago_pjcode_value = ''
    
    # 集計列情報を取得（英字形式）
    column_sagyou = get_column_letter(col_num['作業時間'])
    column_syainmei = get_column_letter(col_num['社員名'])
    column_pjcode = get_column_letter(col_num['PJコード'])
    column_memo = get_column_letter(col_num['メモ'])
    column_memo_kubun = get_column_letter(col_num['メモ_区分'])
    keyword_syain = get_column_letter(col_num[sum_name_label])
    keyword_pj = get_column_letter(col_num[sum_pj_label])
    
    #行ごと処理
    for index, row in df_filtered_sorted.iterrows():

        syainmmei_value = row['社員名']
        pjcode_value = row['PJコード']
        excel_row = index+2

        # 「集計」列に、値や、式を埋め込み
        df_filtered_sorted.at[index, 'メモ_区分'] = f'=IFERROR(LEFT({column_memo}{excel_row}, \
            FIND(",", {column_memo}{excel_row}) - 1), "-")'
        df_filtered_sorted.at[index, 'メモ_詳細'] = f'=IFERROR(MID({column_memo}{excel_row}, \
            FIND(",", {column_memo}{excel_row}) + 1, LEN({column_memo}{excel_row}) - FIND(",", {column_memo}{excel_row})), {column_memo}{excel_row})'
        df_filtered_sorted.at[index, '工程１:名称'] = row['メモ_区分']
        df_filtered_sorted.at[index, '集計_ブロック'] = f'={column_memo_kubun}{excel_row}'
        df_filtered_sorted.at[index, sum_name_label] = syainmmei_value
        df_filtered_sorted.at[index, sum_pj_label] = pjcode_value
        df_filtered_sorted.at[index, sum_time_block1_label] = f"=SUMIFS({column_sagyou}:{column_sagyou}, {column_syainmei}:{column_syainmei}, \
            {keyword_syain}{excel_row}, {column_pjcode}:{column_pjcode}, {keyword_pj}{excel_row}, {column_memo_kubun}:{column_memo_kubun}, {column_memo_kubun}{excel_row})"  

        
        # 1行前と比較:「社員名」と「PJコード」が同一の場合は、以降の処理をスキップ
        if syainmmei_value == one_line_ago_syainmmei_value and pjcode_value == one_line_ago_pjcode_value:
            continue
        
        df_filtered_sorted.at[index, sum_time_label] = f"=SUMIFS({column_sagyou}:{column_sagyou}, {column_syainmei}:{column_syainmei}, \
            {keyword_syain}{excel_row}, {column_pjcode}:{column_pjcode}, {keyword_pj}{excel_row})"  
        
        one_line_ago_syainmmei_value = syainmmei_value
        one_line_ago_pjcode_value = pjcode_value
        

    # ファイルを保存
    with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
        df_filtered_sorted.to_excel(writer, sheet_name=sheet_name_processed, index=False)
        
        # Excelファイルを読み込む
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    # D列を値のみにする
    for row in range(1, ws.max_row + 1):
        ws[f'D{row}'] = ws[f'D{row}'].value
        print(ws[f'D{row}'].value)

    # Excelファイルを保存する
    wb.save(output_path)
        
    print(f'中間ブックが保存ー: {output_path}')


def save_sheet_as_new_file(input_path, sheet_name, output_path):
    """
    指定されたシートを新しいExcelファイルとして保存します。

    :param input_path: 入力ファイルパス
    :param sheet_name: シート名
    :param output_path: 出力ファイルパス
    """
    # ワークブックの読み込み
    wb = openpyxl.load_workbook(input_path)
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    ws = wb[sheet_name]

    # シートの内容を新しいワークブックにコピー
    for row in ws.iter_rows(values_only=True):
        new_ws.append(row)

    # シート名の設定とデフォルトシートの削除
    new_ws.title = sheet_name
    if 'Sheet' in new_wb.sheetnames:
        new_wb.remove(new_wb['Sheet'])

    # ファイルの保存
    new_wb.save(output_path)


def customize_excel_sheet(output_path):
    wb = openpyxl.load_workbook(output_path)
    sheet = wb.active
    
    # ヘッダー設定（固定、背景色）
    sheet.freeze_panes = 'A2'  # 1行目を固定
    right_green='009999'
    right_blue='B8CCE4'
    
    fill_color = PatternFill(start_color=right_green, end_color=right_green, fill_type='solid')
    fill_color_syukei = PatternFill(start_color=right_blue, end_color=right_blue, fill_type='solid')
    
    for i in range(6, 15):
        cell = sheet.cell(row=1, column=i)
        cell.fill = fill_color
    for i in range(1, 6):
        cell = sheet.cell(row=1, column=i)
        cell.fill = fill_color_syukei
        # cell.fill = fill_color_syukei

    # 15列目と16列目の背景色を設定
    for i in range(1, 200):
        cell15 = sheet.cell(row=i, column=15)
        cell16 = sheet.cell(row=i, column=16)
        cell15.fill = fill_color_syukei
        cell16.fill = fill_color_syukei

    # 列幅の調整
    for col in sheet.columns:
        max_length = max((len(str(cell.value)) for cell in col if cell.value is not None), default=0)
        column = get_column_letter(col[0].column)
        if column not in ['C','E','O','J']:  # Excel式の長さにあわせると、横幅が広いため除外して幅を直接指定
            adjusted_width = (max_length + 2) * 1.1
            sheet.column_dimensions[column].width = adjusted_width

    # 列グループ化
    sheet.column_dimensions.group('F', 'I', hidden=True)
    sheet.column_dimensions.group('J', 'N', hidden=True)

    wb.save(output_path)

# ユーザー定義
input_tr_csv ='./input/bugs.csv'
input_timesheet_csv = './input/timesheet.csv'
target_project = '社内ダミー' # 島村楽器　社内ダミー
# sheet_name_detail = 'Detailed Data_' #+ target_project
# sheet_name_processed = 'Detailed Processed Data_' #+ target_project
# sheet_name_monthly_summary = 'Monthly Summary_' #+ target_project

sheet_name_detail = 'Detailed Data' #+ target_project
sheet_name_processed = 'Processed Data' #+ target_project
sheet_name_monthly_summary = 'Monthly Summary' #+ target_project

now = datetime.datetime.now()
random_number = random.randint(10000, 99999)
temp_file01 = f'./output/output_summary_{target_project}_{now:%Y%m%d_%H%M%S}_{random_number}.xlsx'
temp_file02 = './output/temp_file1.xlsx' 
final_output = './output/output_'+target_project+'_aggregate_results.xlsx'  

output_file, tr_file ,df_filtered_sorted= process_timesheet(input_timesheet_csv, input_tr_csv,target_project,temp_file01,sheet_name_detail,sheet_name_processed,sheet_name_monthly_summary)

edit_and_save_df_remake(df_filtered_sorted,temp_file02,sheet_name_processed)

save_sheet_as_new_file(temp_file02, sheet_name_processed, final_output)

customize_excel_sheet(final_output)


print(f'Excelブックが保存されましたー: {final_output}')

