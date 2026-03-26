import pytest
import openpyxl
import pandas as pd
from pathlib import Path
from src.excel_writer import (
    save_initial_report,
    extract_sheet_to_new_file,
    apply_custom_styles,
    add_formulas_and_save,
)
from src.constants import StyleConfig as SC


@pytest.fixture
def simple_df():
    return pd.DataFrame({"A": [1, 2, 3], "B": ["x", "y", "z"]})


@pytest.fixture
def saved_excel(tmp_path, simple_df):
    """単一シートのExcelファイルを返す"""
    path = tmp_path / "test.xlsx"
    save_initial_report(path, [("Sheet1", simple_df)])
    return path


class TestSaveInitialReport:

    def test_creates_file(self, tmp_path, simple_df):
        path = tmp_path / "out.xlsx"
        save_initial_report(path, [("Data", simple_df)])
        assert path.exists()

    def test_multiple_sheets(self, tmp_path, simple_df):
        path = tmp_path / "out.xlsx"
        save_initial_report(path, [("S1", simple_df), ("S2", simple_df)])
        wb = openpyxl.load_workbook(path)
        assert "S1" in wb.sheetnames
        assert "S2" in wb.sheetnames


class TestExtractSheetToNewFile:

    def test_extracts_sheet(self, tmp_path, simple_df):
        src = tmp_path / "src.xlsx"
        dst = tmp_path / "dst.xlsx"
        save_initial_report(src, [("Sheet1", simple_df), ("Sheet2", simple_df)])
        extract_sheet_to_new_file(src, "Sheet1", dst)
        assert dst.exists()
        wb = openpyxl.load_workbook(dst)
        assert "Sheet1" in wb.sheetnames

    def test_raises_on_missing_sheet(self, tmp_path, simple_df):
        """存在しないシート名を指定したら ValueError を raise する（RF-10）"""
        src = tmp_path / "src.xlsx"
        dst = tmp_path / "dst.xlsx"
        save_initial_report(src, [("Sheet1", simple_df)])
        with pytest.raises(ValueError, match="存在しません"):
            extract_sheet_to_new_file(src, "NoSuchSheet", dst)

    def test_no_leftover_default_sheet(self, tmp_path, simple_df):
        """デフォルトの 'Sheet' が残らない"""
        src = tmp_path / "src.xlsx"
        dst = tmp_path / "dst.xlsx"
        save_initial_report(src, [("MySheet", simple_df)])
        extract_sheet_to_new_file(src, "MySheet", dst)
        wb = openpyxl.load_workbook(dst)
        assert "Sheet" not in wb.sheetnames


class TestApplyCustomStyles:

    def test_freeze_pane_set(self, tmp_path, simple_df):
        path = tmp_path / "styled.xlsx"
        save_initial_report(path, [("Sheet1", simple_df)])
        apply_custom_styles(path)
        wb = openpyxl.load_workbook(path)
        assert wb.active.freeze_panes == "A2"

    def test_large_data_fills_all_rows(self, tmp_path):
        """200行超でもO・P列の背景色が全行に適用される（RF-13）"""
        df = pd.DataFrame({"col": range(250)})
        path = tmp_path / "large.xlsx"
        # 列数が少ないのでStyleConfigの列番号に合わせてダミー列を足す
        for i in range(17):
            df[f"pad_{i}"] = ""
        save_initial_report(path, [("Sheet1", df)])
        apply_custom_styles(path)
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        # 201行目（データ200行目）のO列が塗られているか確認
        cell = ws.cell(row=202, column=SC.FILL_BLUE_COL_1)
        assert cell.fill.fgColor.rgb != "00000000"


class TestAddFormulasAndSave:

    def test_output_file_created(self, tmp_path, sample_timesheet_df):
        """add_formulas_and_save がファイルを生成する"""
        from src.processor import process_details
        processed = process_details(sample_timesheet_df)
        out = tmp_path / "formulas.xlsx"
        add_formulas_and_save(processed, out, "Processed Data")
        assert out.exists()
