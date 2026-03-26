import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd
from . import config as cfg
from .constants import InputCols as IC, InternalCols as RC, StyleConfig as SC
import sys

def save_initial_report(temp_file_path, sheet_data_list):
    """
    初期データをExcelに保存する
    sheet_data_list: [(sheet_name, dataframe), ...]
    """
    with pd.ExcelWriter(temp_file_path, engine='xlsxwriter') as writer:
        for sheet_name, df in sheet_data_list:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

def _insert_formula_columns(df):
    """集計用の補助列をDataFrameの先頭に挿入する"""
    labels = [RC.SUM_EMPLOYEE, RC.SUM_PJ, RC.SUM_HOURS, RC.SUM_BLOCK, RC.SUM_HOURS_EXTRA]
    for i, label in enumerate(labels):
        df.insert(i, label, '')
    return df


def _get_col_letters(df):
    """数式生成に必要な列レターの辞書を返す"""
    required_cols = [
        IC.EMPLOYEE_NAME, RC.PJ_CODE, RC.WORK_HOURS,
        RC.SUM_EMPLOYEE, RC.SUM_PJ, RC.SUM_HOURS,
        RC.MEMO_TYPE, RC.MEMO_DETAIL, IC.MEMO
    ]
    col_num = {col: None for col in required_cols}
    for col in df.columns:
        if col in col_num:
            col_num[col] = df.columns.get_loc(col) + 1

    if any(v is None for v in col_num.values()):
        raise ValueError("必要なヘッダーが見つかりませんでした。")

    return {col: get_column_letter(num) for col, num in col_num.items()}


def _embed_formulas(df, letters):
    """各行にExcel数式を埋め込む"""
    prev_name = ''
    prev_pj = ''

    for index, row in df.iterrows():
        name = row[IC.EMPLOYEE_NAME]
        pj = row[RC.PJ_CODE]
        r = index + 2

        df.at[index, RC.MEMO_TYPE] = (
            f'=IFERROR(LEFT({letters[IC.MEMO]}{r}, FIND(",", {letters[IC.MEMO]}{r}) - 1), "-")'
        )
        df.at[index, RC.MEMO_DETAIL] = (
            f'=IFERROR(MID({letters[IC.MEMO]}{r}, FIND(",", {letters[IC.MEMO]}{r}) + 1, '
            f'LEN({letters[IC.MEMO]}{r}) - FIND(",", {letters[IC.MEMO]}{r})), {letters[IC.MEMO]}{r})'
        )
        df.at[index, IC.PROCESS_1_NAME] = row[RC.MEMO_TYPE]
        df.at[index, RC.SUM_BLOCK] = f'={letters[RC.MEMO_TYPE]}{r}'
        df.at[index, RC.SUM_EMPLOYEE] = name
        df.at[index, RC.SUM_PJ] = pj
        df.at[index, RC.SUM_HOURS_EXTRA] = (
            f"=SUMIFS({letters[RC.WORK_HOURS]}:{letters[RC.WORK_HOURS]}, "
            f"{letters[IC.EMPLOYEE_NAME]}:{letters[IC.EMPLOYEE_NAME]}, {letters[RC.SUM_EMPLOYEE]}{r}, "
            f"{letters[RC.PJ_CODE]}:{letters[RC.PJ_CODE]}, {letters[RC.SUM_PJ]}{r}, "
            f"{letters[RC.MEMO_TYPE]}:{letters[RC.MEMO_TYPE]}, {letters[RC.MEMO_TYPE]}{r})"
        )

        if name == prev_name and pj == prev_pj:
            continue

        df.at[index, RC.SUM_HOURS] = (
            f"=SUMIFS({letters[RC.WORK_HOURS]}:{letters[RC.WORK_HOURS]}, "
            f"{letters[IC.EMPLOYEE_NAME]}:{letters[IC.EMPLOYEE_NAME]}, {letters[RC.SUM_EMPLOYEE]}{r}, "
            f"{letters[RC.PJ_CODE]}:{letters[RC.PJ_CODE]}, {letters[RC.SUM_PJ]}{r})"
        )
        prev_name = name
        prev_pj = pj


def _save_with_fixed_column_d(df, output_path, sheet_name):
    """DataFrameをExcel保存後、D列を数式から値に確定させる"""
    try:
        with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        for row in range(1, ws.max_row + 1):
            ws[f'D{row}'] = ws[f'D{row}'].value
        wb.save(output_path)
    except PermissionError:
        print(f"エラー: {output_path} に書き込めませんでした。ファイルが開かれている可能性があります。閉じてから再実行してください。")
        sys.exit(1)


def add_formulas_and_save(df, output_path, sheet_name):
    """データフレームにExcel数式を追加して保存し、D列を値として確定させる"""
    df = _insert_formula_columns(df)
    letters = _get_col_letters(df)
    _embed_formulas(df, letters)
    _save_with_fixed_column_d(df, output_path, sheet_name)

def extract_sheet_to_new_file(input_path, sheet_name, output_path):
    """指定されたシートを新しいExcelファイルとして抽出する"""
    wb = openpyxl.load_workbook(input_path)
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    
    if sheet_name not in wb.sheetnames:
        return
        
    ws = wb[sheet_name]

    for row in ws.iter_rows(values_only=True):
        new_ws.append(row)

    new_ws.title = sheet_name
    if 'Sheet' in new_wb.sheetnames:
        new_wb.remove(new_wb['Sheet'])

    try:
        new_wb.save(output_path)
    except PermissionError:
        print(f"エラー: {output_path} に書き込めませんでした。ファイルが開かれている可能性があります。閉じてから再実行してください。")
        sys.exit(1)

def apply_custom_styles(file_path):
    """Excelシートにスタイル（固定、背景色、列幅、グループ化）を適用する"""
    try:
        wb = openpyxl.load_workbook(file_path)
    except PermissionError:
         print(f"エラー: {file_path} を開けませんでした。ファイルが開かれている可能性があります。閉じてから再実行してください.")
         sys.exit(1)
         
    sheet = wb.active
    
    sheet.freeze_panes = 'A2'

    fill_green = PatternFill(start_color=cfg.COLOR_RIGHT_GREEN, end_color=cfg.COLOR_RIGHT_GREEN, fill_type='solid')
    fill_blue = PatternFill(start_color=cfg.COLOR_RIGHT_BLUE, end_color=cfg.COLOR_RIGHT_BLUE, fill_type='solid')

    # ヘッダー背景色
    for i in SC.HEADER_BLUE_COLS:
        sheet.cell(row=1, column=i).fill = fill_blue
    for i in SC.HEADER_GREEN_COLS:
        sheet.cell(row=1, column=i).fill = fill_green

    # 特定列の背景色
    for i in range(1, 200):
        sheet.cell(row=i, column=SC.FILL_BLUE_COL_1).fill = fill_blue
        sheet.cell(row=i, column=SC.FILL_BLUE_COL_2).fill = fill_blue

    # 列幅の自動調整（一部除外）
    for col in sheet.columns:
        column_letter = get_column_letter(col[0].column)
        if column_letter not in SC.WIDTH_SKIP_COLS:
            max_length = max((len(str(cell.value)) for cell in col if cell.value is not None), default=0)
            adjusted_width = (max_length + 2) * 1.1
            sheet.column_dimensions[column_letter].width = adjusted_width

    # 列グループ化
    sheet.column_dimensions.group(SC.GROUP_1_START, SC.GROUP_1_END, hidden=True)
    sheet.column_dimensions.group(SC.GROUP_2_START, SC.GROUP_2_END, hidden=True)

    try:
        wb.save(file_path)
    except PermissionError:
        print(f"エラー: {file_path} に書き込めませんでした。ファイルが開かれている可能性があります。閉じてから再実行してください。")
        sys.exit(1)
