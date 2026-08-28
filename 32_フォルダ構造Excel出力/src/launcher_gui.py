# 32_フォルダ構造Excel出力 ランチャー GUI
# generate_drive_structure.py を、実行ログをリアルタイムに見ながら実行できるようにする
# 単一ツール向けランチャー。Python 標準ライブラリ（Tkinter）のみで動作する。
# 31_CSVまとめ閲覧Excel化/src/launcher_gui.py の設計を踏襲している。
# テーマ・フォント・ボタン3段階スタイル・タイトルバー処理は theme.py に集約
# （sv-ttk / pywinstyles は theme.py 側で任意 import。未導入でも標準ttkで動作する）。
# 古い出力削除の確認ダイアログは generate_drive_structure.py 側のものがそのまま表示される。

import json
import os
import re
import sys
import time
import queue
import zipfile
import threading
import subprocess
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from tkinter.scrolledtext import ScrolledText
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

import theme
from theme import (
    ACCENT_FG,
    BTN_PRIMARY,
    BTN_SECONDARY,
    BTN_TERTIARY,
    ERROR_SOFT_FG,
    HEADER_FONT,
    LOG_BG,
    LOG_FG,
    LOG_FONT,
    MUTED_FG,
    SUCCESS_FG,
    UI_FONT_BOLD,
    WARN_FG,
)

# src/ の1つ上（プロジェクトルート）を基準にする。cwdに依存しないため、
# run.batから起動しても直接 `python src/launcher_gui.py` を実行しても同じ場所を指す。
BASE_DIR = Path(__file__).resolve().parent.parent
CONFIG_PATH = BASE_DIR / "config" / "config.json"
OUTPUT_DIR = BASE_DIR / "data" / "output"
LOGS_DIR = BASE_DIR / "data" / "logs"
LOG_CLEANUP_DAYS = 30

ABOUT_TEXT = (
    "準備するもの： 左上でスキャン対象フォルダを指定（「参照...」）。出力先・除外条件は「設定を編集...」で調整\n"
    "出力されるもの： data/output/ に drive_structure_<日時>.xlsx（階層構造付きExcel）"
)

_LOG_TAGS = {
    "header": {"foreground": theme.ACCENT_FG},
    "debug": {"foreground": theme.MUTED_FG},
    "pass": {"foreground": theme.SUCCESS_FG},
    "fail": {"foreground": theme.ERROR_FG, "font": LOG_FONT + ("bold",)},
    "error": {"foreground": theme.ERROR_SOFT_FG},
    "warn": {"foreground": theme.WARN_FG},
}

_SUMMARY_COLORS = {
    "idle": theme.MUTED_FG,
    "running": theme.ACCENT_FG,
    "done": theme.SUCCESS_FG,
    "empty": theme.WARN_FG,
}

_SENTINEL_RUN_DONE = object()

# 巨大なフォルダ（Google Drive同期フォルダ等）を除外設定込みでスキャンすると、
# 除外ログだけで数万行に達することがある。1回のflushで無制限に処理すると
# Text部品への大量insert/see呼び出しでメインスレッドが長時間ブロックされ、
# Windowsに「応答なし」と判定される（実測: 3万行を1回で処理すると60秒以上）。
# 1回のflushで処理する行数の上限と、表示欄自体の保持行数の上限で対策する。
MAX_LOG_LINES_PER_FLUSH = 500
MAX_LOG_DISPLAY_LINES = 5000


# ------------------------------------------------------------------
# 汎用ヘルパー
# ------------------------------------------------------------------

def _format_filesize(size_bytes):
    if size_bytes < 1024:
        return f"{size_bytes} B"
    if size_bytes < 1024 * 1024:
        return f"{size_bytes / 1024:.1f} KB"
    return f"{size_bytes / 1024 / 1024:.1f} MB"


def _open_in_explorer(path):
    try:
        os.startfile(str(path))
    except OSError as exc:
        messagebox.showerror("エラー", f"開けませんでした:\n{path}\n\n{exc}")


def cleanup_old_logs():
    """data/logs/ 内の script_YYYYMMDD.log のうち古いものを zip 化して削除する。"""
    if not LOGS_DIR.is_dir():
        return
    archive_dir = LOGS_DIR / "archive"
    cutoff = datetime.now() - timedelta(days=LOG_CLEANUP_DAYS)
    for log_file in LOGS_DIR.glob("script_*.log"):
        try:
            mtime = datetime.fromtimestamp(log_file.stat().st_mtime)
            if mtime >= cutoff:
                continue
            archive_dir.mkdir(exist_ok=True)
            zip_path = archive_dir / f"{log_file.name}.zip"
            with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
                zf.write(log_file, arcname=log_file.name)
            log_file.unlink()
        except OSError:
            continue


def load_config():
    """config/config.json を読み込む。存在しない・壊れている場合は None。"""
    try:
        return json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None


def _snapshot_output_dir():
    """data/output/ の {ファイル名: mtime} スナップショットを返す。"""
    snapshot = {}
    if OUTPUT_DIR.is_dir():
        for p in OUTPUT_DIR.iterdir():
            if p.is_file() and p.suffix.lower() == ".xlsx":
                try:
                    snapshot[p.name] = p.stat().st_mtime
                except OSError:
                    pass
    return snapshot


def _detect_new_output_files(snapshot):
    """スナップショット以降に作成・更新されたxlsxを新しい順で返す。"""
    if not OUTPUT_DIR.is_dir():
        return []
    results = []
    for p in OUTPUT_DIR.iterdir():
        if not (p.is_file() and p.suffix.lower() == ".xlsx"):
            continue
        try:
            mtime = p.stat().st_mtime
        except OSError:
            continue
        if p.name not in snapshot or mtime > snapshot[p.name] + 0.5:
            results.append((mtime, p))
    return [p for _mtime, p in sorted(results, reverse=True)]


def _count_output_rows(path):
    """生成されたExcelの行数（ヘッダー除く）を返す。読めなければ None。"""
    try:
        wb = load_workbook(path, read_only=True)
        try:
            ws = wb.active
            return max(0, (ws.max_row or 1) - 1)
        finally:
            wb.close()
    except Exception:
        return None


# ------------------------------------------------------------------
# 設定編集サブウィンドウ
# ------------------------------------------------------------------

_EXT_PRESETS = [
    ".log", ".tmp", ".bak", ".png", ".jpg", ".jpeg", ".gif",
    ".zip", ".exe", ".pyc", ".DS_Store", "desktop.ini",
]
_FOLDER_PRESETS = [
    "OLD", "old", "backup", "バックアップ", "アーカイブ", "不要",
    "output", "logs", "__pycache__", ".git", "node_modules", ".venv", "venv",
]


def split_tokens(text):
    """カンマ・空白・改行区切りの文字列をトークン列に分解する（空要素は除外）。"""
    return [t for t in re.split(r"[,\s]+", text.strip()) if t]


def merge_unique(existing, additions):
    """existing の順序を保ったまま、未収録の additions を末尾に足したリストを返す。"""
    result = list(existing)
    seen = set(result)
    for item in additions:
        if item not in seen:
            result.append(item)
            seen.add(item)
    return result


class _TagListEditor(ttk.LabelFrame):
    """文字列リスト（除外拡張子・除外フォルダ名など）を追加/削除で編集するウィジェット。"""

    def __init__(self, parent, title, items, presets):
        super().__init__(parent, text=title, padding=8)
        self.columnconfigure(0, weight=1)

        list_row = ttk.Frame(self)
        list_row.grid(row=0, column=0, sticky="ew")
        list_row.columnconfigure(0, weight=1)
        self._listbox = tk.Listbox(list_row, height=5, selectmode="extended",
                                   activestyle="none", exportselection=False)
        self._listbox.grid(row=0, column=0, sticky="ew")
        sb = ttk.Scrollbar(list_row, orient="vertical", command=self._listbox.yview)
        sb.grid(row=0, column=1, sticky="ns")
        self._listbox.config(yscrollcommand=sb.set)
        self._listbox.bind("<Double-Button-1>", lambda e: self._remove_selected())
        self._listbox.bind("<Delete>", lambda e: self._remove_selected())
        for item in items:
            self._listbox.insert("end", item)

        add_row = ttk.Frame(self)
        add_row.grid(row=1, column=0, sticky="ew", pady=(6, 2))
        add_row.columnconfigure(0, weight=1)
        self._entry = ttk.Entry(add_row)
        self._entry.grid(row=0, column=0, sticky="ew")
        self._entry.bind("<Return>", lambda e: self._add_from_entry())
        ttk.Button(add_row, text="追加", width=6, style=BTN_SECONDARY,
                   command=self._add_from_entry).grid(row=0, column=1, padx=(6, 0))
        ttk.Button(add_row, text="選択を削除", style=BTN_TERTIARY,
                   command=self._remove_selected).grid(row=0, column=2, padx=(6, 0))

        preset_row = ttk.Frame(self)
        preset_row.grid(row=2, column=0, sticky="w", pady=(2, 0))
        ttk.Label(preset_row, text="プリセットから追加:", foreground=MUTED_FG).pack(
            side="left", padx=(0, 6))
        self._preset_combo = ttk.Combobox(preset_row, state="readonly",
                                          values=presets, width=18)
        self._preset_combo.pack(side="left")
        self._preset_combo.bind("<<ComboboxSelected>>", self._add_from_preset)
        ttk.Button(preset_row, text="すべて追加", style=BTN_TERTIARY,
                   command=lambda: self._add_items(presets)).pack(side="left", padx=(6, 0))

    def _add_from_preset(self, _event=None):
        value = self._preset_combo.get()
        if value:
            self._add_items([value])
        self._preset_combo.set("")

    def _add_from_entry(self):
        self._add_items(split_tokens(self._entry.get()))
        self._entry.delete(0, "end")

    def _add_items(self, additions):
        merged = merge_unique(self.get_items(), additions)
        self._listbox.delete(0, "end")
        for item in merged:
            self._listbox.insert("end", item)

    def _remove_selected(self):
        for idx in reversed(self._listbox.curselection()):
            self._listbox.delete(idx)

    def get_items(self):
        return list(self._listbox.get(0, "end"))


class ConfigEditorWindow(tk.Toplevel):
    """config/config.json をGUIで編集する。"""

    def __init__(self, master, on_saved=None):
        super().__init__(master)
        self.on_saved = on_saved
        self.title("設定編集 (config/config.json)")
        self.geometry("700x640")
        self.minsize(600, 560)
        self.resizable(True, True)
        self.transient(master)
        self.grab_set()
        theme.style_titlebar(self)

        config = load_config()
        if config is None:
            messagebox.showerror(
                "エラー",
                f"設定ファイルを読み込めません:\n{CONFIG_PATH}\n\n"
                f"config/config.example.json をコピーして config/config.json を作成してください。",
                parent=master)
            self.destroy()
            return

        frame = ttk.Frame(self, padding=16)
        frame.pack(fill="both", expand=True)
        frame.columnconfigure(1, weight=1)

        ttk.Label(frame, foreground=MUTED_FG,
                  text="スキャン対象フォルダはメイン画面の「スキャン対象フォルダ」で指定します。").grid(
            row=0, column=0, columnspan=2, sticky="w", pady=(0, 12))

        ttk.Label(frame, text="出力先フォルダ (output_base_dir)").grid(row=1, column=0, sticky="w", pady=4)
        self.output_base_dir_var = tk.StringVar(value=config.get("output_base_dir", "data/output"))
        ttk.Entry(frame, textvariable=self.output_base_dir_var).grid(
            row=1, column=1, sticky="ew", padx=(12, 0), pady=4)

        ttk.Label(frame, text="出力ファイル名 (output_filename)").grid(row=2, column=0, sticky="w", pady=4)
        self.output_filename_var = tk.StringVar(value=config.get("output_filename", "drive_structure.xlsx"))
        ttk.Entry(frame, textvariable=self.output_filename_var).grid(
            row=2, column=1, sticky="ew", padx=(12, 0), pady=4)

        self.ext_editor = _TagListEditor(
            frame, "除外拡張子 (excluded_extensions)",
            config.get("excluded_extensions", []), _EXT_PRESETS)
        self.ext_editor.grid(row=3, column=0, columnspan=2, sticky="ew", pady=(12, 6))

        self.folder_editor = _TagListEditor(
            frame, "除外フォルダ名 (excluded_folder_names・大文字小文字は区別しない)",
            config.get("excluded_folder_names", []), _FOLDER_PRESETS)
        self.folder_editor.grid(row=4, column=0, columnspan=2, sticky="ew", pady=(6, 6))

        ttk.Label(frame, foreground=MUTED_FG,
                  text="※ 入力欄はカンマ・空白区切りでまとめて追加できます。項目のダブルクリックで削除。").grid(
            row=5, column=0, columnspan=2, sticky="w", pady=(2, 0))

        btn_frame = ttk.Frame(frame)
        btn_frame.grid(row=6, column=0, columnspan=2, sticky="e", pady=(16, 0))
        ttk.Button(btn_frame, text="保存", style=BTN_SECONDARY,
                   command=self._save).pack(side="left", padx=(0, 8))
        ttk.Button(btn_frame, text="キャンセル", style=BTN_TERTIARY,
                   command=self.destroy).pack(side="left")

    def _save(self):
        config = load_config() or {}
        config.setdefault("root_dir", ".")  # スキャン対象はメイン画面側で管理する
        config["output_base_dir"] = self.output_base_dir_var.get().strip() or "data/output"
        config["output_filename"] = self.output_filename_var.get().strip() or "drive_structure.xlsx"
        config["excluded_extensions"] = self.ext_editor.get_items()
        config["excluded_folder_names"] = self.folder_editor.get_items()
        try:
            CONFIG_PATH.parent.mkdir(parents=True, exist_ok=True)
            CONFIG_PATH.write_text(
                json.dumps(config, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        except OSError as exc:
            messagebox.showerror("エラー", f"保存に失敗しました:\n{exc}", parent=self)
            return
        if self.on_saved:
            self.on_saved()
        self.destroy()


# ------------------------------------------------------------------
# メインウィンドウ
# ------------------------------------------------------------------

class LauncherApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("フォルダ構造Excel出力 ランチャー")
        self.geometry("1000x660")
        self.minsize(780, 540)

        self.log_queue = queue.Queue()
        self.proc = None
        self.is_running = False
        self._last_output_file = None
        self._run_started_at = None
        self._scanned_count = 0
        self._elapsed_job = None

        self._build_ui()
        self.style = theme.apply_theme(self)

        self.refresh_config_summary()
        self.after(100, self._drain_log_queue)
        self.after(300, lambda: threading.Thread(target=cleanup_old_logs, daemon=True).start())
        self.protocol("WM_DELETE_WINDOW", self._on_close)
        theme.style_titlebar(self)

    # -------------------------------------------------- UI 構築

    def _build_ui(self):
        about = ttk.LabelFrame(self, text="このツールについて", padding=10)
        about.pack(fill="x", padx=8, pady=(8, 0))
        ttk.Label(about, text=ABOUT_TEXT, justify="left").pack(anchor="w")

        paned = ttk.PanedWindow(self, orient="horizontal")
        paned.pack(fill="both", expand=True, padx=8, pady=8)

        # ---- 左ペイン
        left = ttk.Frame(paned, padding=(0, 0, 8, 0))
        paned.add(left, weight=0)

        # スキャン対象フォルダ（メイン画面で直接指定・確認できる）
        ttk.Label(left, text="スキャン対象フォルダ", font=HEADER_FONT).pack(anchor="w")
        self.root_dir_var = tk.StringVar()
        rd_row = ttk.Frame(left)
        rd_row.pack(fill="x", pady=(4, 2))
        self.root_dir_entry = ttk.Entry(rd_row, textvariable=self.root_dir_var)
        self.root_dir_entry.pack(side="left", fill="x", expand=True)
        ttk.Button(rd_row, text="参照...", width=8, style=BTN_SECONDARY,
                   command=self._browse_root_dir).pack(side="left", padx=(4, 0))
        self.root_dir_hint = ttk.Label(left, text="", foreground=MUTED_FG,
                                       wraplength=340, justify="left")
        self.root_dir_hint.pack(anchor="w", pady=(0, 12))
        self.root_dir_var.trace_add("write", lambda *_: self._update_root_dir_hint())

        # 出力先・除外設定（詳細は「設定を編集...」）
        ttk.Label(left, text="出力先・除外設定", font=HEADER_FONT).pack(anchor="w")
        self.config_label = ttk.Label(left, text="", justify="left", wraplength=340)
        self.config_label.pack(anchor="w", fill="x", pady=(4, 8))

        cfg_btns = ttk.Frame(left)
        cfg_btns.pack(fill="x", pady=(0, 8))
        ttk.Button(cfg_btns, text="設定を編集...", style=BTN_SECONDARY,
                   command=self._open_config_editor).pack(side="left", fill="x", expand=True, padx=(0, 4))
        ttk.Button(cfg_btns, text="↻", width=3, style=BTN_TERTIARY,
                   command=self.refresh_config_summary).pack(side="left", padx=(4, 0))

        run_frame = ttk.Frame(left)
        run_frame.pack(fill="x", pady=(4, 8))
        self.run_btn = ttk.Button(run_frame, text="▶ 実行", command=self._on_run,
                                  style=BTN_PRIMARY)
        self.run_btn.pack(side="left", fill="x", expand=True, padx=(0, 4))
        self.stop_btn = ttk.Button(run_frame, text="■ Stop", command=self._on_stop,
                                   state="disabled", style=BTN_SECONDARY)
        self.stop_btn.pack(side="left", fill="x", expand=True, padx=(4, 0))

        folder_frame = ttk.Frame(left)
        folder_frame.pack(fill="x")
        ttk.Button(folder_frame, text="出力フォルダを開く", style=BTN_SECONDARY,
                   command=lambda: (OUTPUT_DIR.mkdir(parents=True, exist_ok=True), _open_in_explorer(OUTPUT_DIR))
                   ).pack(side="left", fill="x", expand=True, padx=(0, 4))
        ttk.Button(folder_frame, text="ログフォルダを開く", style=BTN_SECONDARY,
                   command=lambda: (LOGS_DIR.mkdir(parents=True, exist_ok=True), _open_in_explorer(LOGS_DIR))
                   ).pack(side="left", fill="x", expand=True, padx=(4, 0))

        status_box = ttk.Frame(left)
        status_box.pack(fill="x", pady=(12, 0))
        self.status_label = ttk.Label(status_box, text="待機中", foreground=MUTED_FG)
        self.status_label.pack(anchor="w")
        self.progress = ttk.Progressbar(status_box, mode="indeterminate")
        self.progress_detail = ttk.Label(status_box, text="", foreground=MUTED_FG)

        # ---- 右ペイン
        right = ttk.Frame(paned)
        paned.add(right, weight=1)
        right.rowconfigure(0, weight=1)
        right.columnconfigure(0, weight=1)

        self.log_text = ScrolledText(right, font=LOG_FONT, state="disabled",
                                     background=LOG_BG, foreground=LOG_FG,
                                     insertbackground=LOG_FG)
        self.log_text.grid(row=0, column=0, sticky="nsew")
        for tag, opts in _LOG_TAGS.items():
            self.log_text.tag_configure(tag, **opts)

        log_menu = tk.Menu(self, tearoff=0)
        log_menu.add_command(label="Save Log...", command=self._save_log_as)
        log_menu.add_command(label="Clear Log", command=self._clear_log)
        self.log_text.bind("<Button-3>", lambda e: log_menu.tk_popup(e.x_root, e.y_root))

        out_frame = ttk.LabelFrame(right, text="出力（data/output/）", padding=6)
        out_frame.grid(row=1, column=0, sticky="ew", pady=(6, 0))
        out_frame.columnconfigure(0, weight=1)

        self.summary_var = tk.StringVar(value="（まだ実行していません）")
        self.summary_label = ttk.Label(out_frame, textvariable=self.summary_var,
                                       anchor="w", justify="left", font=UI_FONT_BOLD,
                                       foreground=_SUMMARY_COLORS["idle"])
        self.summary_label.grid(row=0, column=0, sticky="ew")
        self._summary_flash_job = None
        self.open_btn = ttk.Button(out_frame, text="開く", state="disabled",
                                   style=BTN_SECONDARY, command=self._open_last_output)
        self.open_btn.grid(row=0, column=1, padx=(6, 0))

    # -------------------------------------------------- 設定サマリー

    def refresh_config_summary(self):
        config = load_config()
        if config is None:
            self.config_label.config(
                text="config/config.json が見つかりません。\n「設定を編集...」から作成してください。",
                foreground=ERROR_SOFT_FG)
            self._update_root_dir_hint()
            return
        # config を正としてスキャン対象欄を同期（"." は未指定扱い）
        cfg_root = str(config.get("root_dir", "")).strip()
        self.root_dir_var.set("" if cfg_root in ("", ".") else cfg_root)
        excluded_ext = ", ".join(config.get("excluded_extensions", [])) or "(なし)"
        excluded_dir = ", ".join(config.get("excluded_folder_names", [])) or "(なし)"
        self.config_label.config(foreground="", text=(
            f"出力先:\n  {config.get('output_base_dir', '?')}/{config.get('output_filename', '?')}\n\n"
            f"除外拡張子:\n  {excluded_ext}\n\n"
            f"除外フォルダ名:\n  {excluded_dir}"
        ))

    def _update_root_dir_hint(self):
        path = self.root_dir_var.get().strip()
        if not path:
            self.root_dir_hint.config(
                text="未指定です。実行すると従来どおりフォルダ選択ダイアログが開きます。",
                foreground=WARN_FG)
        elif os.path.isdir(path):
            self.root_dir_hint.config(
                text="実行するとこのフォルダをスキャンします。", foreground=MUTED_FG)
        else:
            self.root_dir_hint.config(
                text="⚠ このパスは見つかりません。「参照...」で選び直してください。",
                foreground=ERROR_SOFT_FG)

    def _browse_root_dir(self):
        current = self.root_dir_var.get().strip()
        initial = current if os.path.isdir(current) else os.path.expanduser("~")
        path = filedialog.askdirectory(
            title="スキャンするフォルダを選択してください", initialdir=initial, parent=self)
        if path:
            self.root_dir_var.set(path)
            self._persist_root_dir(path)

    def _persist_root_dir(self, path):
        """スキャン対象を config.json に書き戻す（次回起動時の初期表示用）。"""
        config = load_config()
        if config is None:
            return
        config["root_dir"] = path
        try:
            CONFIG_PATH.write_text(
                json.dumps(config, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        except OSError:
            pass

    def _open_config_editor(self):
        ConfigEditorWindow(self, on_saved=self.refresh_config_summary)

    # -------------------------------------------------- 実行制御

    def _on_run(self):
        if self.is_running:
            return
        if load_config() is None:
            messagebox.showerror(
                "実行できません",
                f"config/config.json が見つかりません。\n先に「設定を編集...」から作成してください。")
            return

        cmd = [sys.executable, "-u", str(BASE_DIR / "src" / "generate_drive_structure.py")]
        root_dir = self.root_dir_var.get().strip()
        if root_dir and not os.path.isdir(root_dir):
            messagebox.showerror(
                "実行できません", f"スキャン対象フォルダが見つかりません:\n{root_dir}")
            return
        if root_dir:
            cmd += ["--root-dir", root_dir]

        snapshot = _snapshot_output_dir()

        self._reset_run_summary()
        self._set_running(True)
        self._append_log(f"=== 実行開始 {datetime.now():%Y-%m-%d %H:%M:%S} ===\n")
        if root_dir:
            self._append_log(f"[launcher] スキャン対象: {root_dir}\n", tag="debug")
        else:
            self._append_log("[launcher] フォルダ選択ダイアログが開きます（前面に表示されない場合はタスクバーを確認）\n", tag="debug")

        threading.Thread(target=self._run_process, args=(cmd, snapshot), daemon=True).start()

    def _run_process(self, cmd, snapshot):
        """バックグラウンドスレッド。tkinter ウィジェットは直接触らず log_queue 経由で渡す。"""
        exit_code = None
        try:
            env = os.environ.copy()
            env["PYTHONIOENCODING"] = "utf-8"
            env["PYTHONUTF8"] = "1"
            proc = subprocess.Popen(
                cmd, cwd=str(BASE_DIR),
                stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                text=True, encoding="utf-8", errors="replace", bufsize=1,
                env=env, creationflags=subprocess.CREATE_NO_WINDOW)
            self.proc = proc
            for line in proc.stdout:
                self.log_queue.put(line)
            exit_code = proc.wait()
        except Exception as exc:
            self.log_queue.put(f"[launcher] 実行エラー: {exc}\n")
        finally:
            self.proc = None
            self.log_queue.put((_SENTINEL_RUN_DONE, (exit_code, snapshot)))

    def _on_run_finished(self, exit_code, snapshot):
        if exit_code == 0:
            self._append_log("=== 終了 (exit code: 0) ===\n", tag="pass")
        elif exit_code is None:
            self._append_log("=== 異常終了（プロセスを起動できませんでした） ===\n", tag="fail")
        else:
            self._append_log(f"=== 終了 (exit code: {exit_code}) ===\n", tag="fail")
        self._set_running(False)

        new_files = _detect_new_output_files(snapshot)
        self._show_run_summary(new_files)

    def _on_stop(self):
        proc = self.proc
        if proc is not None:
            self._append_log("[launcher] Stop 要求を送信しました\n", tag="warn")
            try:
                proc.terminate()
            except OSError:
                pass

    def _set_running(self, running):
        self.is_running = running
        self.run_btn.config(state="disabled" if running else "normal")
        self.stop_btn.config(state="normal" if running else "disabled")
        self.status_label.config(
            text="実行中..." if running else "待機中",
            foreground=ACCENT_FG if running else MUTED_FG)

        if running:
            self._run_started_at = time.monotonic()
            self._scanned_count = 0
            self.progress.pack(fill="x", pady=(4, 2))
            self.progress.start(15)
            self.progress_detail.pack(anchor="w")
            self._tick_elapsed()
        else:
            self.progress.stop()
            self.progress.pack_forget()
            if self._elapsed_job is not None:
                self.after_cancel(self._elapsed_job)
                self._elapsed_job = None
            self.progress_detail.config(text=self._elapsed_text(final=True))

    def _elapsed_text(self, final=False):
        if self._run_started_at is None:
            return ""
        elapsed = int(time.monotonic() - self._run_started_at)
        mm, ss = divmod(elapsed, 60)
        count_part = f" ・ 収集 {self._scanned_count:,} 件" if self._scanned_count else ""
        label = "所要" if final else "経過"
        return f"{label} {mm:02d}:{ss:02d}{count_part}"

    def _tick_elapsed(self):
        if not self.is_running or not self.winfo_exists():
            return
        self.progress_detail.config(text=self._elapsed_text())
        self._elapsed_job = self.after(1000, self._tick_elapsed)

    _SCAN_COUNT_RE = re.compile(r"([\d,]+)\s*件")

    def _note_progress(self, text):
        """スキャン進捗ログから収集件数を拾う（経過表示の「収集 N 件」に使う）。"""
        if "スキャン中" in text or "取得しました" in text:
            m = self._SCAN_COUNT_RE.search(text)
            if m:
                try:
                    self._scanned_count = int(m.group(1).replace(",", ""))
                except ValueError:
                    pass

    def _on_close(self):
        proc = self.proc
        if proc is not None:
            if not messagebox.askyesno("確認", "実行中のプロセスがあります。停止して終了しますか？"):
                return
            try:
                proc.terminate()
            except OSError:
                pass
        self.destroy()

    # -------------------------------------------------- ログ表示

    @staticmethod
    def _get_log_tag(line):
        stripped = line.lstrip()
        if stripped.startswith("==="):
            return "header"
        lower = stripped.lower()
        if "critical" in lower or "traceback" in lower:
            return "fail"
        if stripped.startswith(("エラー", "Error", "ERROR")) or "エラー" in stripped or "error:" in lower:
            return "error"
        if "警告" in stripped or "warn" in lower:
            return "warn"
        if "完了" in stripped or "出力しました" in stripped:
            return "pass"
        return None

    def _append_log(self, text, tag=None):
        """1行だけを即座に反映する（コード内から直接呼ぶ少数呼び出し用）。"""
        self._append_log_lines([(text, tag)])

    def _append_log_lines(self, entries):
        """複数行をまとめて1回のstate切替・see("end")で反映する。

        1行ごとにconfigure(state=...)やsee("end")を呼ぶと、行数が多いとText部品の
        処理コストが積み上がり、メインスレッドが長時間ブロックされる
        （実測: 3万行を1行ずつ処理すると60秒以上かかりWindowsに「応答なし」と判定された）。
        entries は (text, tag_or_None) のリスト。tagがNoneなら内容から自動判定する。
        """
        if not entries:
            return
        self.log_text.configure(state="normal")
        for text, tag in entries:
            if self.is_running:
                self._note_progress(text)
            if tag is None:
                tag = self._get_log_tag(text)
            if tag:
                self.log_text.insert("end", text, tag)
            else:
                self.log_text.insert("end", text)
        self._trim_log_if_needed()
        self.log_text.see("end")
        self.log_text.configure(state="disabled")

    def _trim_log_if_needed(self):
        """表示行数が上限を超えたら先頭から間引く。全文は data/logs/ のログファイルに残るため実害はない。
        呼び出し側で既に state="normal" にしていることが前提。"""
        line_count = int(self.log_text.index("end-1c").split(".")[0])
        if line_count > MAX_LOG_DISPLAY_LINES:
            excess = line_count - MAX_LOG_DISPLAY_LINES
            self.log_text.delete("1.0", f"{excess + 1}.0")

    def _flush_log_queue(self):
        """1回の呼び出しで処理する行数に上限を設ける。上限に達したら残りは次のタイマー
        （100ms後）に持ち越す。バーストが続く間はログ表示が実行に追いつくまで遅延するが、
        メインスレッドが長時間ブロックされて「応答なし」になる事態を防げる。"""
        pending = []
        processed = 0
        while processed < MAX_LOG_LINES_PER_FLUSH:
            try:
                item = self.log_queue.get_nowait()
            except queue.Empty:
                break
            processed += 1
            if isinstance(item, tuple) and item and item[0] is _SENTINEL_RUN_DONE:
                if pending:
                    self._append_log_lines(pending)
                    pending = []
                self._on_run_finished(*item[1])
            else:
                pending.append((item, None))
        if pending:
            self._append_log_lines(pending)

    def _drain_log_queue(self):
        self._flush_log_queue()
        self.after(100, self._drain_log_queue)

    def _clear_log(self):
        self.log_text.configure(state="normal")
        self.log_text.delete("1.0", "end")
        self.log_text.configure(state="disabled")

    def _save_log_as(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".log", initialdir=str(LOGS_DIR),
            filetypes=[("Log", "*.log"), ("すべて", "*.*")])
        if not path:
            return
        try:
            Path(path).write_text(self.log_text.get("1.0", "end-1c"), encoding="utf-8")
        except OSError as exc:
            messagebox.showerror("エラー", f"保存に失敗しました:\n{exc}")

    # -------------------------------------------------- 実行結果サマリー

    def _reset_run_summary(self):
        self._cancel_summary_flash()
        self._last_output_file = None
        self.summary_var.set("実行中...")
        self.summary_label.config(foreground=_SUMMARY_COLORS["running"])
        self.open_btn.config(state="disabled")

    def _show_run_summary(self, new_files):
        if not new_files:
            self.summary_var.set("新しい出力ファイルはありません（キャンセル、または対象0件）")
            self.open_btn.config(state="disabled")
            self._flash_summary(_SUMMARY_COLORS["empty"])
            return

        output_file = new_files[0]
        self._last_output_file = output_file
        try:
            size = _format_filesize(output_file.stat().st_size)
        except OSError:
            size = "?"
        row_count = _count_output_rows(output_file)
        row_part = f" / {row_count:,} 件" if row_count is not None else ""
        self.summary_var.set(f"✔ {output_file.name}{row_part} / {size}")
        self.open_btn.config(state="normal")
        self._flash_summary(_SUMMARY_COLORS["done"])
        self._append_log(f"[launcher] 出力ファイルを検出しました: {output_file.name}\n", tag="debug")

    def _flash_summary(self, color, blinks=3, interval=180):
        self._cancel_summary_flash()
        total_steps = blinks * 2 - 1

        def step(n):
            self._summary_flash_job = None
            on = (n % 2 == 0)
            self.summary_label.config(foreground=color if on else _SUMMARY_COLORS["idle"])
            if n + 1 < total_steps:
                self._summary_flash_job = self.after(interval, step, n + 1)

        step(0)

    def _cancel_summary_flash(self):
        if self._summary_flash_job is not None:
            self.after_cancel(self._summary_flash_job)
            self._summary_flash_job = None

    def _open_last_output(self):
        if self._last_output_file is not None:
            _open_in_explorer(self._last_output_file)


def main():
    app = LauncherApp()
    app.mainloop()


if __name__ == "__main__":
    main()
