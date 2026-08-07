import json
import os
import sys

import pandas as pd
import pytest
from openpyxl import load_workbook

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import generate_drive_structure
from generate_drive_structure import (
    BASE_DIR,
    ConfigError,
    ConfigManager,
    DirectoryScanner,
    ScannedItem,
    create_dataframe_with_fullpath,
    format_size,
    manage_old_output_files,
    save_dataframe_to_excel,
)


# ---- format_size ----

def test_format_size_none():
    assert format_size(None) == ''


def test_format_size_bytes():
    assert format_size(0) == '0 B'
    assert format_size(500) == '500 B'


def test_format_size_kb():
    assert format_size(1536) == '1.5 KB'


def test_format_size_mb():
    assert format_size(2 * 1024 * 1024) == '2.0 MB'


def test_format_size_gb():
    assert format_size(3 * 1024 ** 3) == '3.0 GB'


# ---- ConfigManager ----

def _write_config(tmp_path, data):
    p = tmp_path / "config.json"
    p.write_text(json.dumps(data), encoding="utf-8")
    return str(p)


def test_config_manager_missing_file(tmp_path):
    with pytest.raises(ConfigError):
        ConfigManager(str(tmp_path / "no_such.json"))


def test_config_manager_invalid_json(tmp_path):
    p = tmp_path / "config.json"
    p.write_text("{not valid json", encoding="utf-8")
    with pytest.raises(ConfigError):
        ConfigManager(str(p))


def test_config_manager_missing_required_key(tmp_path):
    path = _write_config(tmp_path, {"root_dir": "C:/tmp"})
    with pytest.raises(ConfigError):
        ConfigManager(path)


def test_config_manager_root_dir_wrong_type(tmp_path):
    path = _write_config(tmp_path, {
        "root_dir": 123,
        "output_base_dir": "output",
        "output_filename": "out.xlsx",
    })
    with pytest.raises(ConfigError):
        ConfigManager(path)


def test_config_manager_defaults(tmp_path):
    path = _write_config(tmp_path, {
        "root_dir": "C:/tmp",
        "output_base_dir": "output",
        "output_filename": "out.xlsx",
    })
    cm = ConfigManager(path)
    assert cm.get_excluded_extensions() == []
    assert cm.get_excluded_folder_names() == []


def test_config_manager_excluded_lists_lowercased(tmp_path):
    path = _write_config(tmp_path, {
        "root_dir": "C:/tmp",
        "output_base_dir": "output",
        "output_filename": "out.xlsx",
        "excluded_extensions": [".LOG", ".Tmp"],
        "excluded_folder_names": ["OLD", "Backup"],
    })
    cm = ConfigManager(path)
    assert cm.get_excluded_extensions() == ['.log', '.tmp']
    assert cm.get_excluded_folder_names() == ['old', 'backup']


def test_config_manager_output_base_dir_relative_resolves_under_base_dir(tmp_path):
    path = _write_config(tmp_path, {
        "root_dir": "C:/tmp",
        "output_base_dir": "some_output",
        "output_filename": "out.xlsx",
    })
    cm = ConfigManager(path)
    assert cm.output_base_dir == os.path.join(BASE_DIR, "some_output")


def test_config_manager_output_base_dir_absolute_kept_as_is(tmp_path):
    abs_dir = str(tmp_path / "abs_output")
    path = _write_config(tmp_path, {
        "root_dir": "C:/tmp",
        "output_base_dir": abs_dir,
        "output_filename": "out.xlsx",
    })
    cm = ConfigManager(path)
    assert cm.output_base_dir == abs_dir


def test_config_manager_save_root_dir_persists(tmp_path):
    path = _write_config(tmp_path, {
        "root_dir": "C:/tmp",
        "output_base_dir": "output",
        "output_filename": "out.xlsx",
    })
    cm = ConfigManager(path)
    cm.save_root_dir("D:/new_dir")
    assert cm.root_dir == "D:/new_dir"

    reloaded = json.loads(open(path, encoding="utf-8").read())
    assert reloaded["root_dir"] == "D:/new_dir"


# ---- DirectoryScanner ----

def _build_sample_tree(tmp_path):
    root = tmp_path / "root"
    (root / "sub").mkdir(parents=True)
    (root / "OLD").mkdir()
    (root / "OLD" / "hidden.txt").write_text("secret", encoding="utf-8")
    (root / "a.txt").write_text("hello", encoding="utf-8")  # 5 bytes
    (root / "b.log").write_text("should be excluded", encoding="utf-8")
    (root / "sub" / "c.txt").write_text("world!", encoding="utf-8")  # 6 bytes
    return root


def test_scan_nonexistent_root_returns_empty(tmp_path):
    scanner = DirectoryScanner(str(tmp_path / "does_not_exist"))
    assert scanner.scan() == []


def test_scan_collects_files_and_folders(tmp_path):
    root = _build_sample_tree(tmp_path)
    scanner = DirectoryScanner(
        str(root),
        excluded_extensions=['.log'],
        excluded_folder_names=['old'],
    )
    items = scanner.scan()
    names = {item.name for item in items}

    assert 'a.txt' in names
    assert 'sub' in names
    assert 'c.txt' in names
    # 除外対象は含まれない
    assert 'b.log' not in names
    assert 'OLD' not in names
    assert 'hidden.txt' not in names


def test_scan_file_size_bytes_populated(tmp_path):
    root = _build_sample_tree(tmp_path)
    scanner = DirectoryScanner(str(root), excluded_extensions=['.log'], excluded_folder_names=['old'])
    items = {item.name: item for item in scanner.scan()}

    assert items['a.txt'].size_bytes == 5
    assert items['a.txt'].item_type == 'file'


def test_scan_folder_size_is_none(tmp_path):
    root = _build_sample_tree(tmp_path)
    scanner = DirectoryScanner(str(root), excluded_extensions=['.log'], excluded_folder_names=['old'])
    items = {item.name: item for item in scanner.scan()}

    assert items['sub'].item_type == 'folder'
    assert items['sub'].size_bytes is None


def test_scan_levels_reflect_depth(tmp_path):
    root = _build_sample_tree(tmp_path)
    scanner = DirectoryScanner(str(root), excluded_extensions=['.log'], excluded_folder_names=['old'])
    items = {item.name: item for item in scanner.scan()}

    assert items['a.txt'].levels == []
    assert items['c.txt'].levels == ['sub']


def test_scan_children_immediately_follow_parent_folder(tmp_path):
    """Excel側のフォルダ単位アウトライン（行折りたたみ）が成立するには、親フォルダの行の
    直後にその子孫が全て連続している必要がある。"Apple"/"Apple2" のような、文字列としての
    フルパスでソートすると区切り文字より前に来てしまう名前の兄弟があっても、
    scan() が返す順序自体は正しいツリー順（親→子孫全部→次の兄弟）であることを確認する。"""
    root = tmp_path / "root"
    (root / "Apple").mkdir(parents=True)
    (root / "Apple" / "notes.txt").write_text("x", encoding="utf-8")
    (root / "Apple2").mkdir()
    (root / "Apple2" / "y.txt").write_text("y", encoding="utf-8")

    scanner = DirectoryScanner(str(root))
    ordered = [(item.name, len(item.levels)) for item in scanner.scan()]

    assert ordered == [
        ("Apple", 0),
        ("notes.txt", 1),
        ("Apple2", 0),
        ("y.txt", 1),
    ]


def test_scan_deep_nesting_keeps_depth_first_order(tmp_path):
    root = tmp_path / "root"
    (root / "A" / "A1" / "A1a").mkdir(parents=True)
    (root / "A" / "A2").mkdir()
    (root / "B").mkdir()
    (root / "A" / "A1" / "A1a" / "deep.txt").write_text("d", encoding="utf-8")
    (root / "A" / "A1" / "mid.txt").write_text("m", encoding="utf-8")
    (root / "A" / "A2" / "leaf.txt").write_text("l", encoding="utf-8")
    (root / "A" / "top.txt").write_text("t", encoding="utf-8")
    (root / "B" / "only.txt").write_text("o", encoding="utf-8")

    scanner = DirectoryScanner(str(root))
    ordered = [(item.name, len(item.levels)) for item in scanner.scan()]

    assert ordered == [
        ("A", 0),
        ("A1", 1),
        ("A1a", 2),
        ("deep.txt", 3),
        ("mid.txt", 2),
        ("A2", 1),
        ("leaf.txt", 2),
        ("top.txt", 1),
        ("B", 0),
        ("only.txt", 1),
    ]


# ---- create_dataframe_with_fullpath ----

def test_create_dataframe_empty_input():
    df = create_dataframe_with_fullpath([])
    assert df.empty


def test_create_dataframe_columns_and_values():
    # 'sub' はルート直下(深さ1)のフォルダ、'c.txt' はその配下(深さ2)のファイル。
    # 深さの異なる階層は「階段状」に列へ配置される仕様のため、'sub' の名前は
    # Level1列に入り、アイテム名列は空になる（アイテム名列は最大深度の行のみ埋まる）。
    items = [
        ScannedItem(r"C:\root\sub", r"C:\root", [], "sub", "folder", None),
        ScannedItem(r"C:\root\sub\c.txt", r"C:\root\sub", ["sub"], "c.txt", "file", 6),
    ]
    df = create_dataframe_with_fullpath(items)

    assert list(df.columns) == [
        '_item_self_path', '_depth', 'タイプ', 'フルパス', 'Level1', 'アイテム名', 'サイズ(バイト)', 'サイズ'
    ]
    df = df.set_index('_item_self_path')

    folder_row = df.loc[r"C:\root\sub"]
    assert folder_row['_depth'] == 0
    assert folder_row['タイプ'] == 'folder'
    assert folder_row['フルパス'] == r"C:\root"
    assert folder_row['サイズ(バイト)'] == ''
    assert folder_row['サイズ'] == ''
    assert folder_row['Level1'] == 'sub'
    assert folder_row['アイテム名'] == ''

    file_row = df.loc[r"C:\root\sub\c.txt"]
    assert file_row['_depth'] == 1
    assert file_row['タイプ'] == 'file'
    assert file_row['フルパス'] == ''  # ファイル行は親フルパスを表示しない
    assert file_row['サイズ(バイト)'] == 6
    assert file_row['サイズ'] == format_size(6)
    assert file_row['Level1'] == 'sub'
    assert file_row['アイテム名'] == 'c.txt'


def test_create_dataframe_no_level_columns_when_flat():
    items = [ScannedItem(r"C:\root\a.txt", r"C:\root", [], "a.txt", "file", 5)]
    df = create_dataframe_with_fullpath(items)
    assert 'Level1' not in df.columns
    assert list(df.columns) == [
        '_item_self_path', '_depth', 'タイプ', 'フルパス', 'アイテム名', 'サイズ(バイト)', 'サイズ'
    ]


# ---- manage_old_output_files ----

def test_manage_old_output_files_deletes_on_confirm(tmp_path, monkeypatch):
    out_dir = tmp_path / "output"
    out_dir.mkdir()
    old_file = out_dir / "drive_structure_20250101_000000.xlsx"
    old_file.write_text("dummy", encoding="utf-8")
    current_file = "drive_structure_20260101_000000.xlsx"

    monkeypatch.setattr(generate_drive_structure, 'confirm_delete_old_files', lambda _files: True)
    manage_old_output_files(str(out_dir), "drive_structure.xlsx", current_file)

    assert not old_file.exists()


def test_manage_old_output_files_keeps_on_decline(tmp_path, monkeypatch):
    out_dir = tmp_path / "output"
    out_dir.mkdir()
    old_file = out_dir / "drive_structure_20250101_000000.xlsx"
    old_file.write_text("dummy", encoding="utf-8")
    current_file = "drive_structure_20260101_000000.xlsx"

    monkeypatch.setattr(generate_drive_structure, 'confirm_delete_old_files', lambda _files: False)
    manage_old_output_files(str(out_dir), "drive_structure.xlsx", current_file)

    assert old_file.exists()


def test_manage_old_output_files_ignores_non_matching_names(tmp_path, monkeypatch):
    out_dir = tmp_path / "output"
    out_dir.mkdir()
    unrelated = out_dir / "notes.txt"
    unrelated.write_text("keep me", encoding="utf-8")

    def _fail(_files):
        raise AssertionError("confirm_delete_old_files should not be called")

    monkeypatch.setattr(generate_drive_structure, 'confirm_delete_old_files', _fail)
    manage_old_output_files(str(out_dir), "drive_structure.xlsx", "drive_structure_20260101_000000.xlsx")


# ---- save_dataframe_to_excel ----

def _sample_nested_df():
    items = [
        ScannedItem(r"C:\root\A", r"C:\root", [], "A", "folder", None),
        ScannedItem(r"C:\root\A\x_very_long_file_name_for_width_check.txt", r"C:\root\A", ["A"], "x_very_long_file_name_for_width_check.txt", "file", 12345),
        ScannedItem(r"C:\root\B", r"C:\root", [], "B", "folder", None),
    ]
    return create_dataframe_with_fullpath(items)


def test_save_dataframe_to_excel_sets_column_widths(tmp_path):
    out_path = tmp_path / "out.xlsx"
    save_dataframe_to_excel(_sample_nested_df(), str(out_path))

    ws = load_workbook(out_path).active
    for col_letter in ("A", "B", "C", "D", "E", "F"):
        dim = ws.column_dimensions.get(col_letter)
        assert dim is not None and dim.width, f"{col_letter}列の幅が設定されていない"
        assert dim.width <= generate_drive_structure.MAX_COLUMN_WIDTH + 2  # マージン込みの上限チェック


def test_save_dataframe_to_excel_freezes_header_row(tmp_path):
    out_path = tmp_path / "out.xlsx"
    save_dataframe_to_excel(_sample_nested_df(), str(out_path))

    ws = load_workbook(out_path).active
    assert ws.freeze_panes == "A2"


def test_save_dataframe_to_excel_sets_autofilter(tmp_path):
    out_path = tmp_path / "out.xlsx"
    save_dataframe_to_excel(_sample_nested_df(), str(out_path))

    ws = load_workbook(out_path).active
    assert ws.auto_filter.ref == "A1:F4"


def test_save_dataframe_to_excel_outline_levels_and_symbols_below(tmp_path):
    out_path = tmp_path / "out.xlsx"
    save_dataframe_to_excel(_sample_nested_df(), str(out_path))

    ws = load_workbook(out_path).active
    # symbols_below=False（openpyxl上は summaryBelow）で、開閉コントロールが
    # 折りたたみ対象の直前（＝親フォルダ自身の行）に来るようにしている
    assert ws.sheet_properties.outlinePr.summaryBelow is False

    # 行2="A"(深さ0) 行3="x...txt"(深さ1、Aの子) 行4="B"(深さ0)
    assert ws.row_dimensions[2].outlineLevel == 0
    assert ws.row_dimensions[3].outlineLevel == 1
    assert ws.row_dimensions[4].outlineLevel == 0
