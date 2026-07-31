import unicodedata

import pandas as pd
from openpyxl.comments import Comment
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.utils.fixed_format import REC_TYPE_DATA, REC_TYPE_END, REC_TYPE_HEADER, REC_TYPE_TRAILER

# 区切り列の見出し文字列（表示上は空白。列ごとに空白の個数を変えて一意な列名にする）
SEPARATOR_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
SEPARATOR_WIDTH = 3

HEADER_FILL = PatternFill(start_color="44546A", end_color="44546A", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)

# 開始位置・文字数の参考表示行（2,3行目）の書式
POSITION_ROW_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
POSITION_ROW_FONT = Font(italic=True, color="595959")

# シート上の行レイアウト: 1=項目名, 2=開始位置, 3=文字数, 4以降=実データ
POSITION_ROWS = (2, 3)
DATA_START_ROW = 4

# レコード種別ごとの行の背景色（データはExcelの既定色のまま）
ROW_FILLS = {
    REC_TYPE_HEADER: PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
    REC_TYPE_TRAILER: PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    REC_TYPE_END: PatternFill(start_color="E2D9F3", end_color="E2D9F3", fill_type="solid"),
}

THIN_SIDE = Side(style="thin", color="B7B7B7")
CELL_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

# 同じレコード種別が連続する行/列をグループ化する際の最小連続数（1つだけの折りたたみは無意味なので対象外）
MIN_GROUP_SIZE = 2

MAX_COLUMN_WIDTH = 60


def _display_width(text):
    """全角文字を2、半角文字を1として表示幅を概算する"""
    width = 0
    for ch in text:
        width += 2 if unicodedata.east_asian_width(ch) in ("W", "F") else 1
    return width


def _comment_text(entry):
    """項目コメントを組み立てる。column_name_forで種別ごとに別カラムに分けているため、
    entryは常に(レコード種別ラベル, rule)の1件のみ"""
    _, rule = entry
    return f"開始位置:{rule['start'] + 1} 文字数:{rule['length']}"


def _group_contiguous_ranges(types, start_index):
    """typesの値が連続する区間を求める（start_indexから採番、1-indexed）

    Noneは常に区切りとして扱い、グループに含めない（行番号・区分等、種別を持たない列向け）。
    戻り値: [(開始, 終了), ...]（MIN_GROUP_SIZE未満の連続は対象外）
    """
    ranges = []
    run_start = None
    prev = None
    idx = start_index
    for t in types:
        if t != prev:
            if run_start is not None and idx - run_start >= MIN_GROUP_SIZE:
                ranges.append((run_start, idx - 1))
            run_start = idx if t is not None else None
            prev = t
        idx += 1
    if run_start is not None and idx - run_start >= MIN_GROUP_SIZE:
        ranges.append((run_start, idx - 1))
    return ranges


def _group_row_ranges(rec_types):
    """レコード種別が連続している行範囲を求める（1-indexed, データ部分はDATA_START_ROWから）"""
    return _group_contiguous_ranges(rec_types, start_index=DATA_START_ROW)


def _column_type(name, field_rules):
    """列名からレコード種別ラベルを引く（field_rulesに無い列はNone）。
    field_rulesはcolumn_name_forで既にレコード種別ごとに分かれた列名をキーにしているため、
    列名1つに対して定義は常に1件（(ラベル, rule)）。"""
    entry = field_rules.get(name)
    return entry[0] if entry else None


def is_separator_column(name):
    """insert_group_separatorsが挿入した空白の区切り列かどうか"""
    return isinstance(name, str) and name != "" and name.strip() == ""


def insert_group_separators(df, field_rules):
    """レコード種別が切り替わる境目に空白の区切り列を挿入したDataFrameを返す

    項目列同士が隙間なく隣接していると、Excelのアウトライン機能は種別ごとに独立した
    折りたたみグループを作れず、全体が1つの範囲に融合してしまう。区切り列を挟むことで
    ヘッダー/データ/トレーラーの各ブロックを個別に開閉できるようにする。
    行番号・区分・レコード種別などfield_rulesに無い列と最初のブロックの間には挿入しない
    （種別ブロック同士が切り替わる境目にのみ挿入する）。
    """
    columns = list(df.columns)
    types = [_column_type(name, field_rules) for name in columns]

    new_order = []
    sep_count = 0
    prev_type = None
    for name, t in zip(columns, types):
        if t is not None and prev_type is not None and t != prev_type:
            sep_count += 1
            new_order.append(" " * sep_count)  # 空白のみ・列ごとに個数を変えて一意にする
        new_order.append(name)
        prev_type = t

    result = pd.DataFrame(index=df.index)
    for col in new_order:
        result[col] = df[col] if col in df.columns else ""
    return result


def _group_column_ranges(columns, field_rules):
    """項目が属するレコード種別ごとに連続する列範囲を求める（1-indexed）

    行番号・区分・レコード種別・区切り列のように field_rules に定義のない列は区切りとして扱う。
    """
    types = [_column_type(name, field_rules) for name in columns]
    return _group_contiguous_ranges(types, start_index=1)


def _insert_position_rows(ws, columns, field_rules):
    """開始位置・文字数を2,3行目に実際の値として書き出す（今までコメントでしか見えなかった情報を
    セルの値として確認できるようにする）。項目に紐づかない列（行番号・区分・レコード種別・区切り列）は空欄にする。
    ヘッダー行(1行目)より下を2行分挿入するため、以降の実データはDATA_START_ROW(4行目)から始まる。
    """
    ws.insert_rows(POSITION_ROWS[0], amount=len(POSITION_ROWS))
    for col_idx, name in enumerate(columns, start=1):
        entry = field_rules.get(name)
        if not entry:
            continue
        _, rule = entry
        ws.cell(row=POSITION_ROWS[0], column=col_idx, value=rule["start"] + 1)
        ws.cell(row=POSITION_ROWS[1], column=col_idx, value=rule["length"])

    for row_idx in POSITION_ROWS:
        for cell in ws[row_idx]:
            cell.fill = POSITION_ROW_FILL
            cell.font = POSITION_ROW_FONT
            cell.border = CELL_BORDER


def style_output_sheet(ws, df, field_rules, rec_type_col="レコード種別"):
    """ヘッダー行の書式・項目コメント・開始位置/文字数の参考表示行、レコード種別ごとの行色・
    行/列グループ化、罫線、列幅自動調整をまとめて適用する

    field_rules: {出力Excel上のカラム名: (レコード種別ラベル, {"start", "length"})}
    （column_name_forで既にレコード種別ごとに別カラム名へ分けているため、カラム名は必ず1件の定義に対応する）
    """
    _insert_position_rows(ws, df.columns, field_rules)

    # 項目名・開始位置・文字数の帯(1〜3行目)を常に表示したまま実データをスクロールできるようにする
    ws.freeze_panes = f"A{DATA_START_ROW}"

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = CELL_BORDER
        entry = field_rules.get(cell.value)
        if entry:
            cell.comment = Comment(_comment_text(entry), "Fixed2Excel")

    for row in ws.iter_rows(min_row=DATA_START_ROW):
        for cell in row:
            cell.border = CELL_BORDER
            # 数字だけの項目値（会員番号・区分コード等）をExcel上で編集すると、既定の「標準」書式では
            # 入力時に数値と解釈されて先頭のゼロが消えてしまう（"000"→"001"と打っても"1"になる）。
            # 文字列書式に固定してこの事故を防ぐ。
            cell.number_format = "@"

    if rec_type_col in df.columns:
        rec_types = list(df[rec_type_col])

        for row_idx, rec_type in enumerate(rec_types, start=DATA_START_ROW):
            fill = ROW_FILLS.get(rec_type)
            if fill:
                for cell in ws[row_idx]:
                    cell.fill = fill

        # 同じレコード種別が連続する行を折りたたみ可能なグループにする（Excelのアウトライン機能）。
        # データ部は開いた状態のまま、それ以外（ヘッダー/トレーラー/エンドレコード等が複数行に
        # なる場合）は初期状態で折りたたんでおく（データ部だけをすぐ見たいという要望に対応）。
        for start_row, end_row in _group_row_ranges(rec_types):
            group_type = rec_types[start_row - DATA_START_ROW]
            collapse = group_type != REC_TYPE_DATA
            for row_idx in range(start_row, end_row + 1):
                ws.row_dimensions[row_idx].outlineLevel = 1
                if collapse:
                    ws.row_dimensions[row_idx].hidden = True

    for col_idx, col_name in enumerate(df.columns, start=1):
        letter = get_column_letter(col_idx)
        if is_separator_column(col_name):
            ws.column_dimensions[letter].width = SEPARATOR_WIDTH
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    cell.fill = SEPARATOR_FILL
            continue
        values = [str(col_name)] + [str(v) for v in df[col_name].fillna("")]
        max_width = max(_display_width(v) for v in values)
        ws.column_dimensions[letter].width = min(max_width + 2, MAX_COLUMN_WIDTH)

    # レコード種別ごとに連続する項目列を折りたたみ可能なグループにする（Excelのアウトライン機能）。
    # データ部の列は開いた状態のまま、それ以外（ヘッダー/トレーラー/エンドレコード部）は
    # 初期状態で折りたたんでおく（データ部だけをすぐ見たいという要望に対応）。
    columns = list(df.columns)
    for start_col, end_col in _group_column_ranges(columns, field_rules):
        group_type = _column_type(columns[start_col - 1], field_rules)
        collapse = group_type != REC_TYPE_DATA
        for col_idx in range(start_col, end_col + 1):
            letter = get_column_letter(col_idx)
            ws.column_dimensions[letter].outlineLevel = 1
            if collapse:
                ws.column_dimensions[letter].hidden = True
