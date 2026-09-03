import logging
import os
import sys

import openpyxl
import pandas as pd
import pytest

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from src.app_context import AppContext, create_context
from src import app_context as app_context_module
from src.handlers import excel_to_fixed as excel_to_fixed_module
from src.handlers import fixed_to_excel as fixed_to_excel_module
from src.handlers.diff_checker import check_all, diff_rows, restore_and_check, restored_name_for
from src.handlers.excel_to_fixed import build_fixed_line, pad_value_to_bytes, restore_all
from src.handlers.fixed_to_excel import _flatten_field_rules, convert_all, process_file
from src.handlers.mapping_handler import (
    add_mapping_entry,
    build_or_update_mapping,
    find_existing_config,
    load_mapping,
    remove_mapping_entry,
)
from src.handlers.setup_handler import _with_padding_filler, init_environment
from src.utils.excel_style import (
    _comment_text,
    _group_column_ranges,
    _group_row_ranges,
    _insert_position_rows,
    insert_group_separators,
    is_separator_column,
    style_output_sheet,
)
from src.utils.fixed_format import (
    AmbiguousKeywordMatchError,
    analysis_excel_name,
    build_field_columns,
    load_config_rules,
    match_config,
    read_mapping_csv,
    resolve_config_path,
    restored_txt_name,
    validate_config_rules,
)

ENCODING = "cp932"
RECORD_TYPE_CODES = {"header": ["1"], "trailer": ["8", "9"]}

logger = logging.getLogger("test")


def _make_ctx(dirs, mapping_csv, mapping_columns, **overrides):
    """テスト用にAppContextを組み立てる（record_type_codes等は既定値を使い回す）"""
    fields = {
        "app_name": "test",
        "encoding": ENCODING,
        "dirs": dirs,
        "mapping_csv": str(mapping_csv),
        "mapping_columns": mapping_columns,
        "record_type_codes": RECORD_TYPE_CODES,
        "logger": logger,
    }
    fields.update(overrides)
    return AppContext(**fields)


def _write_config_excel(path):
    data_sheet = pd.DataFrame([
        ["開始位置", 1, 5],
        ["文字数", 1, 4],
    ], columns=["区分", "会員番号", "名前"])

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        data_sheet.to_excel(writer, sheet_name="データ", index=False)


# ---- process_file ----

def test_process_file_parses_data_rows(tmp_path):
    config_path = tmp_path / "config.xlsx"
    _write_config_excel(config_path)

    txt_path = tmp_path / "sample.txt"
    with open(txt_path, "wb") as f:
        f.write("2XXXABCD".encode(ENCODING) + b"\r\n")

    from src.utils.fixed_format import load_config_rules
    rules = load_config_rules(config_path)

    df = process_file(txt_path, rules, ENCODING, RECORD_TYPE_CODES)

    assert list(df["会員番号"]) == ["2"]
    assert list(df["名前"]) == ["ABCD"]
    assert list(df["レコード種別"]) == ["データ"]


def test_process_file_skips_blank_lines(tmp_path):
    config_path = tmp_path / "config.xlsx"
    _write_config_excel(config_path)

    txt_path = tmp_path / "sample.txt"
    with open(txt_path, "wb") as f:
        f.write(b"\r\n")
        f.write("2ABCDテスト".encode(ENCODING) + b"\r\n")

    from src.utils.fixed_format import load_config_rules
    rules = load_config_rules(config_path)

    df = process_file(txt_path, rules, ENCODING, RECORD_TYPE_CODES)
    assert len(df) == 1


# ---- pad_value_to_bytes / build_fixed_line ----

def test_pad_value_to_bytes_right_justifies_digits():
    assert pad_value_to_bytes("12", 5, ENCODING) == b"   12"


def test_pad_value_to_bytes_left_justifies_text():
    assert pad_value_to_bytes("AB", 5, ENCODING) == b"AB   "


def test_pad_value_to_bytes_truncates_when_too_long():
    assert pad_value_to_bytes("ABCDEF", 3, ENCODING) == b"ABC"


def test_build_fixed_line_places_fields_at_start_positions():
    rules = [
        {"name": "会員番号", "start": 0, "length": 1},
        {"name": "名前", "start": 1, "length": 4},
    ]
    row = {"会員番号": "2", "名前": "AB"}
    line = build_fixed_line(row, rules, ENCODING)
    assert line == b"2AB  "


# ---- match_config ----

def test_match_config_finds_matching_keyword():
    columns = {"keyword": "kw", "config_name": "cfg"}
    df_map = pd.DataFrame([{"kw": "KAIIN", "cfg": "kaiin.xlsx"}])
    assert match_config("KAIIN_SAMPLE.txt", df_map, columns) == "kaiin.xlsx"


def test_match_config_returns_none_when_no_match():
    columns = {"keyword": "kw", "config_name": "cfg"}
    df_map = pd.DataFrame([{"kw": "OTHER", "cfg": "other.xlsx"}])
    assert match_config("KAIIN_SAMPLE.txt", df_map, columns) is None


# 部分一致(in)判定のため、'SAMPLE'と'KAIIN_SAMPLE'のように複数キーワードが同じファイル名に
# 一致し得る。どちらを採用すべきか機械的に決められないため、静かにどちらかを選ぶのではなく
# エラーとして検知する（設定ミスの兆候として扱う）。

def test_match_config_raises_on_ambiguous_multiple_matches():
    columns = {"keyword": "kw", "config_name": "cfg"}
    df_map = pd.DataFrame([
        {"kw": "SAMPLE", "cfg": "sample.xlsx"},
        {"kw": "KAIIN_SAMPLE", "cfg": "kaiin.xlsx"},
    ])
    with pytest.raises(AmbiguousKeywordMatchError):
        match_config("KAIIN_SAMPLE_20260731.txt", df_map, columns)


# ---- resolve_config_path: convert_all/restore_all/check_all共通の設定解決処理 ----
# キーワード不一致・複数一致・設定ファイル欠落のいずれもここでログして返り値をNoneにするため、
# 呼び出し側は`if not config_path: continue`とするだけでよい。

def test_resolve_config_path_returns_full_path_on_unambiguous_match(tmp_path):
    configs_dir = tmp_path / "configs"
    configs_dir.mkdir()
    (configs_dir / "config.xlsx").write_bytes(b"")

    columns = {"keyword": "kw", "config_name": "cfg"}
    df_map = pd.DataFrame([{"kw": "KAIIN", "cfg": "config.xlsx"}])

    result = resolve_config_path("KAIIN_SAMPLE.txt", df_map, columns, str(configs_dir), logger)
    assert result == str(configs_dir / "config.xlsx")


def test_resolve_config_path_returns_none_when_no_keyword_matches(tmp_path, caplog):
    columns = {"keyword": "kw", "config_name": "cfg"}
    df_map = pd.DataFrame([{"kw": "OTHER", "cfg": "config.xlsx"}])

    with caplog.at_level(logging.INFO):
        result = resolve_config_path("KAIIN_SAMPLE.txt", df_map, columns, str(tmp_path), logger)

    assert result is None
    assert "スキップ" in caplog.text


def test_resolve_config_path_returns_none_on_ambiguous_match(tmp_path, caplog):
    columns = {"keyword": "kw", "config_name": "cfg"}
    df_map = pd.DataFrame([
        {"kw": "SAMPLE", "cfg": "sample.xlsx"},
        {"kw": "KAIIN_SAMPLE", "cfg": "kaiin.xlsx"},
    ])

    with caplog.at_level(logging.INFO):
        result = resolve_config_path("KAIIN_SAMPLE_20260731.txt", df_map, columns, str(tmp_path), logger)

    assert result is None
    assert "複数のキーワードが部分一致しました" in caplog.text


def test_resolve_config_path_returns_none_when_config_file_missing(tmp_path, caplog):
    configs_dir = tmp_path / "configs"
    configs_dir.mkdir()  # config.xlsx自体は作らない

    columns = {"keyword": "kw", "config_name": "cfg"}
    df_map = pd.DataFrame([{"kw": "KAIIN", "cfg": "config.xlsx"}])

    with caplog.at_level(logging.INFO):
        result = resolve_config_path("KAIIN_SAMPLE.txt", df_map, columns, str(configs_dir), logger)

    assert result is None
    assert "設定ファイル不明" in caplog.text


# ---- build_or_update_mapping ----

def test_build_or_update_mapping_appends_new_input_file(tmp_path):
    configs_dir = tmp_path / "configs"
    input_dir = tmp_path / "input"
    configs_dir.mkdir()
    input_dir.mkdir()
    (configs_dir / "sample.xlsx").write_bytes(b"")
    (input_dir / "KAIIN_SAMPLE.txt").write_bytes(b"")

    mapping_csv = tmp_path / "mapping.csv"
    columns = {"keyword": "kw", "config_name": "cfg", "note": "note"}

    ctx = _make_ctx({"configs": str(configs_dir), "input": str(input_dir)}, mapping_csv, columns)
    added = build_or_update_mapping(ctx)

    assert added == 1
    df_map = pd.read_csv(mapping_csv, encoding=ENCODING)
    assert df_map.iloc[0]["kw"] == "KAIIN_SAMPLE"
    assert df_map.iloc[0]["cfg"] == "sample.xlsx"


# 過去に見つかったバグ: mapping.csv を data/configs/ 配下に置く運用に変えた際、
# 設定ファイル一覧のスキャンが拡張子(.xlsx/.csv)だけで判定していたため、mapping.csv
# 自身を「設定ファイル」として拾ってしまい、新規input行の既定Configにmapping.csvが
# 提案されることがあった。

def test_build_or_update_mapping_excludes_mapping_csv_itself_when_colocated(tmp_path):
    configs_dir = tmp_path / "configs"
    input_dir = tmp_path / "input"
    configs_dir.mkdir()
    input_dir.mkdir()
    (configs_dir / "sample.xlsx").write_bytes(b"")
    (input_dir / "KAIIN_SAMPLE.txt").write_bytes(b"")

    mapping_csv = configs_dir / "mapping.csv"  # configs/ 直下に配置
    columns = {"keyword": "kw", "config_name": "cfg", "note": "note"}

    ctx = _make_ctx({"configs": str(configs_dir), "input": str(input_dir)}, mapping_csv, columns)
    build_or_update_mapping(ctx)

    df_map = pd.read_csv(mapping_csv, encoding=ENCODING)
    assert df_map.iloc[0]["cfg"] == "sample.xlsx"  # mapping.csv自身が既定Configになっていない


def test_build_or_update_mapping_skips_already_registered(tmp_path):
    configs_dir = tmp_path / "configs"
    input_dir = tmp_path / "input"
    configs_dir.mkdir()
    input_dir.mkdir()
    (input_dir / "KAIIN_SAMPLE.txt").write_bytes(b"")

    mapping_csv = tmp_path / "mapping.csv"
    columns = {"keyword": "kw", "config_name": "cfg", "note": "note"}
    pd.DataFrame([{"kw": "KAIIN_SAMPLE", "cfg": "x.xlsx", "note": ""}]).to_csv(
        mapping_csv, index=False, encoding=ENCODING
    )

    ctx = _make_ctx({"configs": str(configs_dir), "input": str(input_dir)}, mapping_csv, columns)
    added = build_or_update_mapping(ctx)

    assert added == 0


def test_build_or_update_mapping_backs_up_existing_file_before_overwrite(tmp_path):
    configs_dir = tmp_path / "configs"
    input_dir = tmp_path / "input"
    configs_dir.mkdir()
    input_dir.mkdir()
    (input_dir / "NEW_FILE.txt").write_bytes(b"")

    mapping_csv = tmp_path / "mapping.csv"
    columns = {"keyword": "kw", "config_name": "cfg", "note": "note"}
    pd.DataFrame([{"kw": "EXISTING", "cfg": "x.xlsx", "note": ""}]).to_csv(
        mapping_csv, index=False, encoding=ENCODING
    )

    ctx = _make_ctx({"configs": str(configs_dir), "input": str(input_dir)}, mapping_csv, columns)
    build_or_update_mapping(ctx)

    backups = list(tmp_path.glob("mapping.csv.bak_*"))
    assert len(backups) == 1, f"バックアップが作成されていない: {list(tmp_path.iterdir())}"

    backup_df = pd.read_csv(backups[0], encoding=ENCODING)
    assert list(backup_df["kw"]) == ["EXISTING"]  # バックアップは追記前の内容を保持


def test_build_or_update_mapping_no_backup_when_nothing_to_add(tmp_path):
    configs_dir = tmp_path / "configs"
    input_dir = tmp_path / "input"
    configs_dir.mkdir()
    input_dir.mkdir()
    (input_dir / "EXISTING.txt").write_bytes(b"")

    mapping_csv = tmp_path / "mapping.csv"
    columns = {"keyword": "kw", "config_name": "cfg", "note": "note"}
    pd.DataFrame([{"kw": "EXISTING", "cfg": "x.xlsx", "note": ""}]).to_csv(
        mapping_csv, index=False, encoding=ENCODING
    )

    ctx = _make_ctx({"configs": str(configs_dir), "input": str(input_dir)}, mapping_csv, columns)
    build_or_update_mapping(ctx)  # 追記対象なし

    assert list(tmp_path.glob("mapping.csv.bak_*")) == []


# ---- 固定長 -> Excel -> 固定長 の往復変換（回帰テスト） ----
# 過去に見つかったバグ: (1) 16桁の数字文字列がpd.read_excelの型推論でfloatに
# 化けて桁が欠落する、(2) 先頭1バイトのレコード種別コードが復元時に失われる。

def test_roundtrip_preserves_record_code_and_long_numeric_field(tmp_path):
    configs_dir = tmp_path / "configs"
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    recreated_dir = tmp_path / "recreated"
    for d in (configs_dir, input_dir, output_dir, recreated_dir):
        d.mkdir()

    header_sheet = pd.DataFrame([["開始位置", 2], ["文字数", 8]], columns=["区分", "作成年月日"])
    data_sheet = pd.DataFrame([["開始位置", 2], ["文字数", 16]], columns=["区分", "会員番号"])
    config_path = configs_dir / "config.xlsx"
    with pd.ExcelWriter(config_path, engine="openpyxl") as writer:
        data_sheet.to_excel(writer, sheet_name="データ", index=False)
        header_sheet.to_excel(writer, sheet_name="ヘッダー", index=False)

    h_line = "1" + "20260730"
    d_line = "2" + "1000000000000001"  # 16桁の数字文字列(会員番号)
    input_path = input_dir / "SAMPLE.txt"
    with open(input_path, "wb") as f:
        for line in [h_line, d_line]:
            f.write(line.encode(ENCODING) + b"\r\n")

    mapping_csv = tmp_path / "mapping.csv"
    columns = {"keyword": "判定キーワード(input側)", "config_name": "設定ファイル名(configs内)"}
    pd.DataFrame([{
        "判定キーワード(input側)": "SAMPLE",
        "設定ファイル名(configs内)": "config.xlsx",
    }]).to_csv(mapping_csv, index=False, encoding=ENCODING)

    dirs = {
        "configs": str(configs_dir), "input": str(input_dir),
        "output": str(output_dir), "recreated": str(recreated_dir),
    }
    ctx = _make_ctx(dirs, mapping_csv, columns)
    convert_all(ctx)
    restore_all(ctx)

    restored_path = recreated_dir / "RESTORED_SAMPLE.txt"
    with open(restored_path, "rb") as f:
        restored_lines = [l.rstrip(b"\r\n").decode(ENCODING) for l in f if l.strip()]

    assert restored_lines == [h_line, d_line]


# ---- 出力Excelの項目コメント・列分割（同名項目がレコード種別をまたぐケース） ----
# 過去に見つかったバグ: 「予備」のように同じ列名がヘッダー/データ/トレーラーで別々の
# 開始位置・文字数を持つとき、素朴に辞書へ詰めると後勝ちで上書きされ、片方の情報が消えて
# 誤ったコメントになっていた。
# その後、開始位置・文字数を実際の行(2,3行目)として書き出す機能を追加したことで、1カラムに
# 複数の定義を混在させられなくなったため、column_name_forで「項目名（種別）」の別カラムに
# 分割する方式に変更した。

def test_flatten_field_rules_splits_colliding_name_into_separate_columns():
    config_rules = {
        "D": [{"name": "予備", "start": 66, "length": 4}],
        "H": [{"name": "予備", "start": 49, "length": 20}],
        "T": None,
    }
    flattened = _flatten_field_rules(config_rules)
    assert flattened["予備（データ）"] == ("データ", config_rules["D"][0])
    assert flattened["予備（ヘッダー）"] == ("ヘッダー", config_rules["H"][0])
    assert "予備" not in flattened


def test_comment_text_formats_single_entry():
    entry = ("データ", {"start": 66, "length": 4})
    assert _comment_text(entry) == "開始位置:67 文字数:4"


# ---- build_field_columns: 同一レコード種別内で同名項目が繰り返されるケース ----
# 実データ(SMBC請求データ)で発覚したバグ: 設定Excelのトレーラーシートで、本来別々の13項目が
# すべて同じ名前「区分コード」で定義されていた（開始位置・文字数はそれぞれ異なる）。
# 種別ごとの列分割（「項目名（種別）」）だけでは同種別内の重複を区別できず、13個の値が
# 1カラムに上書きされて実データが12個分消えていた。同種別内で同名が繰り返される場合は
# 出現順の連番も付けて完全に一意化する。

def test_build_field_columns_numbers_duplicate_names_within_same_record_type():
    config_rules = {
        "T": [
            {"name": "区分コード", "start": 2, "length": 2},
            {"name": "区分コード", "start": 4, "length": 4},
            {"name": "区分コード", "start": 8, "length": 6},
        ],
        "D": None,
        "H": None,
        "E": None,
    }
    field_columns = build_field_columns(config_rules)
    columns = [e["column"] for e in field_columns["T"]]
    assert columns == ["区分コード（トレーラー）1", "区分コード（トレーラー）2", "区分コード（トレーラー）3"]
    # ruleの実体（開始位置・文字数）は元のまま保持されている
    assert [e["rule"]["start"] for e in field_columns["T"]] == [2, 4, 8]


def test_build_field_columns_keeps_plain_name_when_no_collision():
    config_rules = {
        "D": [{"name": "会員番号", "start": 1, "length": 16}],
        "H": None,
        "T": None,
        "E": None,
    }
    field_columns = build_field_columns(config_rules)
    assert [e["column"] for e in field_columns["D"]] == ["会員番号"]


# ---- 開始位置・文字数の参考表示行（2,3行目）を書き出す ----
# 依頼: 今までコメント（ホバー）でしか見えなかった開始位置・文字数を、出力Excelの2,3行目に
# 実際の値として書き出してほしい。項目に紐づかない列（行番号・区分・レコード種別・区切り列）は空欄にする。

def test_insert_position_rows_writes_start_and_length(tmp_path):
    df = pd.DataFrame([{"行番号": 1, "区分": "2", "会員番号": "1"}])
    out_path = tmp_path / "position_rows.xlsx"
    field_rules = {"会員番号": ("データ", {"start": 1, "length": 16})}
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sheet1")
        ws = writer.sheets["Sheet1"]
        _insert_position_rows(ws, df.columns, field_rules)

    wb = openpyxl.load_workbook(out_path)
    ws = wb.active

    member_col = list(df.columns).index("会員番号") + 1
    assert ws.cell(row=2, column=member_col).value == 2  # start(0-index1) + 1
    assert ws.cell(row=3, column=member_col).value == 16

    row_no_col = list(df.columns).index("行番号") + 1
    assert ws.cell(row=2, column=row_no_col).value is None  # 項目に紐づかない列は空欄
    assert ws.cell(row=3, column=row_no_col).value is None

    # 元データは2行分シフトして4行目から
    assert ws.cell(row=4, column=member_col).value == "1"


# ---- 同名項目のレコード種別ごとの列分割（開始位置・文字数の書き出しと両立させるため） ----
# 依頼: 「予備」のように複数レコード種別で共有される項目名があっても、ヘッダー部・データ部で
# それぞれ独立した列として出力してほしい（1列に混在させると開始位置・文字数を2,3行目に
# 書き出せない）。

def test_convert_all_splits_colliding_field_name_and_writes_position_rows(tmp_path):
    configs_dir = tmp_path / "configs"
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    for d in (configs_dir, input_dir, output_dir):
        d.mkdir()

    header_sheet = pd.DataFrame([
        ["開始位置", 2, 10],
        ["文字数", 8, 5],
    ], columns=["区分", "作成年月日", "予備"])
    data_sheet = pd.DataFrame([
        ["開始位置", 2, 20],
        ["文字数", 16, 3],
    ], columns=["区分", "会員番号", "予備"])
    config_path = configs_dir / "config.xlsx"
    with pd.ExcelWriter(config_path, engine="openpyxl") as writer:
        data_sheet.to_excel(writer, sheet_name="データ", index=False)
        header_sheet.to_excel(writer, sheet_name="ヘッダー", index=False)

    h_line = "1" + "20260730" + "HHHHH"
    d_line = "2" + "1000000000000001" + "XX" + "DDD"
    input_path = input_dir / "SAMPLE.txt"
    with open(input_path, "wb") as f:
        for line in [h_line, d_line]:
            f.write(line.encode(ENCODING) + b"\r\n")

    mapping_csv = tmp_path / "mapping.csv"
    columns = {"keyword": "判定キーワード(input側)", "config_name": "設定ファイル名(configs内)"}
    pd.DataFrame([{
        "判定キーワード(input側)": "SAMPLE",
        "設定ファイル名(configs内)": "config.xlsx",
    }]).to_csv(mapping_csv, index=False, encoding=ENCODING)

    dirs = {
        "configs": str(configs_dir), "input": str(input_dir),
        "output": str(output_dir), "recreated": str(tmp_path / "recreated"),
    }
    ctx = _make_ctx(dirs, mapping_csv, columns)

    convert_all(ctx)

    wb = openpyxl.load_workbook(output_dir / "解析結果_SAMPLE.xlsx")
    ws = wb.active
    header = [cell.value for cell in ws[1]]

    assert "予備（ヘッダー）" in header
    assert "予備（データ）" in header
    assert "予備" not in header  # 分割前の名前は残らない

    header_col = header.index("予備（ヘッダー）") + 1
    data_col = header.index("予備（データ）") + 1

    # 2行目=開始位置, 3行目=文字数（それぞれ別カラムなので別々の値を持てる）
    assert ws.cell(row=2, column=header_col).value == 10
    assert ws.cell(row=3, column=header_col).value == 5
    assert ws.cell(row=2, column=data_col).value == 20
    assert ws.cell(row=3, column=data_col).value == 3

    # 実データはrow4から。ヘッダー行では「予備（データ）」が空欄、データ行では「予備（ヘッダー）」が空欄
    assert ws.cell(row=4, column=header_col).value == "HHHHH"
    assert ws.cell(row=4, column=data_col).value in (None, "")
    assert ws.cell(row=5, column=data_col).value == "DDD"
    assert ws.cell(row=5, column=header_col).value in (None, "")


# ---- 同一レコード種別内で同名項目が繰り返される設定Excelでの往復変換（回帰テスト） ----
# 実データ(SMBC請求データのトレーラー: 13個の項目が全て「区分コード」という同じ名前)で発覚した
# バグの再現。同種別内の重複を連番で区別しないと、13個の値が1カラムに上書きされて消える。

def test_roundtrip_preserves_all_values_when_same_name_repeats_within_record_type(tmp_path):
    configs_dir = tmp_path / "configs"
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    recreated_dir = tmp_path / "recreated"
    for d in (configs_dir, input_dir, output_dir, recreated_dir):
        d.mkdir()

    # トレーラーの「区分コード」を3つ、開始位置・文字数だけ違えて同名で定義（実データの再現）
    trailer_sheet = pd.DataFrame([
        ["開始位置", 2, 4, 10],
        ["文字数", 2, 6, 6],
    ], columns=["区分", "区分コード", "区分コード", "区分コード"])
    config_path = configs_dir / "config.xlsx"
    with pd.ExcelWriter(config_path, engine="openpyxl") as writer:
        pd.DataFrame([["開始位置", 1], ["文字数", 1]], columns=["区分", "値"]).to_excel(
            writer, sheet_name="データ", index=False
        )
        trailer_sheet.to_excel(writer, sheet_name="トレーラー", index=False)

    t_line = "8" + "93" + "701931" + "000002"  # 区分コード1='93', 2='701931', 3='000002'
    input_path = input_dir / "SAMPLE.txt"
    with open(input_path, "wb") as f:
        f.write(t_line.encode(ENCODING) + b"\r\n")

    mapping_csv = tmp_path / "mapping.csv"
    columns = {"keyword": "判定キーワード(input側)", "config_name": "設定ファイル名(configs内)"}
    pd.DataFrame([{
        "判定キーワード(input側)": "SAMPLE",
        "設定ファイル名(configs内)": "config.xlsx",
    }]).to_csv(mapping_csv, index=False, encoding=ENCODING)

    dirs = {
        "configs": str(configs_dir), "input": str(input_dir),
        "output": str(output_dir), "recreated": str(recreated_dir),
    }
    ctx = _make_ctx(dirs, mapping_csv, columns)

    convert_all(ctx)

    wb = openpyxl.load_workbook(output_dir / "解析結果_SAMPLE.xlsx")
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    values = {name: ws.cell(row=4, column=idx + 1).value for idx, name in enumerate(header)}

    # 3つとも別カラムとして残り、値が上書きで消えていないこと
    assert values["区分コード（トレーラー）1"] == "93"
    assert values["区分コード（トレーラー）2"] == "701931"
    assert values["区分コード（トレーラー）3"] == "000002"

    restore_all(ctx)
    restored_path = recreated_dir / "RESTORED_SAMPLE.txt"
    with open(restored_path, "rb") as f:
        restored_lines = [l.rstrip(b"\r\n").decode(ENCODING) for l in f if l.strip()]

    assert restored_lines == [t_line]  # 往復後も完全一致（値の欠落がない）


# ---- 往復変換時に開始位置/文字数の参考表示行を実データとして誤読しない ----
# 2,3行目に開始位置・文字数を書き出すようになったため、Excel→固定長復元時にこれを
# 実データ行として読み込んでしまうと、存在しないレコードが復元結果に混入してしまう。

def test_restore_all_skips_position_rows(tmp_path):
    configs_dir = tmp_path / "configs"
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    recreated_dir = tmp_path / "recreated"
    for d in (configs_dir, input_dir, output_dir, recreated_dir):
        d.mkdir()

    data_sheet = pd.DataFrame([["開始位置", 2], ["文字数", 1]], columns=["区分", "値"])
    config_path = configs_dir / "config.xlsx"
    with pd.ExcelWriter(config_path, engine="openpyxl") as writer:
        data_sheet.to_excel(writer, sheet_name="データ", index=False)

    input_path = input_dir / "SAMPLE.txt"
    with open(input_path, "wb") as f:
        f.write(b"2A\r\n")

    mapping_csv = tmp_path / "mapping.csv"
    columns = {"keyword": "判定キーワード(input側)", "config_name": "設定ファイル名(configs内)"}
    pd.DataFrame([{
        "判定キーワード(input側)": "SAMPLE",
        "設定ファイル名(configs内)": "config.xlsx",
    }]).to_csv(mapping_csv, index=False, encoding=ENCODING)

    dirs = {
        "configs": str(configs_dir), "input": str(input_dir),
        "output": str(output_dir), "recreated": str(recreated_dir),
    }
    ctx = _make_ctx(dirs, mapping_csv, columns)

    convert_all(ctx)
    restore_all(ctx)

    restored_path = recreated_dir / "RESTORED_SAMPLE.txt"
    with open(restored_path, "rb") as f:
        restored_lines = [l.rstrip(b"\r\n") for l in f if l.strip()]

    assert restored_lines == [b"2A"]  # 開始位置/文字数の行が復元結果に混入していない


# ---- 出力先ファイルが他アプリ（Excel等）で開いている場合のエラーハンドリング ----
# 実際に開発中、出力Excelをexcelで開いたまま再変換してPermissionErrorでクラッシュした。
# 1ファイルの書き込み/読み込み失敗で処理全体を止めず、他のファイルは続行できるようにする。

def _write_simple_config(config_path):
    data_sheet = pd.DataFrame([["開始位置", 1], ["文字数", 1]], columns=["区分", "値"])
    with pd.ExcelWriter(config_path, engine="openpyxl") as writer:
        data_sheet.to_excel(writer, sheet_name="データ", index=False)


def test_convert_all_skips_locked_output_and_continues_other_files(tmp_path, monkeypatch):
    configs_dir = tmp_path / "configs"
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    configs_dir.mkdir()
    input_dir.mkdir()

    _write_simple_config(configs_dir / "config.xlsx")
    for name in ("LOCKED.txt", "OK.txt"):
        with open(input_dir / name, "wb") as f:
            f.write(b"2A\r\n")

    mapping_csv = tmp_path / "mapping.csv"
    columns = {"keyword": "kw", "config_name": "cfg"}
    pd.DataFrame([
        {"kw": "LOCKED", "cfg": "config.xlsx"},
        {"kw": "OK", "cfg": "config.xlsx"},
    ]).to_csv(mapping_csv, index=False, encoding=ENCODING)

    dirs = {
        "configs": str(configs_dir), "input": str(input_dir),
        "output": str(output_dir), "recreated": str(tmp_path / "recreated"),
    }
    ctx = _make_ctx(dirs, mapping_csv, columns)

    real_excel_writer = pd.ExcelWriter

    def flaky_excel_writer(path, *args, **kwargs):
        if "LOCKED" in str(path):
            raise PermissionError("mock: file is open in Excel")
        return real_excel_writer(path, *args, **kwargs)

    monkeypatch.setattr(fixed_to_excel_module.pd, "ExcelWriter", flaky_excel_writer)

    convert_all(ctx)  # 例外を送出せず完走すること

    assert not (output_dir / "解析結果_LOCKED.xlsx").exists()
    assert (output_dir / "解析結果_OK.xlsx").exists()


def test_restore_all_skips_locked_input_and_continues_other_files(tmp_path, monkeypatch):
    configs_dir = tmp_path / "configs"
    output_dir = tmp_path / "output"
    recreated_dir = tmp_path / "recreated"
    configs_dir.mkdir()
    output_dir.mkdir()

    _write_simple_config(configs_dir / "config.xlsx")
    for name in ("解析結果_LOCKED.xlsx", "解析結果_OK.xlsx"):
        pd.DataFrame([{"レコード種別": "データ", "区分": "2", "値": "A"}]).to_excel(
            output_dir / name, index=False
        )

    mapping_csv = tmp_path / "mapping.csv"
    columns = {"keyword": "kw", "config_name": "cfg"}
    pd.DataFrame([
        {"kw": "LOCKED", "cfg": "config.xlsx"},
        {"kw": "OK", "cfg": "config.xlsx"},
    ]).to_csv(mapping_csv, index=False, encoding=ENCODING)

    dirs = {
        "configs": str(configs_dir), "input": str(tmp_path / "input"),
        "output": str(output_dir), "recreated": str(recreated_dir),
    }
    ctx = _make_ctx(dirs, mapping_csv, columns)

    real_read_excel = pd.read_excel

    def flaky_read_excel(path, *args, **kwargs):
        if "LOCKED" in str(path):
            raise PermissionError("mock: file is open in Excel")
        return real_read_excel(path, *args, **kwargs)

    monkeypatch.setattr(excel_to_fixed_module.pd, "read_excel", flaky_read_excel)

    restore_all(ctx)  # 例外を送出せず完走すること

    assert not (recreated_dir / "RESTORED_LOCKED.txt").exists()
    assert (recreated_dir / "RESTORED_OK.txt").exists()


# ---- 複数キーワードが部分一致した場合のエラーハンドリング ----
# match_configは部分一致(in)判定のため、'SAMPLE'と'KAIIN_SAMPLE'のように複数キーワードが
# 同時に一致し得る。どちらを採用すべきか機械的に決められないため、AmbiguousKeywordMatchErrorを
# 送出してその1ファイルだけスキップし、他のファイルの処理は続行する。

def _ambiguous_mapping_df():
    return pd.DataFrame([
        {"kw": "SAMPLE", "cfg": "config.xlsx"},
        {"kw": "KAIIN_SAMPLE", "cfg": "config.xlsx"},
        {"kw": "OK", "cfg": "config.xlsx"},
    ])


def test_convert_all_skips_ambiguous_keyword_match_and_continues_other_files(tmp_path):
    configs_dir = tmp_path / "configs"
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    configs_dir.mkdir()
    input_dir.mkdir()

    _write_simple_config(configs_dir / "config.xlsx")
    for name in ("KAIIN_SAMPLE.txt", "OK.txt"):
        with open(input_dir / name, "wb") as f:
            f.write(b"2A\r\n")

    mapping_csv = tmp_path / "mapping.csv"
    columns = {"keyword": "kw", "config_name": "cfg"}
    _ambiguous_mapping_df().to_csv(mapping_csv, index=False, encoding=ENCODING)

    dirs = {
        "configs": str(configs_dir), "input": str(input_dir),
        "output": str(output_dir), "recreated": str(tmp_path / "recreated"),
    }
    ctx = _make_ctx(dirs, mapping_csv, columns)

    convert_all(ctx)  # 例外を送出せず完走すること

    assert not (output_dir / "解析結果_KAIIN_SAMPLE.xlsx").exists()
    assert (output_dir / "解析結果_OK.xlsx").exists()


def test_restore_all_skips_ambiguous_keyword_match_and_continues_other_files(tmp_path):
    configs_dir = tmp_path / "configs"
    output_dir = tmp_path / "output"
    recreated_dir = tmp_path / "recreated"
    configs_dir.mkdir()
    output_dir.mkdir()

    _write_simple_config(configs_dir / "config.xlsx")
    for name in ("解析結果_KAIIN_SAMPLE.xlsx", "解析結果_OK.xlsx"):
        pd.DataFrame([{"レコード種別": "データ", "区分": "2", "値": "A"}]).to_excel(
            output_dir / name, index=False
        )

    mapping_csv = tmp_path / "mapping.csv"
    columns = {"keyword": "kw", "config_name": "cfg"}
    _ambiguous_mapping_df().to_csv(mapping_csv, index=False, encoding=ENCODING)

    dirs = {
        "configs": str(configs_dir), "input": str(tmp_path / "input"),
        "output": str(output_dir), "recreated": str(recreated_dir),
    }
    ctx = _make_ctx(dirs, mapping_csv, columns)

    restore_all(ctx)  # 例外を送出せず完走すること

    assert not (recreated_dir / "RESTORED_KAIIN_SAMPLE.txt").exists()
    assert (recreated_dir / "RESTORED_OK.txt").exists()


# ---- 出力Excelの行グループ化・罫線 ----

def test_group_row_ranges_groups_contiguous_runs_of_length_2_or_more():
    rec_types = ["ヘッダー", "データ", "データ", "データ", "トレーラー"]
    assert _group_row_ranges(rec_types) == [(5, 7)]


def test_group_row_ranges_skips_single_row_runs():
    rec_types = ["ヘッダー", "データ", "トレーラー"]
    assert _group_row_ranges(rec_types) == []


def test_group_row_ranges_handles_multiple_groups():
    rec_types = ["データ", "データ", "ヘッダー", "データ", "データ", "データ"]
    assert _group_row_ranges(rec_types) == [(4, 5), (7, 9)]


def test_style_output_sheet_applies_border_and_groups_data_rows(tmp_path):
    df = pd.DataFrame([
        {"行番号": 1, "区分": "1", "レコード種別": "ヘッダー", "値": "H"},
        {"行番号": 2, "区分": "2", "レコード種別": "データ", "値": "D1"},
        {"行番号": 3, "区分": "2", "レコード種別": "データ", "値": "D2"},
        {"行番号": 4, "区分": "9", "レコード種別": "トレーラー", "値": "T"},
    ])
    out_path = tmp_path / "styled.xlsx"
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sheet1")
        style_output_sheet(writer.sheets["Sheet1"], df, {})

    wb = openpyxl.load_workbook(out_path)
    ws = wb.active

    assert ws.cell(row=1, column=1).border.left.style == "thin"
    assert ws.cell(row=5, column=1).border.right.style == "thin"  # データ1行目(2,3行目は開始位置/文字数の分だけシフト)

    assert ws.row_dimensions[4].outlineLevel == 0  # ヘッダー単独行
    assert ws.row_dimensions[5].outlineLevel == 1  # データ1行目
    assert ws.row_dimensions[6].outlineLevel == 1  # データ2行目
    assert ws.row_dimensions[7].outlineLevel == 0  # トレーラー単独行


# ---- 出力Excelのデータセルを文字列書式(@)に固定（先頭ゼロ消失トラップの防止） ----
# 実際の運用中に発覚した事故: 出力Excelのセルが既定の「標準」書式のままだと、"000"を"001"に
# 手で打ち直したときにExcelが数値と解釈し、先頭のゼロが消えて"1"になってしまう
# （復元後も"  1"のようにスペース埋めされ、元の桁数・ゼロ埋めのどちらも失われる）。

def test_style_output_sheet_forces_text_format_on_data_cells(tmp_path):
    df = pd.DataFrame([
        {"行番号": 1, "区分": "2", "レコード種別": "データ", "会員番号": "000"},
    ])
    out_path = tmp_path / "styled_textformat.xlsx"
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sheet1")
        style_output_sheet(writer.sheets["Sheet1"], df, {})

    wb = openpyxl.load_workbook(out_path)
    ws = wb.active

    assert ws.cell(row=4, column=4).number_format == "@"  # 会員番号セル（データはDATA_START_ROW=4から）
    assert ws.cell(row=1, column=4).number_format != "@"  # ヘッダー行は対象外


# ---- 出力Excelの帯(項目名/開始位置/文字数)をウィンドウ枠固定にする ----
# 依頼: データ行数が多いファイルをスクロールしても、項目名・開始位置・文字数の帯が
# 常に見えるようにしてほしい。

def test_style_output_sheet_freezes_panes_below_position_rows(tmp_path):
    df = pd.DataFrame([
        {"行番号": 1, "区分": "2", "レコード種別": "データ", "値": "A"},
    ])
    out_path = tmp_path / "styled_freeze.xlsx"
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sheet1")
        style_output_sheet(writer.sheets["Sheet1"], df, {})

    wb = openpyxl.load_workbook(out_path)
    ws = wb.active

    assert ws.freeze_panes == "A4"  # 1〜3行目(項目名/開始位置/文字数)を固定し、4行目以降をスクロール対象にする


# ---- 出力Excelの列グループ化（同じレコード種別が連続する項目列を折りたたむ） ----

def test_group_column_ranges_groups_contiguous_same_type_columns():
    columns = ["行番号", "区分", "レコード種別", "作成年月日", "送信元コード", "会員番号", "有効期限", "合計件数"]
    field_rules = {
        "作成年月日": ("ヘッダー", {"start": 1, "length": 8}),
        "送信元コード": ("ヘッダー", {"start": 9, "length": 10}),
        "会員番号": ("データ", {"start": 1, "length": 16}),
        "有効期限": ("データ", {"start": 17, "length": 4}),
        "合計件数": ("トレーラー", {"start": 1, "length": 10}),
    }
    # 作成年月日・送信元コード(4,5列目)がヘッダー連続、会員番号・有効期限(6,7列目)がデータ連続。
    # 合計件数(8列目)は単独なので対象外。行番号・区分・レコード種別はfield_rulesに無く区切り扱い。
    assert _group_column_ranges(columns, field_rules) == [(4, 5), (6, 7)]


def test_group_column_ranges_empty_field_rules_groups_nothing():
    columns = ["行番号", "区分", "レコード種別", "値"]
    assert _group_column_ranges(columns, {}) == []


def test_style_output_sheet_groups_columns_by_record_type(tmp_path):
    df = pd.DataFrame([
        {"行番号": 1, "区分": "1", "レコード種別": "ヘッダー", "作成年月日": "20260730",
         "送信元コード": "SRC001", "会員番号": None, "有効期限": None},
        {"行番号": 2, "区分": "2", "レコード種別": "データ", "作成年月日": None,
         "送信元コード": None, "会員番号": "1", "有効期限": "2801"},
    ])
    field_rules = {
        "作成年月日": ("ヘッダー", {"start": 1, "length": 8}),
        "送信元コード": ("ヘッダー", {"start": 9, "length": 10}),
        "会員番号": ("データ", {"start": 1, "length": 16}),
        "有効期限": ("データ", {"start": 17, "length": 4}),
    }
    out_path = tmp_path / "styled_cols.xlsx"
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sheet1")
        style_output_sheet(writer.sheets["Sheet1"], df, field_rules)

    wb = openpyxl.load_workbook(out_path)
    ws = wb.active

    for letter in ("A", "B", "C"):  # 行番号/区分/レコード種別は種別を持たず区切り扱い
        assert ws.column_dimensions[letter].outlineLevel == 0
    assert ws.column_dimensions["D"].outlineLevel == 1  # 作成年月日
    assert ws.column_dimensions["E"].outlineLevel == 1  # 送信元コード
    assert ws.column_dimensions["F"].outlineLevel == 1  # 会員番号
    assert ws.column_dimensions["G"].outlineLevel == 1  # 有効期限


# ---- データ部以外の列/行グループを初期状態で折りたたむ ----
# 依頼: 出力Excelを開いたとき、データ部だけがすぐ見える状態にしてほしい（ヘッダー/トレーラー等の
# 列・行グループは初期状態で折りたたんでおく）。

def test_style_output_sheet_collapses_non_data_column_groups_by_default(tmp_path):
    df = pd.DataFrame([
        {"行番号": 1, "区分": "1", "レコード種別": "ヘッダー", "作成年月日": "20260730",
         "送信元コード": "SRC001", "会員番号": None, "有効期限": None},
        {"行番号": 2, "区分": "2", "レコード種別": "データ", "作成年月日": None,
         "送信元コード": None, "会員番号": "1", "有効期限": "2801"},
    ])
    field_rules = {
        "作成年月日": ("ヘッダー", {"start": 1, "length": 8}),
        "送信元コード": ("ヘッダー", {"start": 9, "length": 10}),
        "会員番号": ("データ", {"start": 1, "length": 16}),
        "有効期限": ("データ", {"start": 17, "length": 4}),
    }
    out_path = tmp_path / "styled_collapse_cols.xlsx"
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sheet1")
        style_output_sheet(writer.sheets["Sheet1"], df, field_rules)

    wb = openpyxl.load_workbook(out_path)
    ws = wb.active

    assert ws.column_dimensions["D"].hidden is True  # 作成年月日(ヘッダー) は折りたたむ
    assert ws.column_dimensions["E"].hidden is True  # 送信元コード(ヘッダー) は折りたたむ
    assert not ws.column_dimensions["F"].hidden  # 会員番号(データ) は開いたまま
    assert not ws.column_dimensions["G"].hidden  # 有効期限(データ) は開いたまま


def test_style_output_sheet_collapses_non_data_row_groups_by_default(tmp_path):
    df = pd.DataFrame([
        {"行番号": 1, "区分": "1", "レコード種別": "ヘッダー", "値": "H1"},
        {"行番号": 2, "区分": "1", "レコード種別": "ヘッダー", "値": "H2"},
        {"行番号": 3, "区分": "2", "レコード種別": "データ", "値": "D1"},
        {"行番号": 4, "区分": "2", "レコード種別": "データ", "値": "D2"},
    ])
    out_path = tmp_path / "styled_collapse_rows.xlsx"
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sheet1")
        style_output_sheet(writer.sheets["Sheet1"], df, {})

    wb = openpyxl.load_workbook(out_path)
    ws = wb.active

    # データはDATA_START_ROW=4から。ヘッダー2行(4,5)は折りたたみ、データ2行(6,7)は開いたまま
    assert ws.row_dimensions[4].hidden is True
    assert ws.row_dimensions[5].hidden is True
    assert not ws.row_dimensions[6].hidden
    assert not ws.row_dimensions[7].hidden


# ---- 区切り列の挿入（列グループを隙間なしの1範囲に融合させないため） ----
# Excelのアウトライン機能は、隙間なく隣接する同レベルの列を1つの折りたたみ範囲に融合して
# しまい、種別ごとに個別開閉できない（実機のExcelで確認済み）。ブロック間に空白列を挟むことで
# 独立して開閉できるグループにする。

def _sample_df_for_separator_test():
    return pd.DataFrame([
        {"行番号": 1, "区分": "1", "レコード種別": "ヘッダー", "作成年月日": "20260730",
         "送信元コード": "SRC001", "会員番号": None, "有効期限": None, "合計件数": None},
        {"行番号": 2, "区分": "2", "レコード種別": "データ", "作成年月日": None,
         "送信元コード": None, "会員番号": "1", "有効期限": "2801", "合計件数": None},
        {"行番号": 3, "区分": "9", "レコード種別": "トレーラー", "作成年月日": None,
         "送信元コード": None, "会員番号": None, "有効期限": None, "合計件数": "1"},
    ])


def _sample_field_rules_for_separator_test():
    return {
        "作成年月日": ("ヘッダー", {"start": 1, "length": 8}),
        "送信元コード": ("ヘッダー", {"start": 9, "length": 10}),
        "会員番号": ("データ", {"start": 1, "length": 16}),
        "有効期限": ("データ", {"start": 17, "length": 4}),
        "合計件数": ("トレーラー", {"start": 1, "length": 10}),
    }


def test_insert_group_separators_no_gap_before_first_block():
    df = _sample_df_for_separator_test()
    field_rules = _sample_field_rules_for_separator_test()
    result = insert_group_separators(df, field_rules)
    columns = list(result.columns)
    # レコード種別(識別列, 種別なし) の直後がすぐ最初のブロック(作成年月日)であること
    idx = columns.index("レコード種別")
    assert columns[idx + 1] == "作成年月日"


def test_insert_group_separators_adds_gap_only_between_blocks():
    df = _sample_df_for_separator_test()
    field_rules = _sample_field_rules_for_separator_test()
    result = insert_group_separators(df, field_rules)
    columns = list(result.columns)

    seps = [c for c in columns if is_separator_column(c)]
    assert len(seps) == 2  # ヘッダー→データ、データ→トレーラーの境目にそれぞれ1つ

    idx = columns.index("送信元コード")
    assert is_separator_column(columns[idx + 1])
    idx = columns.index("有効期限")
    assert is_separator_column(columns[idx + 1])


def test_insert_group_separators_preserves_values():
    df = _sample_df_for_separator_test()
    field_rules = _sample_field_rules_for_separator_test()
    result = insert_group_separators(df, field_rules)
    assert list(result["会員番号"]) == list(df["会員番号"])
    assert list(result["合計件数"]) == list(df["合計件数"])


def test_group_column_ranges_after_separators_are_independent_blocks():
    df = _sample_df_for_separator_test()
    field_rules = _sample_field_rules_for_separator_test()
    result = insert_group_separators(df, field_rules)
    ranges = _group_column_ranges(list(result.columns), field_rules)
    # 各グループの間に区切り列(未グループ)を挟み、範囲同士が連続していないこと
    for (_, end), (next_start, _) in zip(ranges, ranges[1:]):
        assert next_start - end >= 2


def test_style_output_sheet_with_separators_creates_independent_outline_gaps(tmp_path):
    df = _sample_df_for_separator_test()
    field_rules = _sample_field_rules_for_separator_test()
    df_display = insert_group_separators(df, field_rules)

    out_path = tmp_path / "styled_sep.xlsx"
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df_display.to_excel(writer, index=False, sheet_name="Sheet1")
        style_output_sheet(writer.sheets["Sheet1"], df_display, field_rules)

    wb = openpyxl.load_workbook(out_path)
    ws = wb.active

    # D,E=ヘッダー(outline1) / F=区切り(outline0) / G,H=データ(outline1) という
    # 「1-0-1」の並びになっていれば、Excel上で独立した折りたたみグループとして認識される
    levels = [ws.column_dimensions[l].outlineLevel for l in ("D", "E", "F", "G", "H")]
    assert levels == [1, 1, 0, 1, 1]


# ---- 設定Excelが「データ」シートのみ（ヘッダー/トレーラーシートなし）の場合の往復変換 ----
# ヘッダー/トレーラーの識別レコードを持たない単純な固定長ファイル（全行データ）を扱いたい
# という要望に対し、既存コードで既に対応できていることを確認する回帰テスト。

def test_roundtrip_works_with_data_sheet_only_config(tmp_path):
    configs_dir = tmp_path / "configs"
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    recreated_dir = tmp_path / "recreated"
    for d in (configs_dir, input_dir, output_dir, recreated_dir):
        d.mkdir()

    data_sheet = pd.DataFrame([
        ["開始位置", 1, 5],
        ["文字数", 4, 13],
    ], columns=["区分", "コード", "氏名"])
    config_path = configs_dir / "config_データのみ.xlsx"
    with pd.ExcelWriter(config_path, engine="openpyxl") as writer:
        data_sheet.to_excel(writer, sheet_name="データ", index=False)  # ヘッダー/トレーラーシートなし

    lines = [
        "0001TANAKA TARO  ",
        "0002SATO   HANAKO",
        "0003SUZUKI ICHIRO",
    ]
    input_path = input_dir / "SIMPLE_DATA.txt"
    with open(input_path, "wb") as f:
        for line in lines:
            f.write(line.encode(ENCODING) + b"\r\n")

    mapping_csv = tmp_path / "mapping.csv"
    columns = {"keyword": "判定キーワード(input側)", "config_name": "設定ファイル名(configs内)"}
    pd.DataFrame([{
        "判定キーワード(input側)": "SIMPLE_DATA",
        "設定ファイル名(configs内)": "config_データのみ.xlsx",
    }]).to_csv(mapping_csv, index=False, encoding=ENCODING)

    dirs = {
        "configs": str(configs_dir), "input": str(input_dir),
        "output": str(output_dir), "recreated": str(recreated_dir),
    }
    ctx = _make_ctx(dirs, mapping_csv, columns)

    convert_all(ctx)
    # skiprows=[1, 2]: 2/3行目は開始位置・文字数の参考表示で実データではないため読み飛ばす
    df = pd.read_excel(output_dir / "解析結果_SIMPLE_DATA.xlsx", dtype=str, skiprows=[1, 2])
    assert list(df["レコード種別"]) == ["データ", "データ", "データ"]  # H/Tシートが無くても全行データ扱い

    restore_all(ctx)
    restored_path = recreated_dir / "RESTORED_SIMPLE_DATA.txt"
    with open(restored_path, "rb") as f:
        restored_lines = [l.rstrip(b"\r\n").decode(ENCODING) for l in f if l.strip()]

    assert restored_lines == lines


# ---- 第4のレコード種別「エンドレコード」（ヘッダー/データ/トレーラーとは別に、ファイル終端の
# ほぼ空白のみのレコードを扱う） ----

END_RECORD_TYPE_CODES = {"header": ["1"], "trailer": ["8"], "end": ["9"]}


def test_process_file_classifies_end_record():
    config_rules = {
        "D": [{"name": "値", "start": 0, "length": 5}],
        "H": None,
        "T": None,
        "E": [{"name": "予備", "start": 0, "length": 5}],
    }
    import tempfile
    with tempfile.TemporaryDirectory() as tmp:
        txt_path = os.path.join(tmp, "sample.txt")
        with open(txt_path, "wb") as f:
            f.write(b"9EMPTY\r\n")
        df = process_file(txt_path, config_rules, ENCODING, END_RECORD_TYPE_CODES)

    assert list(df["レコード種別"]) == ["エンドレコード"]


def test_process_file_end_code_falls_back_to_data_when_no_end_sheet():
    config_rules = {"D": [{"name": "値", "start": 0, "length": 6}], "H": None, "T": None, "E": None}
    import tempfile
    with tempfile.TemporaryDirectory() as tmp:
        txt_path = os.path.join(tmp, "sample.txt")
        with open(txt_path, "wb") as f:
            f.write(b"9EMPTY\r\n")
        df = process_file(txt_path, config_rules, ENCODING, END_RECORD_TYPE_CODES)

    assert list(df["レコード種別"]) == ["データ"]  # Eシートが無ければ従来通りデータ扱い


def test_setup_handler_generates_four_record_types_with_same_total_length(tmp_path):
    # 実際の固定長ファイルは1ファイル内で全レコード種別の総バイト数が揃っているのが一般的
    # （テキストエディタで開いたときに桁がずれて見えないように）。
    dirs = {
        "configs": str(tmp_path / "configs"), "input": str(tmp_path / "input"),
        "output": str(tmp_path / "output"), "recreated": str(tmp_path / "recreated"),
    }
    mapping_csv = tmp_path / "configs" / "mapping.csv"
    columns = {"keyword": "kw", "config_name": "cfg", "note": "note"}
    ctx = _make_ctx(dirs, mapping_csv, columns, record_type_codes=END_RECORD_TYPE_CODES)

    init_environment(ctx)

    input_path = tmp_path / "input" / "KAIIN_SAMPLE.txt"
    with open(input_path, "rb") as f:
        lines = [l.rstrip(b"\r\n") for l in f if l.strip()]

    lengths = {len(l) for l in lines}
    assert len(lengths) == 1, f"レコード種別ごとに桁数が揃っていない: {[len(l) for l in lines]}"
    assert [l[0:1].decode(ENCODING) for l in lines] == ["1", "2", "2", "2", "8", "9"]


def test_full_roundtrip_with_generated_sample_includes_end_record(tmp_path):
    dirs = {
        "configs": str(tmp_path / "configs"), "input": str(tmp_path / "input"),
        "output": str(tmp_path / "output"), "recreated": str(tmp_path / "recreated"),
    }
    mapping_csv = tmp_path / "configs" / "mapping.csv"
    columns = {
        "keyword": "判定キーワード(input側)", "config_name": "設定ファイル名(configs内)", "note": "備考",
    }
    ctx = _make_ctx(dirs, mapping_csv, columns, record_type_codes=END_RECORD_TYPE_CODES)

    init_environment(ctx)
    build_or_update_mapping(ctx)
    convert_all(ctx)

    # skiprows=[1, 2]: 2/3行目は開始位置・文字数の参考表示で実データではないため読み飛ばす
    df = pd.read_excel(tmp_path / "output" / "解析結果_KAIIN_SAMPLE.xlsx", dtype=str, skiprows=[1, 2])
    assert list(df["レコード種別"]) == ["ヘッダー", "データ", "データ", "データ", "トレーラー", "エンドレコード"]

    restore_all(ctx)

    input_path = tmp_path / "input" / "KAIIN_SAMPLE.txt"
    restored_path = tmp_path / "recreated" / "RESTORED_KAIIN_SAMPLE.txt"
    with open(input_path, "rb") as f:
        original_lines = [l.rstrip(b"\r\n") for l in f if l.strip()]
    with open(restored_path, "rb") as f:
        restored_lines = [l.rstrip(b"\r\n") for l in f if l.strip()]

    assert restored_lines == original_lines  # エンドレコード込みで完全一致


# ---- _with_padding_filler のガード（COMMON_RECORD_LENGTHがフィールド定義より短い場合） ----

def test_with_padding_filler_raises_when_total_length_too_short():
    fields = [{"name": "値", "start": 2, "length": 10}]  # 終端position=11
    with pytest.raises(ValueError):
        _with_padding_filler(fields, total_length=10)  # 11 > 10 なので不足


def test_with_padding_filler_ok_when_total_length_exactly_fits():
    fields = [{"name": "値", "start": 2, "length": 10}]  # 終端position=11
    result = _with_padding_filler(fields, total_length=11)
    filler = result[-1]
    assert filler["name"] == "予備"
    assert filler["start"] == 12
    assert filler["length"] == 0


# ---- mapping.csv の個別登録・削除（ランチャーGUIのエディタ用） ----

def _mapping_columns():
    return {"keyword": "kw", "config_name": "cfg", "note": "note"}


def test_load_mapping_returns_empty_df_when_file_missing(tmp_path):
    mapping_csv = tmp_path / "mapping.csv"
    ctx = _make_ctx({}, mapping_csv, _mapping_columns())
    df_map = load_mapping(ctx)
    assert list(df_map.columns) == ["kw", "cfg", "note"]
    assert len(df_map) == 0


def test_add_mapping_entry_creates_file_and_row(tmp_path):
    mapping_csv = tmp_path / "mapping.csv"
    ctx = _make_ctx({}, mapping_csv, _mapping_columns())

    add_mapping_entry(ctx, "KAIIN_SAMPLE", "config_サンプル.xlsx")

    df_map = load_mapping(ctx)
    assert list(df_map["kw"]) == ["KAIIN_SAMPLE"]
    assert list(df_map["cfg"]) == ["config_サンプル.xlsx"]


def test_add_mapping_entry_replaces_existing_keyword(tmp_path):
    mapping_csv = tmp_path / "mapping.csv"
    ctx = _make_ctx({}, mapping_csv, _mapping_columns())

    add_mapping_entry(ctx, "KAIIN_SAMPLE", "config_old.xlsx")
    add_mapping_entry(ctx, "KAIIN_SAMPLE", "config_new.xlsx")

    df_map = load_mapping(ctx)
    assert len(df_map) == 1  # 重複登録されず置き換わる
    assert df_map.iloc[0]["cfg"] == "config_new.xlsx"


def test_add_mapping_entry_backs_up_existing_file(tmp_path):
    mapping_csv = tmp_path / "mapping.csv"
    ctx = _make_ctx({}, mapping_csv, _mapping_columns())

    add_mapping_entry(ctx, "A", "a.xlsx")
    add_mapping_entry(ctx, "B", "b.xlsx")  # 2回目の登録でバックアップが作られる

    assert len(list(tmp_path.glob("mapping.csv.bak_*"))) >= 1


def test_remove_mapping_entry_deletes_matching_row(tmp_path):
    mapping_csv = tmp_path / "mapping.csv"
    ctx = _make_ctx({}, mapping_csv, _mapping_columns())

    add_mapping_entry(ctx, "A", "a.xlsx")
    add_mapping_entry(ctx, "B", "b.xlsx")

    removed = remove_mapping_entry(ctx, "A")

    assert removed is True
    df_map = load_mapping(ctx)
    assert list(df_map["kw"]) == ["B"]


def test_remove_mapping_entry_returns_false_when_not_found(tmp_path):
    mapping_csv = tmp_path / "mapping.csv"
    ctx = _make_ctx({}, mapping_csv, _mapping_columns())

    add_mapping_entry(ctx, "A", "a.xlsx")
    removed = remove_mapping_entry(ctx, "NOT_EXIST")

    assert removed is False
    df_map = load_mapping(ctx)
    assert list(df_map["kw"]) == ["A"]  # 何も変わらない


# ---- find_existing_config: mapping_editor_window の上書き確認ダイアログが使う問い合わせ ----

def test_find_existing_config_returns_none_when_keyword_not_registered(tmp_path):
    mapping_csv = tmp_path / "mapping.csv"
    ctx = _make_ctx({}, mapping_csv, _mapping_columns())

    assert find_existing_config(ctx, "NOT_EXIST") is None


def test_find_existing_config_returns_registered_config_name(tmp_path):
    mapping_csv = tmp_path / "mapping.csv"
    ctx = _make_ctx({}, mapping_csv, _mapping_columns())

    add_mapping_entry(ctx, "KAIIN_SAMPLE", "config_サンプル.xlsx")

    assert find_existing_config(ctx, "KAIIN_SAMPLE") == "config_サンプル.xlsx"


# ---- diff_checker: 入力(原本)と復元後の項目単位差分チェック ----
# 「Excelで特定の列だけ編集したつもりが、他の列やファイル構成まで意図せず変わっていないか」
# を確認したいという要望に対応。既存のprocess_file(固定長→DataFrame)を原本・復元後の
# 両方に適用し、項目単位で突き合わせる。

def test_restored_name_for_matches_excel_to_fixed_naming():
    assert restored_name_for("KAIIN_SAMPLE.txt") == "RESTORED_KAIIN_SAMPLE.txt"


def test_diff_rows_returns_zero_when_identical():
    df = pd.DataFrame([{"行番号": 1, "区分": "2", "店舗名": "A"}])
    assert diff_rows(df, df.copy(), logger, "SAMPLE.txt") == 0


def test_diff_rows_detects_changed_field_value():
    df_before = pd.DataFrame([{"行番号": 1, "区分": "2", "店舗名": "A"}])
    df_after = pd.DataFrame([{"行番号": 1, "区分": "2", "店舗名": "B"}])
    assert diff_rows(df_before, df_after, logger, "SAMPLE.txt") == 1


# 過去に見つかったバグ: レコード種別が異なる行では項目自体が存在せずNaNになるため、
# 素朴な != 比較だとNaN != NaNがTrueになり、無関係の欠損項目まで差分扱いになっていた
# （実データでの動作確認中、変更していないヘッダー/トレーラー行の全項目が差分として大量に出た）。

def test_diff_rows_treats_nan_as_equal_across_different_record_types():
    df_before = pd.DataFrame([
        {"行番号": 1, "レコード種別": "ヘッダー", "作成年月日": "20260730", "会員番号": pd.NA},
        {"行番号": 2, "レコード種別": "データ", "作成年月日": pd.NA, "会員番号": "1"},
    ])
    df_after = df_before.copy()
    assert diff_rows(df_before, df_after, logger, "SAMPLE.txt") == 0


def test_diff_rows_ignores_row_number_column_itself():
    # 行番号(ブランク行スキップ等でズレ得る)自体は差分としてカウントしない
    df_before = pd.DataFrame([{"行番号": 1, "区分": "2", "値": "A"}])
    df_after = pd.DataFrame([{"行番号": 2, "区分": "2", "値": "A"}])
    assert diff_rows(df_before, df_after, logger, "SAMPLE.txt") == 0


def test_diff_rows_returns_none_on_row_count_mismatch():
    df_before = pd.DataFrame([
        {"行番号": 1, "区分": "2", "値": "A"},
        {"行番号": 2, "区分": "2", "値": "B"},
    ])
    df_after = pd.DataFrame([{"行番号": 1, "区分": "2", "値": "A"}])
    assert diff_rows(df_before, df_after, logger, "SAMPLE.txt") is None


def _setup_diff_check_ctx(tmp_path):
    configs_dir = tmp_path / "configs"
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    recreated_dir = tmp_path / "recreated"
    for d in (configs_dir, input_dir, output_dir, recreated_dir):
        d.mkdir()

    data_sheet = pd.DataFrame([
        ["開始位置", 2, 18],
        ["文字数", 16, 10],
    ], columns=["区分", "会員番号", "店舗名"])
    config_path = configs_dir / "config.xlsx"
    with pd.ExcelWriter(config_path, engine="openpyxl") as writer:
        data_sheet.to_excel(writer, sheet_name="データ", index=False)

    d_line = "2" + "1000000000000001" + "TEST_SHOP "
    input_path = input_dir / "SAMPLE.txt"
    with open(input_path, "wb") as f:
        f.write(d_line.encode(ENCODING) + b"\r\n")

    mapping_csv = tmp_path / "mapping.csv"
    columns = {"keyword": "判定キーワード(input側)", "config_name": "設定ファイル名(configs内)"}
    pd.DataFrame([{
        "判定キーワード(input側)": "SAMPLE",
        "設定ファイル名(configs内)": "config.xlsx",
    }]).to_csv(mapping_csv, index=False, encoding=ENCODING)

    dirs = {
        "configs": str(configs_dir), "input": str(input_dir),
        "output": str(output_dir), "recreated": str(recreated_dir),
    }
    return _make_ctx(dirs, mapping_csv, columns)


def test_check_all_reports_no_diff_when_roundtrip_unchanged(tmp_path, caplog):
    ctx = _setup_diff_check_ctx(tmp_path)
    convert_all(ctx)
    restore_all(ctx)

    with caplog.at_level(logging.INFO):
        check_all(ctx)

    assert "SAMPLE.txt: 差分なし（構成・値とも一致）" in caplog.text


def test_check_all_detects_edited_column(tmp_path, caplog):
    ctx = _setup_diff_check_ctx(tmp_path)
    convert_all(ctx)

    # 開始位置・文字数の参考表示行(2,3行目)や書式を壊さずに1セルだけ書き換えるため、
    # DataFrame経由の全体書き直しではなくopenpyxlで対象セルだけ直接編集する。
    excel_path = tmp_path / "output" / "解析結果_SAMPLE.xlsx"
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    shop_col = header.index("店舗名") + 1
    ws.cell(row=4, column=shop_col, value="NEW_SHOP")  # 元の値"TEST_SHOP"と同じく10文字以内に収める（データはrow4から）
    wb.save(excel_path)

    restore_all(ctx)

    with caplog.at_level(logging.INFO):
        check_all(ctx)

    assert "「店舗名」: 「TEST_SHOP」→「NEW_SHOP」" in caplog.text
    assert "SAMPLE.txt: 差分1件検出" in caplog.text


def test_check_all_skips_when_restored_file_missing(tmp_path, caplog):
    ctx = _setup_diff_check_ctx(tmp_path)
    convert_all(ctx)  # restore_allは実行しない -> recreated_inputが空

    with caplog.at_level(logging.INFO):
        check_all(ctx)

    assert "復元後ファイル未検出" in caplog.text


def test_check_all_reports_error_and_continues_on_ambiguous_keyword_match(tmp_path, caplog):
    ctx = _setup_diff_check_ctx(tmp_path)
    convert_all(ctx)
    restore_all(ctx)

    # SAMPLE.txtに部分一致するキーワードをもう1つ追加登録し、あいまいな状態にする
    add_mapping_entry(ctx, "AMPLE", "config.xlsx")

    with caplog.at_level(logging.INFO):
        check_all(ctx)

    assert "複数のキーワードが部分一致しました" in caplog.text


# ---- #169 変換・復元時の警告強化 ----
# レビューで判明: pad_value_to_bytes はフィールドのバイト長を超える値を無言で切り捨てる。
# cp932の2バイト文字の途中で切ると不正バイトが残る。切り捨ては警告し、文字境界で切る。

def test_pad_value_to_bytes_truncates_multibyte_on_char_boundary():
    # "あいう" = cp932で6バイト。5バイトに切ると3文字目の途中で切れる。
    result = pad_value_to_bytes("あいう", 5, ENCODING)
    assert len(result) == 5
    result.decode(ENCODING)  # 不正バイトが残っていれば UnicodeDecodeError
    assert result == "あい".encode(ENCODING) + b" "  # 半端な1バイトを落としスペース埋め


def test_pad_value_to_bytes_calls_on_truncate_callback():
    calls = []
    pad_value_to_bytes("ABCDEF", 3, ENCODING,
                       on_truncate=lambda *a: calls.append(a), field_name="コード")
    assert calls == [("コード", 6, 3)]


def test_pad_value_to_bytes_no_truncate_callback_when_value_fits():
    calls = []
    pad_value_to_bytes("AB", 5, ENCODING, on_truncate=lambda *a: calls.append(a))
    assert calls == []


def test_build_fixed_line_records_truncation_in_diagnostics():
    rules = [{"name": "会員番号", "start": 0, "length": 4}]
    diag = []
    build_fixed_line({"会員番号": "1234567890"}, rules, ENCODING, diagnostics=diag)
    assert diag == [{"field": "会員番号", "actual": 10, "limit": 4}]


def test_process_file_records_record_length_mismatch_in_diagnostics(tmp_path):
    config_path = tmp_path / "config.xlsx"
    data_sheet = pd.DataFrame([["開始位置", 2], ["文字数", 5]], columns=["区分", "値"])
    with pd.ExcelWriter(config_path, engine="openpyxl") as writer:
        data_sheet.to_excel(writer, sheet_name="データ", index=False)
    from src.utils.fixed_format import load_config_rules
    rules = load_config_rules(config_path)  # 想定レコード長 = 1 + 5 = 6

    txt_path = tmp_path / "sample.txt"
    with open(txt_path, "wb") as f:
        f.write(b"2ABCDE\r\n")   # 6バイト(一致)
        f.write(b"2ABC\r\n")     # 4バイト(不足)

    diag = []
    process_file(txt_path, rules, ENCODING, RECORD_TYPE_CODES, diagnostics=diag)
    assert diag == [{"line": 2, "rec_type": "データ", "expected": 6, "actual": 4}]


def _setup_restore_ctx_with_one_row(tmp_path):
    """convert_all で position 行付きの正しい出力Excelを1行だけ作った状態の ctx を返す。"""
    configs_dir = tmp_path / "configs"
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    recreated_dir = tmp_path / "recreated"
    for d in (configs_dir, input_dir, output_dir, recreated_dir):
        d.mkdir()

    _write_simple_config(configs_dir / "config.xlsx")  # 「値」start1 len1
    with open(input_dir / "SAMPLE.txt", "wb") as f:
        f.write(b"2A\r\n")

    mapping_csv = tmp_path / "mapping.csv"
    columns = {"keyword": "kw", "config_name": "cfg"}
    pd.DataFrame([{"kw": "SAMPLE", "cfg": "config.xlsx"}]).to_csv(
        mapping_csv, index=False, encoding=ENCODING
    )
    dirs = {
        "configs": str(configs_dir), "input": str(input_dir),
        "output": str(output_dir), "recreated": str(recreated_dir),
    }
    ctx = _make_ctx(dirs, mapping_csv, columns)
    convert_all(ctx)
    return ctx, output_dir / "解析結果_SAMPLE.xlsx"


def test_restore_all_warns_on_truncation_and_summarizes(tmp_path, caplog):
    ctx, excel_path = _setup_restore_ctx_with_one_row(tmp_path)

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    ws.cell(row=4, column=header.index("値") + 1, value="ABC")  # 1バイト枠に3バイト（データはrow4から）
    wb.save(excel_path)

    with caplog.at_level(logging.INFO):
        restore_all(ctx)

    assert "「値」: 3バイト → 1バイトに切り捨て" in caplog.text
    assert "切り捨て 1件" in caplog.text


def test_restore_all_reports_no_truncation_when_values_fit(tmp_path, caplog):
    ctx, _ = _setup_restore_ctx_with_one_row(tmp_path)

    with caplog.at_level(logging.INFO):
        restore_all(ctx)

    assert "切り捨てなし" in caplog.text


def test_convert_all_warns_on_record_length_mismatch(tmp_path, caplog):
    configs_dir = tmp_path / "configs"
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    for d in (configs_dir, input_dir, output_dir):
        d.mkdir()

    data_sheet = pd.DataFrame([["開始位置", 2], ["文字数", 5]], columns=["区分", "値"])
    with pd.ExcelWriter(configs_dir / "config.xlsx", engine="openpyxl") as writer:
        data_sheet.to_excel(writer, sheet_name="データ", index=False)  # 想定長6

    with open(input_dir / "SAMPLE.txt", "wb") as f:
        f.write(b"2ABCDE\r\n")  # 6バイト(一致)
        f.write(b"2XYZ\r\n")    # 4バイト(不一致)

    mapping_csv = tmp_path / "mapping.csv"
    columns = {"keyword": "kw", "config_name": "cfg"}
    pd.DataFrame([{"kw": "SAMPLE", "cfg": "config.xlsx"}]).to_csv(
        mapping_csv, index=False, encoding=ENCODING
    )
    dirs = {
        "configs": str(configs_dir), "input": str(input_dir),
        "output": str(output_dir), "recreated": str(tmp_path / "recreated"),
    }
    ctx = _make_ctx(dirs, mapping_csv, columns)

    with caplog.at_level(logging.INFO):
        convert_all(ctx)

    assert "レコード長 4バイト （設定の想定は 6バイト）" in caplog.text
    assert "レコード長不一致 1件" in caplog.text


# ---- #167 設定Excelのバリデーション ----

def test_validate_config_rules_detects_overlap():
    rules = {
        "D": [
            {"name": "会員番号", "start": 0, "length": 10},
            {"name": "氏名", "start": 8, "length": 20},  # 会員番号と2バイト重なる
        ],
        "H": None, "T": None, "E": None,
    }
    msgs = validate_config_rules(rules)
    assert any("氏名" in m and "重なって" in m and "2 バイト" in m for m in msgs)


def test_validate_config_rules_detects_gap():
    rules = {
        "D": [
            {"name": "会員番号", "start": 0, "length": 10},
            {"name": "氏名", "start": 15, "length": 20},  # 10〜15に隙間
        ],
        "H": None, "T": None, "E": None,
    }
    msgs = validate_config_rules(rules)
    assert any("未定義のバイト" in m and "氏名" in m for m in msgs)


def test_validate_config_rules_detects_non_positive_length():
    rules = {"D": [{"name": "予備", "start": 5, "length": 0}], "H": None, "T": None, "E": None}
    msgs = validate_config_rules(rules)
    assert any("文字数が 0" in m for m in msgs)


def test_validate_config_rules_detects_record_length_mismatch_between_types():
    rules = {
        "D": [{"name": "会員番号", "start": 0, "length": 16}],   # 終端16
        "H": [{"name": "作成日", "start": 0, "length": 8}],       # 終端8
        "T": None, "E": None,
    }
    msgs = validate_config_rules(rules)
    assert any("終端位置" in m and "揃っていません" in m for m in msgs)


def test_validate_config_rules_clean_config_has_no_warnings():
    rules = {
        "D": [
            {"name": "会員番号", "start": 0, "length": 10},
            {"name": "氏名", "start": 10, "length": 20},
        ],
        "H": None, "T": None, "E": None,
    }
    assert validate_config_rules(rules) == []


def test_parse_sheet_rules_skips_non_numeric_cell_and_records_it(tmp_path):
    config_path = tmp_path / "config.xlsx"
    data_sheet = pd.DataFrame([
        ["開始位置", 1, "8桁"],   # 2列目の開始位置が数値でない
        ["文字数", 4, 8],
    ], columns=["区分", "コード", "名前"])
    with pd.ExcelWriter(config_path, engine="openpyxl") as writer:
        data_sheet.to_excel(writer, sheet_name="データ", index=False)

    from src.utils.fixed_format import parse_sheet_rules
    xl = pd.ExcelFile(config_path)
    invalid = []
    fields = parse_sheet_rules(xl, "データ", invalid)
    assert [f["name"] for f in fields] == ["コード"]  # 「名前」は飛ばされる
    assert invalid == ['データ「名前」']


def test_load_config_rules_logs_config_warnings(tmp_path, caplog):
    config_path = tmp_path / "config.xlsx"
    data_sheet = pd.DataFrame([
        ["開始位置", 1, 8],   # コード(1-4), 氏名(8-...) → 4〜8に隙間
        ["文字数", 4, 10],
    ], columns=["区分", "コード", "氏名"])
    with pd.ExcelWriter(config_path, engine="openpyxl") as writer:
        data_sheet.to_excel(writer, sheet_name="データ", index=False)

    with caplog.at_level(logging.WARNING):
        load_config_rules(config_path, logger=logger)

    assert "未定義のバイト" in caplog.text


# ---- #168 復元 → 差分チェックの自動連結 ----

def test_restore_and_check_runs_both_and_reports_no_diff(tmp_path, caplog):
    ctx = _setup_diff_check_ctx(tmp_path)
    convert_all(ctx)

    with caplog.at_level(logging.INFO):
        restore_and_check(ctx)

    # 復元が走った証跡
    assert (tmp_path / "recreated" / "RESTORED_SAMPLE.txt").exists()
    # そのまま差分チェックも走った証跡
    assert "SAMPLE.txt: 差分なし（構成・値とも一致）" in caplog.text


def test_restore_and_check_detects_edited_column_in_one_action(tmp_path, caplog):
    ctx = _setup_diff_check_ctx(tmp_path)
    convert_all(ctx)

    excel_path = tmp_path / "output" / "解析結果_SAMPLE.xlsx"
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    ws.cell(row=4, column=header.index("店舗名") + 1, value="NEW_SHOP")
    wb.save(excel_path)

    with caplog.at_level(logging.INFO):
        restore_and_check(ctx)

    assert "「店舗名」: 「TEST_SHOP」→「NEW_SHOP」" in caplog.text
    assert "SAMPLE.txt: 差分1件検出" in caplog.text


# ---- #170 GUI/往復まわりの掃除 ----

def test_create_context_raises_clear_error_on_missing_keys(monkeypatch):
    class _StubCM:
        def load_config(self):
            return {"app_name": "x"}  # encoding 等の必須キーが無い

    monkeypatch.setattr(app_context_module, "ConfigManager", _StubCM)
    with pytest.raises(RuntimeError) as exc:
        create_context(logger)
    assert "encoding" in str(exc.value) and "config/main.yaml" in str(exc.value)


def test_create_context_builds_when_all_keys_present(monkeypatch):
    cfg = {
        "encoding": "cp932", "dirs": {"input": "x"}, "mapping_csv": "m.csv",
        "mapping_columns": {"keyword": "k"}, "record_type_codes": {"header": ["1"]},
    }

    class _StubCM:
        def load_config(self):
            return cfg

    monkeypatch.setattr(app_context_module, "ConfigManager", _StubCM)
    ctx = create_context(logger)
    assert ctx.encoding == "cp932"
    assert ctx.app_name == "34_Fixed2Excel"  # 既定値


def test_read_mapping_csv_keeps_numeric_keyword_as_string_and_blank_as_empty(tmp_path):
    path = tmp_path / "mapping.csv"
    pd.DataFrame([
        {"判定キーワード(input側)": "12345", "設定ファイル名(configs内)": "c.xlsx", "備考": ""},
    ]).to_csv(path, index=False, encoding=ENCODING)

    df = read_mapping_csv(path, ENCODING)
    assert df.iloc[0]["判定キーワード(input側)"] == "12345"  # "12345.0" や 12345 にならない
    assert df.iloc[0]["備考"] == ""  # NaN でも "nan" でもない


def test_load_mapping_blank_note_is_empty_string_not_nan(tmp_path):
    mapping_csv = tmp_path / "mapping.csv"
    columns = {"keyword": "kw", "config_name": "cfg", "note": "note"}
    pd.DataFrame([{"kw": "SAMPLE", "cfg": "c.xlsx", "note": ""}]).to_csv(
        mapping_csv, index=False, encoding=ENCODING
    )
    ctx = _make_ctx({}, mapping_csv, columns)
    df = load_mapping(ctx)
    assert df.iloc[0]["note"] == ""
    assert not pd.isna(df.iloc[0]["note"])


@pytest.mark.parametrize("name", ["SAMPLE.txt", "解析結果_SAMPLE.xlsx"])
def test_restored_txt_name_resolves_same_base_from_input_or_excel_name(name):
    assert restored_txt_name(name) == "RESTORED_SAMPLE.txt"


def test_analysis_excel_name_from_input():
    assert analysis_excel_name("KAIIN_SAMPLE.txt") == "解析結果_KAIIN_SAMPLE.xlsx"


def test_restored_txt_name_handles_base_containing_xlsx_like_substring():
    # 旧実装の .replace(".xlsx", "") はファイル名途中の ".xlsx" も消してしまっていた
    assert restored_txt_name("解析結果_a.xlsx.b.xlsx") == "RESTORED_a.xlsx.b.txt"


# ---- #171 「新しいファイルに対応」ウィザードのロジック ----

from src.handlers import config_wizard


def test_parse_field_spec_tab_separated():
    text = "会員番号\t2\t16\n有効期限\t18\t4"
    fields, errors = config_wizard.parse_field_spec(text)
    assert errors == []
    assert fields == [
        {"name": "会員番号", "start": 2, "length": 16},
        {"name": "有効期限", "start": 18, "length": 4},
    ]


def test_parse_field_spec_accepts_comma_and_whitespace_and_skips_comments():
    text = "# コメント行\n\nコード,1,4\n氏名  5  13\n"
    fields, errors = config_wizard.parse_field_spec(text)
    assert errors == []
    assert [f["name"] for f in fields] == ["コード", "氏名"]
    assert fields[1] == {"name": "氏名", "start": 5, "length": 13}


def test_parse_field_spec_reports_bad_lines():
    text = "会員番号\t2\t16\nこわれた行\nコード\tあ\t4"
    fields, errors = config_wizard.parse_field_spec(text)
    assert [f["name"] for f in fields] == ["会員番号"]
    assert len(errors) == 2
    assert "2行目" in errors[0] and "3行目" in errors[1]


def test_parse_field_spec_empty_is_error():
    fields, errors = config_wizard.parse_field_spec("   \n # only comment\n")
    assert fields == []
    assert errors and "空" in errors[0]


def test_build_config_excel_roundtrips_through_load_config_rules(tmp_path):
    path = tmp_path / "config_新規.xlsx"
    config_wizard.build_config_excel(path, {
        "D": [{"name": "会員番号", "start": 2, "length": 16}],
        "H": [{"name": "作成年月日", "start": 2, "length": 8}],
    })
    rules = load_config_rules(path)
    assert rules["D"] == [{"name": "会員番号", "start": 1, "length": 16}]  # 1始まり→0始まり
    assert rules["H"] == [{"name": "作成年月日", "start": 1, "length": 8}]
    assert rules["T"] is None  # 貼り付けなかった種別のシートは作られない


def test_build_config_excel_requires_data_records(tmp_path):
    with pytest.raises(ValueError):
        config_wizard.build_config_excel(tmp_path / "x.xlsx", {"H": [{"name": "a", "start": 1, "length": 1}]})


def test_validation_warnings_converts_1based_and_detects_overlap():
    warns = config_wizard.validation_warnings({
        "D": [
            {"name": "会員番号", "start": 1, "length": 10},
            {"name": "氏名", "start": 9, "length": 20},  # 1始まりで2バイト重なる
        ],
    })
    assert any("重なって" in w for w in warns)


def test_wizard_generated_config_works_end_to_end(tmp_path):
    """ウィザードで作った設定Excelで実際に変換できることを確認する。"""
    configs_dir = tmp_path / "configs"
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    for d in (configs_dir, input_dir, output_dir):
        d.mkdir()

    config_wizard.build_config_excel(configs_dir / "config_NEWTYPE.xlsx", {
        "D": [
            {"name": "コード", "start": 1, "length": 4},
            {"name": "氏名", "start": 5, "length": 13},
        ],
    })
    with open(input_dir / "NEWTYPE.txt", "wb") as f:
        f.write("0001TANAKA TARO  ".encode(ENCODING) + b"\r\n")

    mapping_csv = tmp_path / "mapping.csv"
    columns = {"keyword": "判定キーワード(input側)", "config_name": "設定ファイル名(configs内)"}
    pd.DataFrame([{
        "判定キーワード(input側)": "NEWTYPE",
        "設定ファイル名(configs内)": "config_NEWTYPE.xlsx",
    }]).to_csv(mapping_csv, index=False, encoding=ENCODING)

    dirs = {
        "configs": str(configs_dir), "input": str(input_dir),
        "output": str(output_dir), "recreated": str(tmp_path / "recreated"),
    }
    ctx = _make_ctx(dirs, mapping_csv, columns)
    convert_all(ctx)

    df = pd.read_excel(output_dir / "解析結果_NEWTYPE.xlsx", dtype=str, skiprows=[1, 2])
    assert list(df["コード"]) == ["0001"]
    assert list(df["氏名"]) == ["TANAKA TARO"]


def test_config_wizard_window_on_create_generates_config_and_mapping_and_triggers_convert(tmp_path, monkeypatch):
    tk = pytest.importorskip("tkinter")
    from src.config_wizard_window import ConfigWizardWindow

    configs_dir = tmp_path / "configs"
    input_dir = tmp_path / "input"
    for d in (configs_dir, input_dir):
        d.mkdir()
    (input_dir / "NEWTYPE.txt").write_bytes(b"0001AAA\r\n")

    mapping_csv = tmp_path / "mapping.csv"
    columns = {"keyword": "kw", "config_name": "cfg", "note": "note"}
    ctx = _make_ctx(
        {"configs": str(configs_dir), "input": str(input_dir),
         "output": str(tmp_path / "out"), "recreated": str(tmp_path / "rec")},
        mapping_csv, columns,
    )

    try:
        root = tk.Tk()
    except tk.TclError:
        pytest.skip("Tk 表示環境なし")
    root.withdraw()
    root.is_running = False
    root.logger = logger
    triggered = []
    root._run_action = lambda key: triggered.append(key)

    wiz = ConfigWizardWindow(root, ctx)
    wiz.input_var.set("NEWTYPE.txt")
    wiz._on_input_selected()
    wiz._paste_boxes["D"].insert("1.0", "コード\t1\t4\n名前\t5\t3")
    wiz._on_create()
    root.update()

    assert (configs_dir / "config_NEWTYPE.xlsx").exists()
    df_map = pd.read_csv(mapping_csv, encoding=ENCODING)
    assert df_map.iloc[0]["kw"] == "NEWTYPE"
    assert df_map.iloc[0]["cfg"] == "config_NEWTYPE.xlsx"
    assert triggered == ["to_excel"]

    root.destroy()


# ---- メイン画面レイアウトの見直し（作業順の3ブロック構成） ----

def test_main_window_has_workflow_buttons_and_folder_menu(monkeypatch):
    tk = pytest.importorskip("tkinter")
    from src import gui as gui_module

    monkeypatch.setattr(gui_module, "create_context", lambda logger: _make_ctx(
        {"configs": "x", "input": "x", "output": "x", "recreated": "x"},
        "m.csv", {"keyword": "k", "config_name": "c", "note": "n"},
    ))
    try:
        app = gui_module.Fixed2ExcelApp()
    except tk.TclError:
        pytest.skip("Tk 表示環境なし")
    app.withdraw()
    app.update_idletasks()

    # ふだんの作業＋初回＋その他の6アクションが self.buttons に揃っている
    assert set(app.buttons) == {
        "to_excel", "to_fixed", "init", "config_wizard", "diff_check", "edit_mapping",
    }

    # フォルダを開くはドロップダウン（Menubutton）に集約され、実行中も押せる
    menubuttons = []
    def walk(w):
        for c in w.winfo_children():
            if c.winfo_class() == "TMenubutton":
                menubuttons.append(c)
            walk(c)
    walk(app)
    assert len(menubuttons) == 1
    menu = app.nametowidget(menubuttons[0]["menu"])
    assert menu.index("end") == 3  # 4項目（設定/入力/出力/復元後）

    app._set_running(True)
    assert all(str(b["state"]) == "disabled" for b in app.buttons.values())
    assert str(menubuttons[0]["state"]) == "normal"  # フォルダは処理中でも開ける
    app._set_running(False)

    app.destroy()
