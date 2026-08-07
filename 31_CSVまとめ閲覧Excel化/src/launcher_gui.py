# 31_CSVまとめ閲覧Excel化 ランチャー GUI
# data/input/ フォルダのCSV/TXTをまとめて data/output/view.xlsx に変換する csv_table_viewer.py を、
# 「何を準備し、何が出力されるか」を画面上で示しながら実行できるようにする単一ツール向けランチャー。
# Python 標準ライブラリ（Tkinter）のみで動作する。
# 33_テキスト・CSV前処理サポート/launcher_gui.py の設計（入力プレビュー・リアルタイムログ・
# 出力サマリー＋詳細ウィンドウ）を単一ツール向けに簡略化して踏襲している。
# Optional: pip install sv-ttk      → Windows 11 スタイルのテーマが有効になる
# Optional: pip install pywinstyles → タイトルバーもダークテーマに揃う（sv-ttk併用時）

import os
import sys
import queue
import zipfile
import threading
import subprocess
import tkinter as tk
from tkinter import ttk, messagebox
from tkinter.scrolledtext import ScrolledText
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

import csv_table_viewer as ctv

try:
    import sv_ttk as _sv_ttk
    _SV_TTK = True
except ImportError:
    _SV_TTK = False

try:
    import pywinstyles as _pywinstyles
    _PYWINSTYLES = True
except ImportError:
    _PYWINSTYLES = False

# src/ の1つ上（プロジェクトルート）を基準にする。cwdに依存しないため、
# run.batから起動しても直接 `python src/launcher_gui.py` を実行しても同じ場所を指す。
BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / "data"
CSV_DIR = DATA_DIR / "input"
OUTPUT_PATH = DATA_DIR / "output" / "view.xlsx"
LOGS_DIR = DATA_DIR / "logs"
LOG_CLEANUP_DAYS = 30

ABOUT_TEXT = (
    "準備するもの： data/input/ フォルダに UTF-8 / Shift-JIS の .csv または .txt ファイルを置く\n"
    "　　　　　　（1ファイル＝1シート。区切り文字はカンマ／タブをファイルごとに自動判定）\n"
    "出力されるもの： data/output/view.xlsx （INDEXシート＋ファイルごとのシート。ヘッダ固定・オートフィルタ付き）"
)

UI_FONT_FAMILY = "Yu Gothic UI"
UI_FONT = (UI_FONT_FAMILY, 9)
UI_FONT_BOLD = (UI_FONT_FAMILY, 9, "bold")
HEADER_FONT = (UI_FONT_FAMILY, 11, "bold")
LOG_FONT = ("ＭＳ ゴシック", 10)

# Treeview選択行の強調色。sv_ttkダークテーマの既定（#1c1c1c→#292929）はコントラスト比1.17と
# ほぼ同じ暗さで選択が見分けづらいため明示指定する（このコントラスト比3.75、白文字4.55でWCAG AA相当）。
TREE_SELECT_BG = "#2f6fed"
TREE_SELECT_FG = "#ffffff"

BTN_PRIMARY = "Primary.Accent.TButton"
BTN_SECONDARY = "Secondary.TButton"
BTN_TERTIARY = "Tertiary.Toolbutton"
_BUTTON_SPECS = {
    BTN_PRIMARY:   {"font": (UI_FONT_FAMILY, 14, "bold"), "height": 40, "hpad": 20},
    BTN_SECONDARY: {"font": (UI_FONT_FAMILY, 11),         "height": 32, "hpad": 12},
    BTN_TERTIARY:  {"font": (UI_FONT_FAMILY, 10),         "height": 26, "hpad": 6},
}

_LOG_TAGS = {
    "header": {"foreground": "#4a9eff"},
    "debug": {"foreground": "#888888"},
    "pass": {"foreground": "#4ec94e"},
    "fail": {"foreground": "#ff5555", "font": LOG_FONT + ("bold",)},
    "error": {"foreground": "#ff7777"},
    "warn": {"foreground": "#ffb347"},
}

_SUMMARY_COLORS = {
    "idle": "#888888",
    "running": "#4a9eff",
    "done": "#4ec94e",
    "empty": "#ffb347",
}

_SENTINEL_RUN_DONE = object()
_LOG_NAME_RE_SUFFIX = "_"  # ログファイル名は "run_YYYYMMDD_HHMMSS.log"

# 1回のflushで無制限に処理すると、大量ログ発生時にText部品への大量insert/see呼び出しで
# メインスレッドが長時間ブロックされ、Windowsに「応答なし」と判定される
# （32_フォルダ構造Excel出力で実測: 3万行を1回で処理すると60秒以上。launcher-gui-designスキル参照）。
# 表示行数の間引きは行わない（_autosave_logがウィジェットの表示内容をそのままログファイルへ
# 保存する設計のため、間引くと保存されるログの前半が欠けてしまう）。
MAX_LOG_LINES_PER_FLUSH = 500


# ------------------------------------------------------------------
# 汎用ヘルパー
# ------------------------------------------------------------------

def _format_filesize(size_bytes):
    if size_bytes < 1024:
        return f"{size_bytes} B"
    if size_bytes < 1024 * 1024:
        return f"{size_bytes / 1024:.1f} KB"
    return f"{size_bytes / 1024 / 1024:.1f} MB"


def _read_head_lines(path, max_lines=10, max_bytes=64 * 1024):
    """CSVの先頭 max_lines 行を (行リスト, エンコード表示名) で返す。"""
    with open(path, "rb") as f:
        raw = f.read(max_bytes)
    has_bom = raw.startswith(b"\xef\xbb\xbf")
    for enc, label in (("utf-8-sig", "UTF-8(BOM)" if has_bom else "UTF-8"),
                       ("cp932", "Shift-JIS")):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        text, label = raw.decode("utf-8", errors="replace"), "不明（変換時にエラーになる可能性）"
    lines = text.splitlines()
    if len(raw) == max_bytes and lines:
        lines = lines[:-1]  # 末尾はバイト境界で切れている可能性があるため捨てる
    return lines[:max_lines], label


def _style_titlebar(window):
    if not (_SV_TTK and _PYWINSTYLES):
        return
    try:
        version = sys.getwindowsversion()
        if version.major == 10 and version.build >= 22000:
            _pywinstyles.change_header_color(window, "#1c1c1c")
        elif version.major == 10:
            _pywinstyles.apply_style(window, "dark")
            window.wm_attributes("-alpha", 0.99)
            window.wm_attributes("-alpha", 1)
    except Exception:
        pass


def _open_in_explorer(path):
    try:
        os.startfile(str(path))
    except OSError as exc:
        messagebox.showerror("エラー", f"開けませんでした:\n{path}\n\n{exc}")


def cleanup_old_logs():
    """logs/ 内の run_YYYYMMDD_HHMMSS.log のうち古いものを zip 化して削除する。"""
    if not LOGS_DIR.is_dir():
        return
    archive_dir = LOGS_DIR / "archive"
    cutoff = datetime.now() - timedelta(days=LOG_CLEANUP_DAYS)
    for log_file in LOGS_DIR.glob("run_*.log"):
        try:
            mtime = datetime.fromtimestamp(log_file.stat().st_mtime)
            if mtime >= cutoff:
                continue
            archive_dir.mkdir(exist_ok=True)
            zip_path = archive_dir / f"{log_file.name}.zip"
            with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
                zf.write(log_file, arcname=log_file.name)
            log_file.unlink()
        except OSError:
            continue


def read_output_summary(path):
    """view.xlsx の INDEXシートから (シート名, 行数) の一覧を読む。読めなければ None。"""
    if not path.exists():
        return None
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            if "INDEX" not in wb.sheetnames:
                return None
            ws = wb["INDEX"]
            rows = []
            for row in ws.iter_rows(min_row=6, max_col=2, values_only=True):
                name, count = row
                if not name:
                    break
                rows.append((str(name), int(count) if count is not None else 0))
            return rows
        finally:
            wb.close()
    except Exception:
        return None


# ------------------------------------------------------------------
# CSV冒頭プレビューサブウィンドウ
# ------------------------------------------------------------------

class InputPreviewWindow(tk.Toplevel):
    """選択中のCSVの先頭数行と行数を表示する。実行前に中身を確認したいときに使う。"""

    def __init__(self, master, path):
        super().__init__(master)
        self.title(f"冒頭プレビュー — {path.name}")
        self.geometry("820x320")
        self.minsize(560, 220)
        _style_titlebar(self)

        try:
            lines, enc_label = _read_head_lines(path)
        except OSError as exc:
            messagebox.showerror("エラー", f"読み込めませんでした:\n{path}\n\n{exc}", parent=master)
            self.destroy()
            return

        delimiter = ctv.detect_delimiter("\n".join(lines)) if lines else ","
        delim_label = {",": "カンマ", "\t": "タブ"}.get(delimiter, repr(delimiter))

        frame = ttk.Frame(self, padding=8)
        frame.pack(fill="both", expand=True)
        frame.rowconfigure(1, weight=1)
        frame.columnconfigure(0, weight=1)

        ttk.Label(frame, text=f"エンコード: {enc_label}　／　区切り文字: {delim_label}　（先頭 {len(lines)} 行）").grid(
            row=0, column=0, sticky="w", pady=(0, 4))

        text = tk.Text(frame, font=LOG_FONT, wrap="none", height=12)
        text.grid(row=1, column=0, sticky="nsew")
        sb_y = ttk.Scrollbar(frame, command=text.yview)
        sb_y.grid(row=1, column=1, sticky="ns")
        sb_x = ttk.Scrollbar(frame, orient="horizontal", command=text.xview)
        sb_x.grid(row=2, column=0, sticky="ew")
        text.configure(yscrollcommand=sb_y.set, xscrollcommand=sb_x.set)
        for i, line in enumerate(lines, start=1):
            text.insert("end", f"{i:>3}: {line}\n")
        text.configure(state="disabled")

        if not enc_label.startswith(("UTF-8", "Shift-JIS")):
            hint = ttk.Label(
                frame, foreground="#ff7777",
                text="※ 文字コードを判別できませんでした。実行時にエラーになる可能性があります。UTF-8で保存し直すことをおすすめします。")
        else:
            hint = ttk.Label(
                frame, foreground="#888888",
                text="※ 1行目が列名（ヘッダー）として出力されます。")
        hint.grid(row=3, column=0, columnspan=2, sticky="w", pady=(4, 0))
        ttk.Button(frame, text="閉じる", style=BTN_TERTIARY, command=self.destroy).grid(
            row=4, column=0, columnspan=2, sticky="e", pady=(8, 0))


# ------------------------------------------------------------------
# 出力シート詳細サブウィンドウ
# ------------------------------------------------------------------

class OutputDetailWindow(tk.Toplevel):
    """view.xlsx の各シート（CSV）の行数一覧。"""

    def __init__(self, master, sheet_info, output_path):
        super().__init__(master)
        self.title(f"出力シートの詳細（{len(sheet_info)} シート）")
        self.geometry("640x420")
        self.minsize(480, 280)
        _style_titlebar(self)

        frame = ttk.Frame(self, padding=8)
        frame.pack(fill="both", expand=True)
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)

        columns = ("name", "rows")
        tree = ttk.Treeview(frame, columns=columns)
        tree.heading("#0", text="No.")
        tree.heading("name", text="シート名（元CSV名）")
        tree.heading("rows", text="行数")
        tree.column("#0", width=50, anchor="e", stretch=False)
        tree.column("name", width=380)
        tree.column("rows", width=100, anchor="e", stretch=False)
        tree.grid(row=0, column=0, sticky="nsew")
        sb = ttk.Scrollbar(frame, command=tree.yview)
        sb.grid(row=0, column=1, sticky="ns")
        tree.configure(yscrollcommand=sb.set)

        for no, (name, rows) in enumerate(sheet_info, start=1):
            tree.insert("", "end", text=str(no), values=(name, f"{rows:,}"))

        btns = ttk.Frame(frame)
        btns.grid(row=1, column=0, columnspan=2, sticky="e", pady=(8, 0))
        ttk.Button(btns, text="view.xlsxを開く", style=BTN_SECONDARY,
                   command=lambda: _open_in_explorer(output_path)).pack(side="left", padx=(0, 8))
        ttk.Button(btns, text="閉じる", style=BTN_TERTIARY,
                   command=self.destroy).pack(side="left")


# ------------------------------------------------------------------
# メインウィンドウ
# ------------------------------------------------------------------

class LauncherApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("CSVまとめ閲覧Excel化 ランチャー")
        self.geometry("980x640")
        self.minsize(760, 520)

        self.log_queue = queue.Queue()
        self.proc = None
        self.is_running = False
        self._last_sheet_info = None
        self._browsed_csv = {}  # Treeview item id -> Path

        self._build_ui()

        if _SV_TTK:
            _sv_ttk.set_theme("dark")
        self._setup_style()

        self.refresh_csv_list()
        self._load_existing_summary()
        self.after(100, self._drain_log_queue)
        self.after(300, lambda: threading.Thread(target=cleanup_old_logs, daemon=True).start())
        self.protocol("WM_DELETE_WINDOW", self._on_close)
        _style_titlebar(self)

    # -------------------------------------------------- UI 構築

    def _setup_style(self):
        style = ttk.Style(self)
        for name in ("TLabel", "TButton", "TCheckbutton", "TEntry", "Treeview"):
            style.configure(name, font=UI_FONT)
        style.configure("TLabelframe.Label", font=UI_FONT_BOLD)
        style.configure("Treeview.Heading", font=UI_FONT_BOLD)
        style.configure("Treeview", rowheight=24)
        style.map("Treeview",
                  background=[("selected", TREE_SELECT_BG)],
                  foreground=[("selected", TREE_SELECT_FG)])

        for style_name, spec in _BUTTON_SPECS.items():
            style.configure(style_name, font=spec["font"], padding=(spec["hpad"], 0))
            probe = ttk.Button(self, text="あ", style=style_name)
            self.update_idletasks()
            base_h = probe.winfo_reqheight()
            probe.destroy()
            extra = max(0, spec["height"] - base_h)
            top, bottom = extra // 2, extra - extra // 2
            style.configure(style_name, padding=(spec["hpad"], top, spec["hpad"], bottom))
        style.configure(BTN_TERTIARY, foreground="#888888")

    def _build_ui(self):
        about = ttk.LabelFrame(self, text="このツールについて", padding=10)
        about.pack(fill="x", padx=8, pady=(8, 0))
        ttk.Label(about, text=ABOUT_TEXT, justify="left").pack(anchor="w")

        paned = ttk.PanedWindow(self, orient="horizontal")
        paned.pack(fill="both", expand=True, padx=8, pady=8)

        # ---- 左ペイン
        left = ttk.Frame(paned, padding=(0, 0, 8, 0))
        paned.add(left, weight=0)

        ttk.Label(left, text="data/input/ フォルダの内容（.csv / .txt）", font=HEADER_FONT).pack(anchor="w")
        list_frame = ttk.Frame(left)
        list_frame.pack(fill="both", expand=True, pady=(4, 4))
        columns = ("size",)
        self.csv_tree = ttk.Treeview(list_frame, columns=columns, height=10)
        self.csv_tree.heading("#0", text="ファイル名")
        self.csv_tree.heading("size", text="サイズ")
        self.csv_tree.column("#0", width=220)
        self.csv_tree.column("size", width=80, anchor="e", stretch=False)
        self.csv_tree.pack(side="left", fill="both", expand=True)
        sb = ttk.Scrollbar(list_frame, command=self.csv_tree.yview)
        sb.pack(side="left", fill="y")
        self.csv_tree.configure(yscrollcommand=sb.set)
        self.csv_tree.bind("<Double-1>", lambda _e: self._open_selected_preview())

        list_btns = ttk.Frame(left)
        list_btns.pack(fill="x", pady=(0, 8))
        ttk.Button(list_btns, text="冒頭を確認", style=BTN_SECONDARY,
                   command=self._open_selected_preview).pack(side="left", fill="x", expand=True, padx=(0, 4))
        ttk.Button(list_btns, text="入力フォルダを開く", style=BTN_SECONDARY,
                   command=lambda: (CSV_DIR.mkdir(parents=True, exist_ok=True), _open_in_explorer(CSV_DIR))
                   ).pack(side="left", fill="x", expand=True, padx=(4, 4))
        ttk.Button(list_btns, text="↻", width=3, style=BTN_TERTIARY,
                   command=self.refresh_csv_list).pack(side="left", padx=(4, 0))

        run_frame = ttk.Frame(left)
        run_frame.pack(fill="x", pady=(4, 8))
        self.run_btn = ttk.Button(run_frame, text="▶ 変換実行", command=self._on_run,
                                  style=BTN_PRIMARY)
        self.run_btn.pack(side="left", fill="x", expand=True, padx=(0, 4))
        self.stop_btn = ttk.Button(run_frame, text="■ Stop", command=self._on_stop,
                                   state="disabled", style=BTN_SECONDARY)
        self.stop_btn.pack(side="left", fill="x", expand=True, padx=(4, 0))

        self.status_label = ttk.Label(left, text="待機中", foreground="#888888")
        self.status_label.pack(anchor="w", pady=(4, 0))

        # ---- 右ペイン
        right = ttk.Frame(paned)
        paned.add(right, weight=1)
        right.rowconfigure(0, weight=1)
        right.columnconfigure(0, weight=1)

        self.log_text = ScrolledText(right, font=LOG_FONT, state="disabled",
                                     background="#1e1e1e", foreground="#e0e0e0",
                                     insertbackground="#e0e0e0")
        self.log_text.grid(row=0, column=0, sticky="nsew")
        for tag, opts in _LOG_TAGS.items():
            self.log_text.tag_configure(tag, **opts)

        out_frame = ttk.LabelFrame(right, text="出力（view.xlsx）", padding=6)
        out_frame.grid(row=1, column=0, sticky="ew", pady=(6, 0))
        out_frame.columnconfigure(0, weight=1)

        self.summary_var = tk.StringVar(value="（まだ実行していません）")
        self.summary_label = ttk.Label(out_frame, textvariable=self.summary_var,
                                       anchor="w", justify="left", font=UI_FONT_BOLD,
                                       foreground=_SUMMARY_COLORS["idle"])
        self.summary_label.grid(row=0, column=0, sticky="ew")
        self._summary_flash_job = None
        self.detail_btn = ttk.Button(out_frame, text="詳細...", state="disabled",
                                     style=BTN_SECONDARY, command=self._open_output_detail)
        self.detail_btn.grid(row=0, column=1, padx=(6, 0))
        self.open_btn = ttk.Button(out_frame, text="開く", state="disabled",
                                   style=BTN_SECONDARY,
                                   command=lambda: _open_in_explorer(OUTPUT_PATH))
        self.open_btn.grid(row=0, column=2, padx=(6, 0))

    # -------------------------------------------------- CSV一覧

    def refresh_csv_list(self):
        self.csv_tree.delete(*self.csv_tree.get_children())
        self._browsed_csv.clear()
        if not CSV_DIR.is_dir():
            return
        files = ctv.list_target_files(CSV_DIR)
        for path in files:
            try:
                size = _format_filesize(path.stat().st_size)
            except OSError:
                size = "-"
            item = self.csv_tree.insert("", "end", text=path.name, values=(size,))
            self._browsed_csv[item] = path
        if not files:
            self.csv_tree.insert("", "end", text="(対象ファイル[.csv/.txt]がありません)", values=("",))

    def _selected_csv_path(self):
        selection = self.csv_tree.selection()
        if not selection:
            return None
        return self._browsed_csv.get(selection[0])

    def _open_selected_preview(self):
        path = self._selected_csv_path()
        if path is None:
            messagebox.showinfo("情報", "確認したいCSVを一覧から選択してください。")
            return
        InputPreviewWindow(self, path)

    # -------------------------------------------------- 実行制御

    def _on_run(self):
        if self.is_running:
            return
        CSV_DIR.mkdir(parents=True, exist_ok=True)
        cmd = [sys.executable, "-u", str(BASE_DIR / "src" / "csv_table_viewer.py")]
        prev_mtime = OUTPUT_PATH.stat().st_mtime if OUTPUT_PATH.exists() else None

        self._reset_run_summary()
        self._set_running(True)
        self._append_log(f"=== 変換実行開始 {datetime.now():%Y-%m-%d %H:%M:%S} ===\n")

        threading.Thread(target=self._run_process, args=(cmd, prev_mtime), daemon=True).start()

    def _run_process(self, cmd, prev_mtime):
        """バックグラウンドスレッド。tkinter ウィジェットは直接触らず log_queue 経由で渡す。"""
        exit_code = None
        try:
            env = os.environ.copy()
            env["PYTHONIOENCODING"] = "utf-8"
            env["PYTHONUTF8"] = "1"
            proc = subprocess.Popen(
                cmd, cwd=str(BASE_DIR),
                stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                text=True, encoding="utf-8", errors="replace", bufsize=1,
                env=env, creationflags=subprocess.CREATE_NO_WINDOW)
            self.proc = proc
            for line in proc.stdout:
                self.log_queue.put(line)
            exit_code = proc.wait()
        except Exception as exc:
            self.log_queue.put(f"[launcher] 実行エラー: {exc}\n")
        finally:
            self.proc = None
            self.log_queue.put((_SENTINEL_RUN_DONE, (exit_code, prev_mtime)))

    def _on_run_finished(self, exit_code, prev_mtime):
        if exit_code == 0:
            self._append_log("=== 終了 (exit code: 0) ===\n", tag="pass")
        elif exit_code is None:
            self._append_log("=== 異常終了（プロセスを起動できませんでした） ===\n", tag="fail")
        else:
            self._append_log(f"=== 終了 (exit code: {exit_code}) ===\n", tag="fail")
        self._set_running(False)
        self._autosave_log()

        new_mtime = OUTPUT_PATH.stat().st_mtime if OUTPUT_PATH.exists() else None
        if new_mtime is None:
            self._show_run_summary(None, updated=False)
        elif prev_mtime is None or new_mtime > prev_mtime:
            self._show_run_summary(read_output_summary(OUTPUT_PATH), updated=True)
        else:
            self._show_run_summary(self._last_sheet_info, updated=False)

    def _on_stop(self):
        proc = self.proc
        if proc is not None:
            self._append_log("[launcher] Stop 要求を送信しました\n", tag="warn")
            try:
                proc.terminate()
            except OSError:
                pass

    def _set_running(self, running):
        self.is_running = running
        self.run_btn.config(state="disabled" if running else "normal")
        self.stop_btn.config(state="normal" if running else "disabled")
        self.status_label.config(
            text="実行中..." if running else "待機中",
            foreground="#4a9eff" if running else "#888888")

    def _on_close(self):
        proc = self.proc
        if proc is not None:
            if not messagebox.askyesno("確認", "実行中のプロセスがあります。停止して終了しますか？"):
                return
            try:
                proc.terminate()
            except OSError:
                pass
        self.destroy()

    # -------------------------------------------------- ログ表示

    @staticmethod
    def _get_log_tag(line):
        stripped = line.lstrip()
        if stripped.startswith("==="):
            return "header"
        lower = stripped.lower()
        if stripped.startswith(("エラー", "Error", "ERROR")) or "error:" in lower:
            return "error"
        if "警告" in stripped or "warn" in lower:
            return "warn"
        if "完了" in stripped:
            return "pass"
        return None

    def _append_log(self, text, tag=None):
        """1行だけを即座に反映する（コード内から直接呼ぶ少数呼び出し用）。"""
        self._append_log_lines([(text, tag)])

    def _append_log_lines(self, entries):
        """複数行をまとめて1回のstate切替・see("end")で反映する。

        1行ごとにconfigure(state=...)やsee("end")を呼ぶと、行数が多いとText部品の
        処理コストが積み上がり、メインスレッドが長時間ブロックされる
        （32_フォルダ構造Excel出力での実測: 3万行を1行ずつ処理すると60秒以上かかり
        Windowsに「応答なし」と判定された）。entries は (text, tag_or_None) のリスト。
        tagがNoneなら内容から自動判定する。

        注意: このツールは _autosave_log() がウィジェットの表示内容をそのままログ
        ファイルへ保存する設計のため、32番と違って表示行数の間引きは行わない
        （間引くと保存されるログの前半が欠けてしまう）。
        """
        if not entries:
            return
        self.log_text.configure(state="normal")
        for text, tag in entries:
            if tag is None:
                tag = self._get_log_tag(text)
            if tag:
                self.log_text.insert("end", text, tag)
            else:
                self.log_text.insert("end", text)
        self.log_text.see("end")
        self.log_text.configure(state="disabled")

    def _flush_log_queue(self):
        """1回の呼び出しで処理する行数に上限を設ける。上限に達したら残りは次のタイマー
        （100ms後）に持ち越す。バーストが続く間はログ表示が実行に追いつくまで遅延するが、
        メインスレッドが長時間ブロックされて「応答なし」になる事態を防げる。"""
        pending = []
        processed = 0
        while processed < MAX_LOG_LINES_PER_FLUSH:
            try:
                item = self.log_queue.get_nowait()
            except queue.Empty:
                break
            processed += 1
            if isinstance(item, tuple) and item and item[0] is _SENTINEL_RUN_DONE:
                if pending:
                    self._append_log_lines(pending)
                    pending = []
                self._on_run_finished(*item[1])
            else:
                pending.append((item, None))
        if pending:
            self._append_log_lines(pending)

    def _drain_log_queue(self):
        self._flush_log_queue()
        self.after(100, self._drain_log_queue)

    def _autosave_log(self):
        content = self.log_text.get("1.0", "end-1c")
        if not content.strip():
            return
        try:
            LOGS_DIR.mkdir(parents=True, exist_ok=True)
            log_path = LOGS_DIR / f"run_{datetime.now():%Y%m%d_%H%M%S}.log"
            log_path.write_text(content, encoding="utf-8")
            self._append_log(f"[launcher] ログを保存しました: {log_path}\n", tag="debug")
        except OSError as exc:
            self._append_log(f"[launcher] ログ保存に失敗: {exc}\n", tag="error")

    # -------------------------------------------------- 実行結果サマリー

    def _reset_run_summary(self):
        self._cancel_summary_flash()
        self.summary_var.set("実行中...")
        self.summary_label.config(foreground=_SUMMARY_COLORS["running"])
        self.detail_btn.config(state="disabled")
        self.open_btn.config(state="disabled")

    def _load_existing_summary(self):
        """起動時、既存の view.xlsx があればその内容をサマリーに表示する。"""
        info = read_output_summary(OUTPUT_PATH)
        if info is not None:
            self._show_run_summary(info, updated=False, prefix="（前回の出力）")

    def _show_run_summary(self, sheet_info, updated, prefix=""):
        self._last_sheet_info = sheet_info
        if sheet_info is None:
            self.summary_var.set("出力ファイルがありません")
            self.detail_btn.config(state="disabled")
            self.open_btn.config(state="disabled")
            self._flash_summary(_SUMMARY_COLORS["empty"])
            return
        total_rows = sum(rows for _name, rows in sheet_info)
        line = f"{prefix}シート {len(sheet_info)} 件 / 合計 {total_rows:,} 行"
        if updated:
            self.summary_var.set(f"✔ {line}")
            self._flash_summary(_SUMMARY_COLORS["done"])
            self._append_log(f"[launcher] view.xlsx を更新しました（{line}）\n", tag="debug")
        else:
            self.summary_var.set(f"{line}\n（変化なし。実行に失敗した可能性があります）" if not prefix else line)
            self._flash_summary(_SUMMARY_COLORS["idle"] if prefix else _SUMMARY_COLORS["empty"])
        self.detail_btn.config(state="normal")
        self.open_btn.config(state="normal")

    def _flash_summary(self, color, blinks=3, interval=180):
        self._cancel_summary_flash()
        total_steps = blinks * 2 - 1

        def step(n):
            self._summary_flash_job = None
            on = (n % 2 == 0)
            self.summary_label.config(foreground=color if on else _SUMMARY_COLORS["idle"])
            if n + 1 < total_steps:
                self._summary_flash_job = self.after(interval, step, n + 1)

        step(0)

    def _cancel_summary_flash(self):
        if self._summary_flash_job is not None:
            self.after_cancel(self._summary_flash_job)
            self._summary_flash_job = None

    def _open_output_detail(self):
        if self._last_sheet_info:
            OutputDetailWindow(self, self._last_sheet_info, OUTPUT_PATH)


def main():
    app = LauncherApp()
    app.mainloop()


if __name__ == "__main__":
    main()
