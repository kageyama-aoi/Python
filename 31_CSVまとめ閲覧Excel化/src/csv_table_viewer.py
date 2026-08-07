"""
CSV Table Viewer
================

data/input/ フォルダ内の全CSV/TXTを読み込み、見やすく整形されたExcelファイル（data/output/view.xlsx）を生成する。

準備するもの
------------
data/input/ フォルダに UTF-8 または Shift-JIS の .csv / .txt ファイルを置くこと（1ファイル = 1シートになる）。
区切り文字（カンマ／タブ）はファイルごとに自動判定する。

出力
----
data/output/view.xlsx （INDEXシート + ファイルごとのシート）

詳しい画面構成・トラブル時の見方は launcher_gui.py（GUIランチャー）を参照。
"""

import csv
import io
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment

# src/ の1つ上（プロジェクトルート）を基準にする。cwdに依存しないため、
# run.batから起動しても直接 `python src/csv_table_viewer.py` を実行しても同じ場所を指す。
BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / "data"
CSV_DIR = DATA_DIR / "input"
OUTPUT_PATH = DATA_DIR / "output" / "view.xlsx"

# 対象とするファイル拡張子（大文字小文字は区別しない）
_TARGET_EXTENSIONS = (".csv", ".txt")

# 上位から順に試すエンコード。utf-8-sig は BOM付きUTF-8も無印UTF-8も両方読める。
_ENCODING_CANDIDATES = [
    ("utf-8-sig", "UTF-8"),
    ("cp932", "Shift-JIS"),
]

# 区切り文字の判定候補（この順で優先度を持たせるわけではなく、Snifferに両方を候補として渡す）
_DELIMITER_CANDIDATES = ",\t"


class CsvReadError(RuntimeError):
    """CSV/TXT読み込み時の、原因を人間向けに説明するエラー。"""


def _decode_with_fallback(path: Path) -> str:
    """UTF-8 → Shift-JIS の順でファイル全体のデコードを試みる。全滅したら原因を含めて例外を投げる。"""
    raw = path.read_bytes()
    errors = []
    for encoding, label in _ENCODING_CANDIDATES:
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError as exc:
            errors.append(f"{label}: {exc}")
    detail = "\n".join(f"  - {e}" for e in errors)
    raise CsvReadError(
        f"'{path.name}' の文字コードを判別できませんでした（UTF-8 / Shift-JIS を試行）。\n"
        f"{detail}\n"
        f"→ ファイルをUTF-8で保存し直してから再実行してください。"
    )


def detect_delimiter(text: str) -> str:
    """先頭部分からカンマ／タブ区切りを判定する。判定できなければタブの出現数で決める。"""
    sample = text[:8192]
    try:
        return csv.Sniffer().sniff(sample, delimiters=_DELIMITER_CANDIDATES).delimiter
    except csv.Error:
        first_line = sample.splitlines()[0] if sample else ""
        return "\t" if first_line.count("\t") > first_line.count(",") else ","


def _read_table_with_fallback(path: Path) -> pd.DataFrame:
    """エンコードと区切り文字（カンマ/タブ）を自動判定して読み込む。"""
    text = _decode_with_fallback(path)
    delimiter = detect_delimiter(text)
    return pd.read_csv(io.StringIO(text), dtype=str, sep=delimiter)


def _unique_sheet_name(name: str, used: set) -> str:
    """Excelのシート名制約（31文字・重複不可）を満たす名前を作る。"""
    base = name[:31]
    candidate = base
    suffix = 2
    while candidate in used:
        tail = f"_{suffix}"
        candidate = base[: 31 - len(tail)] + tail
        suffix += 1
    used.add(candidate)
    return candidate


def list_target_files(csv_dir: Path = CSV_DIR) -> list[Path]:
    """csv_dir 直下の対象ファイル（.csv / .txt、大文字小文字不問）を名前順で返す。"""
    if not csv_dir.is_dir():
        return []
    return sorted(p for p in csv_dir.iterdir()
                  if p.is_file() and p.suffix.lower() in _TARGET_EXTENSIONS)


def build_view_excel(csv_dir: Path = CSV_DIR, output_path: Path = OUTPUT_PATH) -> list[tuple[str, int]]:
    """csv_dir 内の全CSV/TXTをExcelに変換する。作成した (シート名, 行数) のリストを返す。"""
    files = list_target_files(csv_dir)
    if not files:
        raise CsvReadError(
            f"'{csv_dir}/' に対象ファイル（.csv / .txt）が見つかりませんでした。\n"
            f"→ 変換したいファイルを '{csv_dir}/' フォルダに置いてから再実行してください。"
        )

    print(f"対象ファイル数: {len(files)}")

    # フェーズ1: 全ファイルを読み込む（出力ファイルには一切触れない）。
    # ここで検証を済ませておくことで、書き出し開始後に読み込みエラーが起きて
    # view.xlsxのファイルハンドルが中途半端な状態で残る事態を避ける。
    used_names: set = set()
    loaded: list[tuple[str, "pd.DataFrame"]] = []
    for i, file in enumerate(files, start=1):
        name = _unique_sheet_name(file.stem, used_names)
        print(f"[{i}/{len(files)}] 読み込み中: {file.name}")
        df = _read_table_with_fallback(file)
        print(f"  行数: {len(df)}")
        loaded.append((name, df))

    # フェーズ2: Excelへ書き出す
    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        writer = pd.ExcelWriter(output_path, engine="openpyxl")
    except PermissionError:
        raise CsvReadError(
            f"'{output_path}' に書き込めませんでした。\n"
            f"→ Excelで開いている場合は閉じてから再実行してください。"
        )

    sheet_info: list[tuple[str, int]] = []
    try:
        for name, df in loaded:
            df.to_excel(writer, sheet_name=name, index=False)
            ws = writer.book[name]
            sheet_info.append((name, len(df)))

            # ヘッダ装飾
            fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            font = Font(color="FFFFFF", bold=True)
            for cell in ws[1]:
                cell.fill = fill
                cell.font = font

            ws.auto_filter.ref = ws.dimensions
            ws.freeze_panes = "A2"

            # 列幅調整
            for column in ws.columns:
                column_letter = column[0].column_letter
                max_length = 0
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                if max_length > 50:
                    ws.column_dimensions[column_letter].width = 60
                    for cell in column:
                        cell.alignment = Alignment(wrap_text=True)
                else:
                    ws.column_dimensions[column_letter].width = min(max_length + 2, 40)

        # INDEXシート作成
        index_ws = writer.book.create_sheet("INDEX", 0)
        index_ws["A1"] = "CSVビューア INDEX"
        index_ws["A1"].font = Font(size=16, bold=True)
        index_ws["A3"] = "作成日時"
        index_ws["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        index_ws["A5"] = "シート名"
        index_ws["B5"] = "行数"

        row = 6
        for name, rows in sheet_info:
            cell = index_ws.cell(row=row, column=1)
            cell.value = name
            cell.hyperlink = f"#{name}!A1"
            cell.style = "Hyperlink"
            index_ws.cell(row=row, column=2).value = rows
            row += 1
    except BaseException:
        # 例外時もファイルハンドルは必ず解放する（1件も書き込めていない状態でのclose()は
        # openpyxl側で別の例外を出すことがあるため、close失敗は無視して元の例外を優先する）。
        # 中途半端な出力を残さないよう、保存されていれば削除する。
        try:
            writer.close()
        except Exception:
            pass
        output_path.unlink(missing_ok=True)
        raise

    writer.close()
    print(f"\nExcel生成完了: {output_path}")
    return sheet_info


def main() -> int:
    try:
        build_view_excel()
    except CsvReadError as exc:
        print(f"\nエラー: {exc}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
