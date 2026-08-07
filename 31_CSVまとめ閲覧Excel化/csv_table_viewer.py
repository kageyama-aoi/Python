"""
CSV Table Viewer
================

csv/ フォルダ内の全CSVを読み込み、見やすく整形されたExcelファイル（view.xlsx）を生成する。

準備するもの
------------
csv/ フォルダに UTF-8 のCSVファイルを置くこと（1CSV = 1シートになる）。

出力
----
view.xlsx （INDEXシート + CSVごとのシート）

詳しい画面構成・トラブル時の見方は launcher_gui.py（GUIランチャー）を参照。
"""

import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment

CSV_DIR = Path("csv")
OUTPUT_PATH = Path("view.xlsx")

# 上位から順に試すエンコード。utf-8-sig は BOM付きUTF-8も無印UTF-8も両方読める。
_ENCODING_CANDIDATES = [
    ("utf-8-sig", "UTF-8"),
    ("cp932", "Shift-JIS"),
]


class CsvReadError(RuntimeError):
    """CSV読み込み時の、原因を人間向けに説明するエラー。"""


def _read_csv_with_fallback(path: Path) -> pd.DataFrame:
    """UTF-8 → Shift-JIS の順で読み込みを試みる。全滅したら原因を含めて例外を投げる。"""
    errors = []
    for encoding, label in _ENCODING_CANDIDATES:
        try:
            return pd.read_csv(path, dtype=str, encoding=encoding)
        except UnicodeDecodeError as exc:
            errors.append(f"{label}: {exc}")
    detail = "\n".join(f"  - {e}" for e in errors)
    raise CsvReadError(
        f"'{path.name}' の文字コードを判別できませんでした（UTF-8 / Shift-JIS を試行）。\n"
        f"{detail}\n"
        f"→ CSVをUTF-8で保存し直してから再実行してください。"
    )


def _unique_sheet_name(name: str, used: set) -> str:
    """Excelのシート名制約（31文字・重複不可）を満たす名前を作る。"""
    base = name[:31]
    candidate = base
    suffix = 2
    while candidate in used:
        tail = f"_{suffix}"
        candidate = base[: 31 - len(tail)] + tail
        suffix += 1
    used.add(candidate)
    return candidate


def build_view_excel(csv_dir: Path = CSV_DIR, output_path: Path = OUTPUT_PATH) -> list[tuple[str, int]]:
    """csv_dir 内の全CSVをExcelに変換する。作成した (シート名, 行数) のリストを返す。"""
    files = sorted(csv_dir.glob("*.csv"))
    if not files:
        raise CsvReadError(
            f"'{csv_dir}/' にCSVファイルが見つかりませんでした。\n"
            f"→ 変換したいCSVを '{csv_dir}/' フォルダに置いてから再実行してください。"
        )

    print(f"CSVファイル数: {len(files)}")

    # フェーズ1: 全CSVを読み込む（出力ファイルには一切触れない）。
    # ここで検証を済ませておくことで、書き出し開始後に読み込みエラーが起きて
    # view.xlsxのファイルハンドルが中途半端な状態で残る事態を避ける。
    used_names: set = set()
    loaded: list[tuple[str, "pd.DataFrame"]] = []
    for i, file in enumerate(files, start=1):
        name = _unique_sheet_name(file.stem, used_names)
        print(f"[{i}/{len(files)}] 読み込み中: {file.name}")
        df = _read_csv_with_fallback(file)
        print(f"  行数: {len(df)}")
        loaded.append((name, df))

    # フェーズ2: Excelへ書き出す
    try:
        writer = pd.ExcelWriter(output_path, engine="openpyxl")
    except PermissionError:
        raise CsvReadError(
            f"'{output_path}' に書き込めませんでした。\n"
            f"→ Excelで開いている場合は閉じてから再実行してください。"
        )

    sheet_info: list[tuple[str, int]] = []
    try:
        for name, df in loaded:
            df.to_excel(writer, sheet_name=name, index=False)
            ws = writer.book[name]
            sheet_info.append((name, len(df)))

            # ヘッダ装飾
            fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            font = Font(color="FFFFFF", bold=True)
            for cell in ws[1]:
                cell.fill = fill
                cell.font = font

            ws.auto_filter.ref = ws.dimensions
            ws.freeze_panes = "A2"

            # 列幅調整
            for column in ws.columns:
                column_letter = column[0].column_letter
                max_length = 0
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                if max_length > 50:
                    ws.column_dimensions[column_letter].width = 60
                    for cell in column:
                        cell.alignment = Alignment(wrap_text=True)
                else:
                    ws.column_dimensions[column_letter].width = min(max_length + 2, 40)

        # INDEXシート作成
        index_ws = writer.book.create_sheet("INDEX", 0)
        index_ws["A1"] = "CSVビューア INDEX"
        index_ws["A1"].font = Font(size=16, bold=True)
        index_ws["A3"] = "作成日時"
        index_ws["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        index_ws["A5"] = "シート名"
        index_ws["B5"] = "行数"

        row = 6
        for name, rows in sheet_info:
            cell = index_ws.cell(row=row, column=1)
            cell.value = name
            cell.hyperlink = f"#{name}!A1"
            cell.style = "Hyperlink"
            index_ws.cell(row=row, column=2).value = rows
            row += 1
    except BaseException:
        # 例外時もファイルハンドルは必ず解放する（1件も書き込めていない状態でのclose()は
        # openpyxl側で別の例外を出すことがあるため、close失敗は無視して元の例外を優先する）。
        # 中途半端な出力を残さないよう、保存されていれば削除する。
        try:
            writer.close()
        except Exception:
            pass
        output_path.unlink(missing_ok=True)
        raise

    writer.close()
    print(f"\nExcel生成完了: {output_path}")
    return sheet_info


def main() -> int:
    try:
        build_view_excel()
    except CsvReadError as exc:
        print(f"\nエラー: {exc}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
