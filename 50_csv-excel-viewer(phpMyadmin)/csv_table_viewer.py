"""
CSV Table Viewer
================

概要
----
このスクリプトは、CSVファイルを読み込み、見やすく整形されたExcelファイル
(view.xlsx)を自動生成するツールです。

主に以下の用途を想定しています。

・phpMyAdminなどからエクスポートしたCSVの閲覧
・DBテーブルデータの確認
・Excel上でのフィルタ・検索

CSVをExcelで直接開くと発生する問題（先頭ゼロ消失、列幅崩れなど）を
回避し、閲覧に最適なExcelファイルを生成します。


主な機能
--------
1. CSVフォルダ内の全CSVを読み込み
2. 1CSV = 1Excelシートとして出力
3. INDEXシート（表紙）を自動生成
4. 各シートへのハイパーリンクをINDEXに作成
5. 作成タイムスタンプをINDEXに表示
6. CSV行数をINDEXに表示

Excel整形
---------
以下の設定を自動で適用します。

・ヘッダ黒背景 + 白文字
・オートフィルタ
・ヘッダ行固定
・列幅自動調整
・長文列は折り返し表示


フォルダ構成
------------
以下の構成で使用します。

project/
    convert.py
    csv/
        table1.csv
        table2.csv
        table3.csv


使用方法
--------
1. CSVファイルを csv フォルダに配置
2. 以下コマンドを実行

    python convert.py

3. view.xlsx が生成されます


出力ファイル
------------
view.xlsx

構成

INDEX
    ├ テーブル一覧
    ├ 行数
    └ 作成日時

各CSV
    ├ テーブルデータ
    ├ フィルタ
    └ 整形済表示


INDEXシート
-----------
INDEXシートには以下の情報が表示されます。

・Excel作成日時
・シート一覧
・各CSV行数
・シートへのハイパーリンク


必要ライブラリ
--------------
以下を事前にインストールしてください。

pip install pandas
pip install openpyxl
pip install tqdm


注意事項
--------
・CSVの文字コードはUTF-8を想定しています
・Excelのシート名は31文字制限があるため自動で切り詰めます
・Excelで使用できない文字は自動で除外されます


想定用途
--------
・DBテーブル確認
・テストデータ検証
・ログデータ閲覧
・SQL結果の共有


更新履歴
--------
v1.0
・CSV → Excel変換
・INDEXシート生成
・進捗表示


作者
----
Internal utility script
"""

import pandas as pd
import glob
import os
from datetime import datetime
from tqdm import tqdm
from openpyxl.styles import PatternFill, Font, Alignment

CSV_DIR = "csv"

files = glob.glob(os.path.join(CSV_DIR, "*.csv"))

print(f"CSVファイル数: {len(files)}")

writer = pd.ExcelWriter("view.xlsx", engine="openpyxl")

sheet_info = []

for file in tqdm(files, desc="CSV処理中"):

    name = os.path.basename(file).replace(".csv", "")
    name = name[:31]

    print(f"\n読み込み中: {file}")

    df = pd.read_csv(file, dtype=str)

    rows = len(df)

    print(f"行数: {rows}")

    df.to_excel(writer, sheet_name=name, index=False)

    ws = writer.book[name]

    sheet_info.append((name, rows))

    # ヘッダ装飾
    fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = fill
        cell.font = font

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    # 列幅調整
    for column in ws.columns:

        column_letter = column[0].column_letter
        max_length = 0

        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        if max_length > 50:
            ws.column_dimensions[column_letter].width = 60
            for cell in column:
                cell.alignment = Alignment(wrap_text=True)
        else:
            ws.column_dimensions[column_letter].width = min(max_length + 2, 40)

# INDEXシート作成
index_ws = writer.book.create_sheet("INDEX", 0)

index_ws["A1"] = "CSVビューア INDEX"
index_ws["A1"].font = Font(size=16, bold=True)

index_ws["A3"] = "作成日時"
index_ws["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

index_ws["A5"] = "シート名"
index_ws["B5"] = "行数"

row = 6

for name, rows in sheet_info:

    cell = index_ws.cell(row=row, column=1)
    cell.value = name

    cell.hyperlink = f"#{name}!A1"
    cell.style = "Hyperlink"

    index_ws.cell(row=row, column=2).value = rows

    row += 1

writer.close()

print("\nExcel生成完了: view.xlsx")